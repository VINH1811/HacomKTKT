from __future__ import annotations

import posixpath
import re
import time
import zipfile
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable, Optional
from defusedxml import ElementTree as ET

from openpyxl import load_workbook

try:
    from python_calamine import CalamineWorkbook
except Exception:  # optional fallback is handled explicitly
    CalamineWorkbook = None  # type: ignore[assignment]

_FORMULA_ERROR = re.compile(r"#(?:REF!|DIV/0!|VALUE!|NAME\?|N/A|NUM!|NULL!)", re.I)
_CELL_REF = re.compile(r"([A-Z]{1,3})(\d+)", re.I)
_MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_PKG_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"


@dataclass(slots=True)
class SheetMatrix:
    name: str
    rows: list[list[Any]]
    row_count: int
    col_count: int


@dataclass(slots=True)
class WorkbookMatrices:
    path: Path
    sheets: list[SheetMatrix]
    engine: str
    elapsed_seconds: float
    warnings: list[str] = field(default_factory=list)


@dataclass(slots=True)
class SpreadsheetIssue:
    sheet: str
    row: int
    column: int
    cell: str
    kind: str
    formula: str = ""
    value: str = ""
    message: str = ""
    severity: str = "CRITICAL"

    def to_dict(self) -> dict[str, Any]:
        return {
            "sheet": self.sheet,
            "row": self.row,
            "column": self.column,
            "cell": self.cell,
            "kind": self.kind,
            "formula": self.formula,
            "value": self.value,
            "message": self.message,
            "severity": self.severity,
        }


@dataclass(slots=True)
class SpreadsheetScan:
    issues: list[SpreadsheetIssue] = field(default_factory=list)
    external_link_count: int = 0
    warnings: list[str] = field(default_factory=list)

    @property
    def issues_by_row(self) -> dict[tuple[str, int], list[str]]:
        result: dict[tuple[str, int], list[str]] = defaultdict(list)
        for issue in self.issues:
            result[(issue.sheet, issue.row)].append(issue.message)
        return dict(result)


def _trim_row(row: Iterable[Any]) -> list[Any]:
    values = list(row)
    while values and values[-1] in (None, ""):
        values.pop()
    return values


def _read_calamine(
    path: Path,
    selected_sheets: Optional[list[str]],
    max_rows: int,
) -> list[SheetMatrix]:
    if CalamineWorkbook is None:
        raise RuntimeError("Chưa cài python-calamine")
    workbook = CalamineWorkbook.from_path(str(path))
    selected = set(selected_sheets or [])
    sheets: list[SheetMatrix] = []
    try:
        for sheet_name in workbook.sheet_names:
            if selected_sheets is not None and sheet_name not in selected:
                continue
            sheet = workbook.get_sheet_by_name(sheet_name)
            raw_rows = sheet.to_python(skip_empty_area=False, nrows=max_rows + 100)
            rows = [_trim_row(row) for row in raw_rows]
            col_count = max((len(row) for row in rows), default=0)
            sheets.append(SheetMatrix(sheet_name, rows, len(rows), col_count))
    finally:
        close = getattr(workbook, "close", None)
        if callable(close):
            close()
    return sheets


def _read_openpyxl(
    path: Path,
    selected_sheets: Optional[list[str]],
    max_rows: int,
) -> list[SheetMatrix]:
    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    selected = set(selected_sheets or [])
    sheets: list[SheetMatrix] = []
    try:
        for ws in workbook.worksheets:
            if selected_sheets is not None and ws.title not in selected:
                continue
            max_row = min(int(ws.max_row or 0), max_rows + 100)
            rows = [_trim_row(row) for row in ws.iter_rows(min_row=1, max_row=max_row, values_only=True)]
            col_count = max((len(row) for row in rows), default=0)
            sheets.append(SheetMatrix(ws.title, rows, len(rows), col_count))
    finally:
        workbook.close()
    return sheets


def read_workbook_matrices(
    path: str | Path,
    *,
    engine: str = "calamine",
    selected_sheets: Optional[list[str]] = None,
    max_rows: int = 1_000_000,
    fallback_openpyxl: bool = True,
) -> WorkbookMatrices:
    """Read cell values quickly while preserving original row coordinates.

    Calamine is the preferred read-only engine. openpyxl is retained as a
    compatibility fallback and remains the writer/annotator elsewhere.
    """
    started = time.perf_counter()
    path = Path(path)
    engine = (engine or "calamine").strip().lower()
    if engine not in {"calamine", "openpyxl", "auto"}:
        engine = "calamine"
    warnings: list[str] = []

    if engine in {"calamine", "auto"}:
        try:
            sheets = _read_calamine(path, selected_sheets, max_rows)
            return WorkbookMatrices(path, sheets, "calamine", time.perf_counter() - started, warnings)
        except Exception as exc:
            if not fallback_openpyxl and engine == "calamine":
                raise
            warnings.append(f"Calamine không đọc được, chuyển sang openpyxl: {type(exc).__name__}: {exc}")

    sheets = _read_openpyxl(path, selected_sheets, max_rows)
    return WorkbookMatrices(path, sheets, "openpyxl", time.perf_counter() - started, warnings)


def list_sheets_fast(path: str | Path, *, engine: str = "calamine") -> list[dict[str, Any]]:
    matrices = read_workbook_matrices(path, engine=engine, max_rows=1_000_000)
    return [{"name": sheet.name, "rows": sheet.row_count, "cols": sheet.col_count} for sheet in matrices.sheets]


def _normalise_target(target: str) -> str:
    target = target.replace("\\", "/")
    if target.startswith("/"):
        return target.lstrip("/")
    return posixpath.normpath(posixpath.join("xl", target))


def _sheet_xml_map(archive: zipfile.ZipFile) -> dict[str, str]:
    workbook_xml = ET.fromstring(archive.read("xl/workbook.xml"))
    rels_xml = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    rel_map: dict[str, str] = {}
    for rel in rels_xml.findall(f"{{{_PKG_REL_NS}}}Relationship"):
        rel_map[str(rel.attrib.get("Id", ""))] = _normalise_target(str(rel.attrib.get("Target", "")))

    result: dict[str, str] = {}
    sheets_node = workbook_xml.find(f"{{{_MAIN_NS}}}sheets")
    if sheets_node is None:
        return result
    for sheet in sheets_node.findall(f"{{{_MAIN_NS}}}sheet"):
        name = str(sheet.attrib.get("name", ""))
        rel_id = str(sheet.attrib.get(f"{{{_REL_NS}}}id", ""))
        target = rel_map.get(rel_id)
        if name and target:
            result[name] = target
    return result


def _column_number(cell_ref: str) -> int:
    match = _CELL_REF.fullmatch(cell_ref.upper())
    if not match:
        return 0
    value = 0
    for char in match.group(1):
        value = value * 26 + (ord(char) - 64)
    return value


def scan_xlsx_issues(
    path: str | Path,
    *,
    selected_sheets: Optional[list[str]] = None,
    scan_formulas: bool = True,
    scan_external_links: bool = True,
) -> SpreadsheetScan:
    """Scan XLSX OOXML directly for broken formulas and external references.

    This avoids loading the workbook a second time with openpyxl and catches
    #REF! even when the cached cell value is empty.
    """
    result = SpreadsheetScan()
    path = Path(path)
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        return result

    selected = set(selected_sheets or [])
    try:
        with zipfile.ZipFile(path) as archive:
            names = set(archive.namelist())
            sheet_map = _sheet_xml_map(archive)
            if scan_external_links:
                result.external_link_count = sum(
                    name.startswith("xl/externalLinks/externalLink") and name.endswith(".xml")
                    for name in names
                )
            for sheet_name, xml_path in sheet_map.items():
                if selected_sheets is not None and sheet_name not in selected:
                    continue
                if xml_path not in names:
                    continue
                root = ET.fromstring(archive.read(xml_path))
                for cell in root.iter(f"{{{_MAIN_NS}}}c"):
                    cell_ref = str(cell.attrib.get("r", ""))
                    match = _CELL_REF.fullmatch(cell_ref.upper())
                    if not match:
                        continue
                    row_number = int(match.group(2))
                    col_number = _column_number(cell_ref)
                    formula_node = cell.find(f"{{{_MAIN_NS}}}f")
                    value_node = cell.find(f"{{{_MAIN_NS}}}v")
                    formula = "" if formula_node is None or formula_node.text is None else formula_node.text.strip()
                    value = "" if value_node is None or value_node.text is None else value_node.text.strip()
                    combined = f"{formula} {value}".strip()

                    if scan_formulas:
                        error_match = _FORMULA_ERROR.search(combined)
                        if error_match:
                            token = error_match.group(0).upper()
                            result.issues.append(SpreadsheetIssue(
                                sheet=sheet_name,
                                row=row_number,
                                column=col_number,
                                cell=cell_ref,
                                kind="FORMULA_ERROR",
                                formula=formula,
                                value=value,
                                message=(
                                    f"Lỗi công thức {token} tại ô {cell_ref}; không sử dụng giá trị ô này để tính toán, "
                                    "cần sửa lại tham chiếu trong file nguồn."
                                ),
                            ))

                    if scan_external_links and formula and "[" in formula and "]" in formula:
                        result.issues.append(SpreadsheetIssue(
                            sheet=sheet_name,
                            row=row_number,
                            column=col_number,
                            cell=cell_ref,
                            kind="EXTERNAL_LINK",
                            formula=formula,
                            value=value,
                            message=f"Công thức tại ô {cell_ref} tham chiếu workbook bên ngoài; cần xác nhận file nguồn liên kết.",
                            severity="REVIEW",
                        ))
    except (zipfile.BadZipFile, KeyError, ET.ParseError, OSError) as exc:
        result.warnings.append(f"Không thể quét công thức OOXML: {type(exc).__name__}: {exc}")
    return result
