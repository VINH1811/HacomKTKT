"""RFI Tracker — theo dõi vòng LÀM RÕ hồ sơ chào giá (HSCG).

Chủ đầu tư gửi mỗi nhà thầu một file "Nội dung làm rõ HSCG" (sheet PL 1 đánh giá
hợp lệ/năng lực, PL 2 kỹ thuật) trong đó cột "Ý kiến CĐT" là các YÊU CẦU làm rõ.
Nhà thầu phản hồi bằng chính file đó với cột "Nhà thầu trả lời làm rõ" được điền.

Tracker ghép từng yêu cầu trong file CĐT với dòng tương ứng trong file phản hồi
(ghép theo nội dung, chịu được việc nhà thầu sửa/chèn dòng) và chấm trạng thái:
ĐÃ TRẢ LỜI / CHƯA TRẢ LỜI / KHÔNG THẤY TRONG PHẢN HỒI. Xuất Excel tổng hợp.
"""

from __future__ import annotations

import unicodedata
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import xlsxwriter
from openpyxl import load_workbook
from rapidfuzz import fuzz

STATUS_ANSWERED = "ĐÃ TRẢ LỜI"
STATUS_UNANSWERED = "CHƯA TRẢ LỜI"
STATUS_NOT_FOUND = "KHÔNG THẤY TRONG PHẢN HỒI"

# Từ khóa nhận diện cột trong header (đã bỏ dấu, chữ thường).
_COL_KEYS = {
    "stt": ("stt",),
    "content": ("noi dung danh gia", "noi dung"),
    "requirement": ("yeu cau",),
    "cdt_request": ("y kien cdt", "y kien chu dau tu"),
    "response": ("tra loi lam ro", "nha thau tra loi", "phan hoi"),
}


def _fold(text: str) -> str:
    text = str(text or "").replace("Đ", "D").replace("đ", "d")
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return " ".join(text.lower().split())


@dataclass
class RfiItem:
    sheet: str
    row_number: int
    stt: str
    content: str        # nội dung đánh giá theo HSYC
    requirement: str    # yêu cầu của HSYC
    cdt_request: str    # ý kiến CĐT = yêu cầu làm rõ
    response: str       # nhà thầu trả lời làm rõ

    @property
    def match_key(self) -> str:
        return _fold(f"{self.stt} {self.content} {self.cdt_request}")


@dataclass
class TrackedRfi:
    request: RfiItem
    response: Optional[RfiItem]

    @property
    def status(self) -> str:
        if self.response is None:
            return STATUS_NOT_FOUND
        return STATUS_ANSWERED if self.response.response.strip() else STATUS_UNANSWERED


@dataclass
class RfiTrackResult:
    bidder: str
    request_path: str
    response_path: str
    items: list[TrackedRfi] = field(default_factory=list)

    def count(self, status: str) -> int:
        return sum(1 for it in self.items if it.status == status)


def _find_header(rows: list[tuple], max_scan: int = 12) -> tuple[Optional[int], dict[str, int]]:
    """Tìm dòng header và ánh xạ cột theo từ khóa; trả về (chỉ số dòng, {khóa: cột})."""
    for idx, row in enumerate(rows[:max_scan]):
        folded = [_fold(c) for c in row]
        mapping: dict[str, int] = {}
        for key, keywords in _COL_KEYS.items():
            for col, cell in enumerate(folded):
                if cell and any(kw in cell for kw in keywords):
                    mapping[key] = col
                    break
        if "cdt_request" in mapping and "content" in mapping:
            return idx, mapping
    return None, {}


def parse_rfi_file(path: str | Path) -> list[RfiItem]:
    """Đọc mọi sheet dạng 'Nội dung làm rõ' trong file; mỗi dòng có Ý kiến CĐT là một yêu cầu."""
    path = Path(path)
    wb = load_workbook(path, read_only=True, data_only=True)
    items: list[RfiItem] = []
    try:
        for ws in wb.worksheets:
            if not hasattr(ws, "iter_rows"):
                continue
            rows = [tuple(row) for row in ws.iter_rows(values_only=True)]
            header_idx, cols = _find_header(rows)
            if header_idx is None:
                continue
            get = lambda row, key: str(row[cols[key]] or "").strip() if key in cols and cols[key] < len(row) else ""
            last_content = ""
            for offset, row in enumerate(rows[header_idx + 1:], header_idx + 2):
                content = get(row, "content")
                if content:
                    last_content = content
                request = get(row, "cdt_request")
                if not request:
                    continue
                items.append(RfiItem(
                    sheet=ws.title,
                    row_number=offset,
                    stt=get(row, "stt"),
                    content=content or last_content,
                    requirement=get(row, "requirement"),
                    cdt_request=request,
                    response=get(row, "response"),
                ))
    finally:
        wb.close()
    return items


def track_rfi(
    request_path: str | Path,
    response_path: str | Path,
    bidder: str,
    similarity_threshold: float = 0.75,
) -> RfiTrackResult:
    """Ghép từng yêu cầu làm rõ của CĐT với dòng tương ứng trong file phản hồi."""
    requests = parse_rfi_file(request_path)
    responses = parse_rfi_file(response_path)
    used: set[int] = set()
    tracked: list[TrackedRfi] = []
    for req in requests:
        best_idx, best_score = None, 0.0
        for idx, resp in enumerate(responses):
            if idx in used:
                continue
            score = fuzz.token_set_ratio(req.match_key, resp.match_key) / 100.0
            if score > best_score:
                best_idx, best_score = idx, score
        if best_idx is not None and best_score >= similarity_threshold:
            used.add(best_idx)
            tracked.append(TrackedRfi(request=req, response=responses[best_idx]))
        else:
            tracked.append(TrackedRfi(request=req, response=None))
    return RfiTrackResult(
        bidder=bidder,
        request_path=str(request_path),
        response_path=str(response_path),
        items=tracked,
    )


def export_rfi_report(results: list[RfiTrackResult], output_path: str | Path) -> str:
    """Xuất Excel: sheet Tổng quan (đếm theo nhà thầu) + sheet chi tiết từng nhà thầu."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = xlsxwriter.Workbook(str(output_path), {"nan_inf_to_errors": True})
    f_title = wb.add_format({"bold": True, "font_size": 14, "font_color": "#FFFFFF",
                             "bg_color": "#17365D", "align": "center", "valign": "vcenter"})
    f_head = wb.add_format({"bold": True, "font_color": "#FFFFFF", "bg_color": "#1F4E78",
                            "align": "center", "valign": "vcenter", "text_wrap": True, "border": 1})
    f_text = wb.add_format({"valign": "top", "text_wrap": True})
    f_ok = wb.add_format({"bold": True, "font_color": "#2E7D32", "bg_color": "#E2EFDA",
                          "align": "center", "valign": "top", "border": 1})
    f_no = wb.add_format({"bold": True, "font_color": "#9C0006", "bg_color": "#F4CCCC",
                          "align": "center", "valign": "top", "border": 1})
    f_nf = wb.add_format({"bold": True, "font_color": "#C65911", "bg_color": "#FCE4D6",
                          "align": "center", "valign": "top", "border": 1})
    status_fmt = {STATUS_ANSWERED: f_ok, STATUS_UNANSWERED: f_no, STATUS_NOT_FOUND: f_nf}

    ov = wb.add_worksheet("Tổng quan RFI")
    ov.set_column(0, 0, 26)
    ov.set_column(1, 4, 18)
    ov.merge_range(0, 0, 0, 4, "THEO DÕI LÀM RÕ HỒ SƠ CHÀO GIÁ (RFI)", f_title)
    ov.set_row(0, 26)
    heads = ["Nhà thầu", "Tổng yêu cầu", STATUS_ANSWERED, STATUS_UNANSWERED, STATUS_NOT_FOUND]
    for col, head in enumerate(heads):
        ov.write(2, col, head, f_head)
    for r, res in enumerate(results, 3):
        ov.write(r, 0, res.bidder, f_text)
        ov.write_number(r, 1, len(res.items))
        ov.write_number(r, 2, res.count(STATUS_ANSWERED))
        ov.write_number(r, 3, res.count(STATUS_UNANSWERED))
        ov.write_number(r, 4, res.count(STATUS_NOT_FOUND))

    for res in results:
        ws = wb.add_worksheet(f"RFI {res.bidder}"[:31])
        headers = ["Sheet", "STT", "Nội dung đánh giá", "Yêu cầu làm rõ của CĐT",
                   "Nhà thầu trả lời", "Trạng thái"]
        widths = [10, 8, 45, 55, 55, 22]
        for col, (head, width) in enumerate(zip(headers, widths)):
            ws.set_column(col, col, width)
            ws.write(0, col, head, f_head)
        ws.freeze_panes(1, 0)
        ws.autofilter(0, 0, 0, len(headers) - 1)
        for r, it in enumerate(res.items, 1):
            req = it.request
            ws.write(r, 0, req.sheet, f_text)
            ws.write(r, 1, req.stt, f_text)
            ws.write(r, 2, req.content[:1500], f_text)
            ws.write(r, 3, req.cdt_request[:1500], f_text)
            ws.write(r, 4, (it.response.response[:1500] if it.response else ""), f_text)
            ws.write(r, 5, it.status, status_fmt[it.status])
    wb.close()
    return str(output_path)
