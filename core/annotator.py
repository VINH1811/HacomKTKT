from __future__ import annotations

import re
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.formula import Tokenizer
from openpyxl.formula.tokenizer import Token
from openpyxl.formula.translate import Translator
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .models import ComparedItem, FieldDifference, Severity, WorkbookData
from .text_normalizer import normalize_stt

_FILL = {
    Severity.INFO: PatternFill("solid", fgColor="DDEBF7"),
    Severity.REVIEW: PatternFill("solid", fgColor="FFF2CC"),
    Severity.WARNING: PatternFill("solid", fgColor="FCE4D6"),
    Severity.CRITICAL: PatternFill("solid", fgColor="F4CCCC"),
}
_FONT = {
    Severity.INFO: "1F4E78",
    Severity.REVIEW: "7F6000",
    Severity.WARNING: "C65911",
    Severity.CRITICAL: "9C0006",
}
_THIN = Side(style="thin", color="D9E1F2")


def _sheet_meta(workbook: WorkbookData) -> dict[str, dict]:
    return {str(info.get("sheet")): info for info in workbook.sheet_info}


def _field_key(diff: FieldDifference) -> str | None:
    field = diff.field.lower()
    if "tên hạng mục" in field or field == "hạng mục":
        return "item_name"
    if "mã hiệu" in field:
        return "item_code"
    if "đơn vị" in field:
        return "unit"
    if "khối lượng mời thầu" in field:
        return "reference_quantity"
    if "khối lượng nhà thầu" in field:
        return "bid_quantity"
    if "đơn giá tổng hợp" in field:
        return "unit_price_total"
    if "thành tiền theo klmt" in field:
        return "reference_amount"
    if "thành tiền nhà thầu" in field:
        return "bid_amount"
    if "vl chính" in field:
        return "price_main"
    if "vl phụ" in field:
        return "price_aux"
    if "nc & máy" in field or "nc&m" in field:
        return "price_labor"
    if "quản lý" in field:
        return "price_management"
    if "lợi nhuận" in field:
        return "price_profit"
    if "vật tư" in field or "quy cách" in field:
        return "material"
    if "thương hiệu" in field:
        return "brand"
    if "xuất xứ" in field:
        return "origin"
    return None


def _comment_text(row: ComparedItem, max_chars: int = 6000) -> str:
    lines = [
        f"AI đánh dấu: {row.severity.value}",
        f"Điểm bất thường: {row.anomaly_score:.1f}/100",
        f"Ghép với PL01: {row.match.kind.value}, độ tin cậy {row.match.score:.1%}",
    ]
    if row.pl2_category:
        lines.append(f"Nhóm PL02: {row.pl2_category}")
        lines.append(f"Yêu cầu PL02: {row.pl2_requirement}")
        lines.append(f"Trạng thái PL02: {row.pl2_status}")
    if row.flags:
        lines.append("Lý do:")
        lines.extend(f"- {flag}" for flag in row.flags)
    text = "\n".join(lines)
    return text[:max_chars]


def _append_comment(cell, text: str) -> None:
    if cell.comment and cell.comment.text:
        text = cell.comment.text + "\n\n--- HSMT Enterprise AI ---\n" + text
    cell.comment = Comment(text, "HSMT Enterprise AI")


_SORTED_SUFFIX = " — sắp xếp"
_GOC_SUFFIX = " — gốc"


def _phatsinh_block_rows(rows: list[ComparedItem]) -> dict[str, set[int]]:
    """Các dòng ĐƯỢC ĐÁNH DẤU phát sinh ngoài (không ghép được PL01) theo sheet.

    Chỉ lấy đúng những dòng bản thân là phát sinh (``reference is None`` — tức có
    đánh dấu "Hạng mục phát sinh ngoài"). KHÔNG kéo theo cả khối: một dòng đã khớp
    PL01 (dù đứng ngay sau một hạng mục phát sinh) sẽ KHÔNG bị dời — tránh cuốn
    nhầm các dòng không đánh dấu xuống mục B.
    """
    result: dict[str, set[int]] = defaultdict(set)
    for row in rows:
        if row.candidate is not None and row.reference is None:
            result[row.candidate.sheet].add(row.candidate.row_number)
    return result


def _suffixed_sheet_name(base: str, suffix: str, used: set[str]) -> str:
    limit = 31
    name = base + suffix
    if len(name) > limit:
        name = base[: limit - len(suffix)] + suffix
    final, index = name, 2
    while final.lower() in used:
        tail = f" ({index})"
        final = name[: limit - len(tail)] + tail
        index += 1
    used.add(final.lower())
    return final


_CELL_REF = re.compile(r"^(\$?)([A-Za-z]{1,3})(\$?)(\d+)$")
_NAME_REF = re.compile(r"^[A-Za-z_\\][A-Za-z0-9_.\\]*$")


def _same_row_formula(formula: str, row: int) -> bool:
    """True nếu công thức chỉ tham chiếu ô CÙNG DÒNG, cùng sheet, dòng tương đối.

    Các công thức này (ví dụ Thành tiền = KL × Đơn giá cùng dòng) dời dòng được
    an toàn bằng cách dịch lại địa chỉ. Mọi trường hợp khác (tham chiếu dòng
    khác, dòng tuyệt đối $5, sheet khác, vùng nhiều dòng) trả False.
    """
    try:
        tokens = Tokenizer(formula).items
    except Exception:
        return False
    for tok in tokens:
        if tok.type != Token.OPERAND or tok.subtype != Token.RANGE:
            continue
        ref = tok.value
        if "!" in ref:
            return False
        for part in ref.split(":"):
            m = _CELL_REF.match(part)
            if not m:
                # Defined name (độc lập vị trí) thì cho phép; còn lại từ chối.
                if _NAME_REF.match(part):
                    continue
                return False
            if m.group(3) == "$" or int(m.group(4)) != row:
                return False
    return True


def _translate_row_formula(formula: str, col_letter: str, src_row: int, dst_row: int):
    try:
        return Translator(formula, origin=f"{col_letter}{src_row}").translate_formula(f"{col_letter}{dst_row}")
    except Exception:
        return None


def _rewrite_sheet_references(wb, renames: dict[str, str]) -> None:
    """Đổi mọi tham chiếu `'tên cũ'!` sang `'tên mới'!` trong toàn workbook.

    openpyxl không tự cập nhật tham chiếu khi đổi tên sheet (khác Excel), nên
    phải tự sửa: công thức, hyperlink nội bộ, defined names và chart series —
    để sheet `Tổng hợp`/biểu đồ tiếp tục trỏ đúng bản `— gốc` còn nguyên bố cục.
    """
    if not renames:
        return

    def fix(text: str) -> str:
        for old, new in renames.items():
            oq, nq = old.replace("'", "''"), new.replace("'", "''")
            text = text.replace(f"'{oq}'!", f"'{nq}'!")
        return text

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if isinstance(v, str) and v.startswith("=") and "!" in v:
                    nv = fix(v)
                    if nv != v:
                        c.value = nv
                hl = getattr(c, "hyperlink", None)
                if hl is not None:
                    for attr in ("target", "location"):
                        val = getattr(hl, attr, None)
                        if isinstance(val, str) and "!" in val:
                            nv = fix(val)
                            if nv != val:
                                setattr(hl, attr, nv)
    try:
        names = wb.defined_names
        for key in list(names.keys()):
            dn = names[key]
            if isinstance(getattr(dn, "value", None), str):
                nv = fix(dn.value)
                if nv != dn.value:
                    dn.value = nv
    except Exception:
        pass
    try:
        for ws in list(wb.worksheets) + list(getattr(wb, "chartsheets", [])):
            for chart in list(getattr(ws, "_charts", []) or []):
                for ser in list(getattr(chart, "series", []) or []):
                    for holder in (getattr(ser, "val", None), getattr(ser, "cat", None), getattr(ser, "tx", None)):
                        if holder is None:
                            continue
                        for kind in ("numRef", "strRef", "multiLvlStrRef"):
                            ref = getattr(holder, kind, None)
                            f = getattr(ref, "f", None)
                            if isinstance(f, str):
                                nf = fix(f)
                                if nf != f:
                                    ref.f = nf
    except Exception:
        pass


def _add_sorted_companion_sheets(
    wb,
    source_path: Path,
    rows: list[ComparedItem],
    meta: dict[str, dict],
) -> None:
    """Với mỗi sheet có hạng mục phát sinh nằm lẫn trong phần A, dựng bản SẮP XẾP
    mang TÊN GỐC theo ĐÚNG cấu trúc A / B / C của file chào giá.

    - Nhận diện phần A (theo KLMT), phần B (phát sinh ngoài KLMT), phần C (tổng
      cộng = A + B) theo STT + tên. Hạng mục phát sinh (không có trong PL01) đang
      nằm lẫn trong phần A được DỜI xuống phần B; giữ nguyên các mục B sẵn có.
      Nếu file chưa có phần B thì tự tạo tiêu đề B.
    - Format y hệt bản gốc: sao chép nguyên style/tiêu đề/đầu mục từng ô.
    - Công thức TRONG-DÒNG (thành tiền = KL × đơn giá...) được dịch địa chỉ theo
      dòng mới nên vẫn "sống". Ba dòng tổng A / B / C được VIẾT LẠI theo dải dòng
      mới (A, B là SUBTOTAL; C = A + B) nên số tiền luôn đúng.
    - Bản gốc được đổi tên '<tên> — gốc', giữ nguyên toàn bộ; mọi tham chiếu chéo
      (sheet Tổng hợp, hyperlink AI_KIEM_TRA, chart) được viết lại trỏ đúng bản gốc.
    """
    ps_rows = _phatsinh_block_rows(rows)
    cand_by_sheet: dict[str, dict[int, object]] = defaultdict(dict)
    for row in rows:
        cand = row.candidate
        if cand is not None:
            cand_by_sheet[cand.sheet].setdefault(cand.row_number, cand)

    candidate_sheets = [s for s in cand_by_sheet if ps_rows.get(s) and s in wb.sheetnames]
    if not candidate_sheets:
        return

    value_wb = load_workbook(source_path, data_only=True)
    used = {name.lower() for name in wb.sheetnames}
    promotions: list[tuple[str, object]] = []

    try:
        for sheet_name in candidate_sheets:
            src = wb[sheet_name]
            vsrc = value_wb[sheet_name] if sheet_name in value_wb.sheetnames else None
            info = meta.get(sheet_name) or {}
            fields = {str(k): int(v) for k, v in (info.get("field_columns") or {}).items()}
            header_end = int(info.get("header_end") or 1)
            stt_col = fields.get("stt", 1)
            name_col = fields.get("item_name", 2)
            ncol = src.max_column
            max_row = src.max_row

            def cell_text(r: int, c: int) -> str:
                v = (vsrc or src).cell(r, c).value
                return str(v if v is not None else "")

            # Nhận diện phần A (theo KLMT) / B (phát sinh ngoài) / C (tổng cộng)
            # theo STT + tên — đúng cấu trúc file chào giá gốc.
            a_row = b_row = c_row = None
            for r in range(header_end + 1, max_row + 1):
                stt = normalize_stt(cell_text(r, stt_col))
                nm = cell_text(r, name_col).lower()
                if a_row is None and stt == "A":
                    a_row = r
                if b_row is None and (stt == "B" or "phát sinh ngoài" in nm or "phat sinh ngoai" in nm):
                    b_row = r
                if c_row is None and (stt == "C" or "tổng cộng" in nm or "tong cong" in nm):
                    c_row = r

            boundary = b_row or c_row or (max_row + 1)
            ps_all = ps_rows.get(sheet_name, set())
            ps_move = sorted(r for r in ps_all if r < boundary)  # phát sinh nằm trong phần A
            if not ps_move:
                continue  # không có phát sinh cần dời -> không cần sheet sắp xếp

            new = wb.create_sheet(_suffixed_sheet_name(sheet_name, _SORTED_SUFFIX, used))
            for col in range(1, ncol + 1):
                letter = get_column_letter(col)
                dim = src.column_dimensions.get(letter)
                if dim is not None and dim.width:
                    new.column_dimensions[letter].width = dim.width

            def copy_row(dst_row: int, src_row: int, *, with_comment: bool = True, translate: bool = True) -> None:
                for col in range(1, ncol + 1):
                    s = src.cell(src_row, col)
                    v = s.value
                    if isinstance(v, str) and v.startswith("="):
                        translated = _translate_row_formula(v, get_column_letter(col), src_row, dst_row) if (translate and _same_row_formula(v, src_row)) else None
                        v = translated if translated is not None else (vsrc.cell(src_row, col).value if vsrc is not None else None)
                    elif type(v).__name__ in ("ArrayFormula", "DataTableFormula"):
                        v = vsrc.cell(src_row, col).value if vsrc is not None else None
                    d = new.cell(dst_row, col, v)
                    d._style = s._style
                    if with_comment and s.comment is not None:
                        d.comment = Comment(s.comment.text, s.comment.author or "HSMT Enterprise AI")
                if src.row_dimensions[src_row].height is not None:
                    new.row_dimensions[dst_row].height = src.row_dimensions[src_row].height

            # Cấu trúc chuẩn không nhận ra -> fallback: giữ nguyên, dồn phát sinh cuối.
            if a_row is None:
                out = 1
                for r in range(1, max_row + 1):
                    if r > header_end and r in ps_all:
                        continue
                    copy_row(out, r, with_comment=(r > header_end), translate=(r > header_end))
                    out += 1
                for r in sorted(ps_all):
                    copy_row(out, r)
                    out += 1
                new.freeze_panes = new.cell(header_end + 1, 1)
                promotions.append((sheet_name, new))
                continue

            move_set = set(ps_move)
            out = 1
            for r in range(1, a_row):  # tiêu đề + mọi dòng trước phần A
                copy_row(out, r, with_comment=(r > header_end), translate=(r > header_end))
                out += 1
            a_header_out = out
            copy_row(out, a_row, with_comment=False, translate=False)  # tiêu đề A (sửa tổng sau)
            out += 1
            a_item_start = out
            a_body_end = (b_row or c_row or (max_row + 1)) - 1
            for r in range(a_row + 1, a_body_end + 1):  # thân A: giữ đầu mục + hạng mục khớp, BỎ phát sinh
                if r in move_set:
                    continue
                copy_row(out, r)
                out += 1
            a_item_end = out - 1

            b_header_out = out
            if b_row is not None:
                copy_row(out, b_row, with_comment=False, translate=False)  # tiêu đề B sẵn có
                out += 1
                b_item_start = out
                for r in range(b_row + 1, (c_row or (max_row + 1))):  # các mục B sẵn có
                    copy_row(out, r)
                    out += 1
            else:
                for col in range(1, ncol + 1):  # tạo tiêu đề B mới, style theo tiêu đề A
                    new.cell(out, col)._style = src.cell(a_row, col)._style
                new.cell(out, stt_col, "B")
                new.cell(out, name_col, "Phát sinh ngoài KLMT (nhà thầu bổ sung)")
                out += 1
                b_item_start = out
            for r in ps_move:  # dời phát sinh của phần A xuống B
                copy_row(out, r)
                out += 1
            b_item_end = out - 1

            c_out = None
            if c_row is not None:
                c_out = out
                copy_row(out, c_row, with_comment=False, translate=False)  # tổng C (sửa sau)
                out += 1
                for r in range(c_row + 1, max_row + 1):  # phần đuôi sau C (nếu có)
                    copy_row(out, r)
                    out += 1

            # Viết lại 3 công thức tổng theo dải dòng MỚI (A, B là SUBTOTAL; C = A + B).
            for col in range(1, ncol + 1):
                letter = get_column_letter(col)
                a_f = src.cell(a_row, col).value
                if isinstance(a_f, str) and "SUBTOTAL" in a_f.upper() and a_item_end >= a_item_start:
                    new.cell(a_header_out, col, f"=SUBTOTAL(9,{letter}{a_item_start}:{letter}{a_item_end})")
                template = src.cell(b_row, col).value if b_row is not None else a_f
                if isinstance(template, str) and "SUBTOTAL" in template.upper() and b_item_end >= b_item_start:
                    new.cell(b_header_out, col, f"=SUBTOTAL(9,{letter}{b_item_start}:{letter}{b_item_end})")
                if c_out is not None:
                    c_f = src.cell(c_row, col).value
                    if isinstance(c_f, str) and c_f.startswith("="):
                        new.cell(c_out, col, f"={letter}{a_header_out}+{letter}{b_header_out}")

            new.freeze_panes = new.cell(header_end + 1, 1)
            promotions.append((sheet_name, new))
    finally:
        value_wb.close()

    # Bản sắp xếp mang TÊN GỐC; bản gốc đổi thành '— gốc' và mọi tham chiếu chéo
    # được viết lại để tiếp tục trỏ đúng bản gốc (openpyxl không tự làm như Excel).
    renames: dict[str, str] = {}
    used_names = {name.lower() for name in wb.sheetnames}
    for original, sorted_ws in promotions:
        old_ws = wb[original]
        goc_name = _suffixed_sheet_name(original, _GOC_SUFFIX, used_names)
        renames[original] = goc_name
        index = wb._sheets.index(old_ws)
        old_ws.title = goc_name
        sorted_ws.title = original
        wb._sheets.remove(sorted_ws)
        wb._sheets.insert(index, sorted_ws)
    _rewrite_sheet_references(wb, renames)


def _prepare_review_sheet(wb, bidder: str):
    for name in ("AI_TONG_QUAN", "AI_KIEM_TRA"):
        if name in wb.sheetnames:
            del wb[name]
    summary = wb.create_sheet("AI_TONG_QUAN", 0)
    review = wb.create_sheet("AI_KIEM_TRA", 1)

    summary["A1"] = "KẾT QUẢ KIỂM TRA HỒ SƠ CHÀO GIÁ"
    summary["A2"] = f"Nhà thầu: {bidder}"
    summary["A4"] = "Màu"
    summary["B4"] = "Ý nghĩa"
    legend = [
        (Severity.REVIEW, "Cần chuyên viên xác nhận"),
        (Severity.WARNING, "Sai lệch đáng kể"),
        (Severity.CRITICAL, "Thiếu, sai công thức hoặc chênh lệch nghiêm trọng"),
    ]
    for index, (severity, meaning) in enumerate(legend, 5):
        summary.cell(index, 1, severity.value)
        summary.cell(index, 1).fill = _FILL[severity]
        summary.cell(index, 1).font = Font(name="Arial", bold=True, color=_FONT[severity])
        summary.cell(index, 2, meaning)
    summary["A10"] = "Lưu ý"
    summary["B10"] = (
        "Các đánh dấu là tín hiệu hỗ trợ rà soát. Thương hiệu ngoài Phụ lục 02 không tự động bị loại; "
        "cần kiểm tra tài liệu chứng minh tương đương hoặc tốt hơn. Khác tên sheet chỉ được ghi chú, "
        "không được tính là cảnh báo khi hạng mục đã khớp."
    )
    summary.column_dimensions["A"].width = 25
    summary.column_dimensions["B"].width = 95
    summary.merge_cells("A1:H1")
    summary["A1"].font = Font(name="Arial", bold=True, size=16, color="FFFFFF")
    summary["A1"].fill = PatternFill("solid", fgColor="17365D")
    summary["A1"].alignment = Alignment(horizontal="center")

    headers = [
        "Mức độ", "Sheet gốc", "Dòng gốc", "STT", "Hạng mục PL01", "Hạng mục nhà thầu",
        "Thông số", "Giá trị yêu cầu/nhóm", "Giá trị nhà thầu", "Chênh lệch", "Chênh lệch (%)",
        "Lý do", "Liên kết tới dòng gốc",
    ]
    for col, header in enumerate(headers, 1):
        cell = review.cell(1, col, header)
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=_THIN)
    widths = [18, 24, 12, 14, 48, 48, 30, 35, 35, 18, 18, 80, 22]
    for col, width in enumerate(widths, 1):
        review.column_dimensions[get_column_letter(col)].width = width
    review.freeze_panes = "A2"
    review.auto_filter.ref = f"A1:M1"
    return summary, review


def annotate_bidder_workbook(
    source_path: str | Path,
    output_path: str | Path,
    bidder_workbook: WorkbookData,
    rows: list[ComparedItem],
) -> str:
    """Create an annotated copy while preserving original formulas and layout.

    The original workbook is never overwritten. Problematic cells are marked IN
    PLACE with a fill colour and a hover comment (chú thích) — NO extra columns
    are added next to the data. Two front sheets (AI_TONG_QUAN overview and
    AI_KIEM_TRA linked findings) are still provided.
    """
    source_path, output_path = Path(source_path), Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(source_path, data_only=False, keep_links=True)
    try:
        summary_ws, review_ws = _prepare_review_sheet(wb, bidder_workbook.bidder)
        meta = _sheet_meta(bidder_workbook)
        # CHỈ đánh dấu các dòng có sai lệch (severity khác OK). Hạng mục khớp —
        # kể cả khớp nhưng khác tên sheet (chỉ là ghi chú, không phải lỗi) — KHÔNG
        # được tô màu hay gắn chú thích trong file nhà thầu.
        grouped_by_location: dict[tuple[str, int], list[ComparedItem]] = defaultdict(list)
        for compared in rows:
            if compared.candidate is None or compared.severity is Severity.OK:
                continue
            location = (compared.candidate.sheet, compared.candidate.row_number)
            grouped_by_location[location].append(compared)

        review_row = 2
        counts = defaultdict(int)
        for compared in rows:
            if compared.severity is Severity.OK:
                continue
            counts[compared.severity.value] += 1
            candidate, reference = compared.candidate, compared.reference
            diffs = compared.differences
            fields = list(dict.fromkeys(diff.field for diff in diffs))
            ref_values = list(dict.fromkeys(str(diff.reference_value) for diff in diffs if diff.reference_value not in (None, "")))
            cand_values = list(dict.fromkeys(str(diff.candidate_value) for diff in diffs if diff.candidate_value not in (None, "")))
            deltas = [diff.delta for diff in diffs if isinstance(diff.delta, (int, float))]
            delta_pcts = [diff.delta_pct for diff in diffs if isinstance(diff.delta_pct, (int, float))]
            reasons = list(dict.fromkeys(compared.flags or [compared.match.reason]))
            values = [
                compared.severity.value,
                candidate.sheet if candidate else "",
                candidate.row_number if candidate else None,
                candidate.stt if candidate else (reference.stt if reference else ""),
                reference.item_name if reference else "",
                candidate.item_name if candidate else "",
                " | ".join(fields),
                " | ".join(ref_values)[:8000],
                " | ".join(cand_values)[:8000],
                max(deltas, key=abs) if deltas else None,
                max(delta_pcts, key=abs) if delta_pcts else None,
                " | ".join(reasons)[:30000],
                "Mở dòng gốc" if candidate else "Không có dòng để mở",
            ]
            for col, value in enumerate(values, 1):
                cell = review_ws.cell(review_row, col, value)
                cell.alignment = Alignment(vertical="top", wrap_text=col in {5, 6, 7, 8, 9, 12})
                if col == 1:
                    cell.fill = _FILL.get(compared.severity, PatternFill())
                    cell.font = Font(name="Arial", bold=True, color=_FONT.get(compared.severity, "000000"))
                if col == 11 and isinstance(value, (int, float)):
                    cell.number_format = "0.00%"
                if col == 10 and isinstance(value, (int, float)):
                    cell.number_format = "#,##0.000;[Red]-#,##0.000"
            if candidate:
                escaped = candidate.sheet.replace("'", "''")
                review_ws.cell(review_row, 13).hyperlink = f"#'{escaped}'!A{candidate.row_number}"
                review_ws.cell(review_row, 13).style = "Hyperlink"
            review_row += 1

        # Add exact spreadsheet formula/external-link issues, including cells
        # outside parsed BOQ rows. These are detected by direct OOXML scanning.
        for issue in bidder_workbook.formula_issues:
            kind = str(issue.get("kind", ""))
            severity = Severity.CRITICAL if kind == "FORMULA_ERROR" else Severity.REVIEW
            counts[severity.value] += 1
            sheet_name = str(issue.get("sheet", ""))
            row_number = int(issue.get("row", 0) or 0)
            cell_ref = str(issue.get("cell", ""))
            message = str(issue.get("message", ""))
            formula = str(issue.get("formula", ""))
            values = [
                severity.value, sheet_name, row_number, "", "", "",
                f"Lỗi Excel tại ô {cell_ref}", formula, str(issue.get("value", "")),
                None, None, message, "Mở ô lỗi",
            ]
            for col, value in enumerate(values, 1):
                cell = review_ws.cell(review_row, col, value)
                cell.alignment = Alignment(vertical="top", wrap_text=col in {7, 8, 9, 12})
                if col == 1:
                    cell.fill = _FILL[severity]
                    cell.font = Font(name="Arial", bold=True, color=_FONT[severity])
            if sheet_name and cell_ref and sheet_name in wb.sheetnames:
                escaped = sheet_name.replace("'", "''")
                review_ws.cell(review_row, 13).hyperlink = f"#'{escaped}'!{cell_ref}"
                review_ws.cell(review_row, 13).style = "Hyperlink"
            review_row += 1

        summary_ws["A12"] = "Số dòng cần kiểm tra"
        summary_ws["B12"] = counts.get(Severity.REVIEW.value, 0)
        summary_ws["A13"] = "Số dòng cảnh báo"
        summary_ws["B13"] = counts.get(Severity.WARNING.value, 0)
        summary_ws["A14"] = "Số dòng bất thường"
        summary_ws["B14"] = counts.get(Severity.CRITICAL.value, 0)
        summary_ws["A16"] = "Tổng dòng trong AI_KIEM_TRA"
        summary_ws["B16"] = max(0, review_row - 2)
        summary_ws["A17"] = "Lỗi công thức/liên kết Excel"
        summary_ws["B17"] = len(bidder_workbook.formula_issues)
        summary_ws["A18"] = "External links trong workbook"
        summary_ws["B18"] = bidder_workbook.external_link_count

        # Bản đồ cột cho từng sheet — KHÔNG thêm bất kỳ cột nào vào file dữ liệu.
        # Chỉ dùng để biết ô giá trị nào cần tô màu và gắn chú thích trực tiếp.
        sheet_fields: dict[str, dict[str, int]] = {}
        for sheet_name, info in meta.items():
            if sheet_name not in wb.sheetnames:
                continue
            sheet_fields[sheet_name] = {str(k): int(v) for k, v in (info.get("field_columns") or {}).items()}

        # CHỈ tô màu + gắn CHÚ THÍCH lên ĐÚNG những ô có sai lệch. Mỗi sai lệch chỉ
        # đánh dấu đúng ô của nó (tên hạng mục khác -> bôi ô tên; khối lượng khác ->
        # bôi ô khối lượng...). KHÔNG bôi ô tên hạng mục nếu tên không sai. Sai lệch
        # không gắn với ô cụ thể (ví dụ một số thông số kỹ thuật) chỉ được liệt kê ở
        # sheet AI_KIEM_TRA, không tô màu bừa lên dòng.
        for (sheet_name, row_number), location_rows in grouped_by_location.items():
            if sheet_name not in wb.sheetnames or sheet_name not in sheet_fields:
                continue
            ws = wb[sheet_name]
            fields = sheet_fields[sheet_name]
            for compared in location_rows:
                for diff in compared.differences:
                    fill = _FILL.get(diff.severity)
                    if fill is None:
                        continue  # sai lệch mức OK -> không đánh dấu
                    key = _field_key(diff)
                    col = fields.get(key) if key else None
                    if not col:
                        continue  # không xác định được ô -> để AI_KIEM_TRA liệt kê
                    cell = ws.cell(row_number, col)
                    cell.fill = fill
                    _append_comment(cell, f"{diff.field}: {diff.message}"[:2000])

        # Highlight exact error cells and also write the AI reason on that row.
        for issue in bidder_workbook.formula_issues:
            sheet_name = str(issue.get("sheet", ""))
            cell_ref = str(issue.get("cell", ""))
            kind = str(issue.get("kind", ""))
            message = str(issue.get("message", ""))
            severity = Severity.CRITICAL if kind == "FORMULA_ERROR" else Severity.REVIEW
            if sheet_name not in wb.sheetnames or not cell_ref:
                continue
            ws = wb[sheet_name]
            target = ws[cell_ref]
            target.fill = _FILL[severity]
            target.font = Font(
                name="Arial", size=target.font.sz, bold=True,
                italic=target.font.italic, color=_FONT[severity], underline=target.font.underline,
            )
            _append_comment(target, message)

        review_ws.auto_filter.ref = f"A1:M{max(1, review_row - 1)}"

        # Với mỗi sheet có hạng mục phát sinh: dựng bản sắp xếp mang TÊN GỐC (dồn
        # phát sinh xuống cuối, công thức trong-dòng sống, tổng dựng lại); bản gốc
        # đổi tên '— gốc' giữ nguyên công thức và mọi tham chiếu được trỏ lại đúng.
        _add_sorted_companion_sheets(wb, source_path, rows, meta)

        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.save(output_path)
    finally:
        wb.close()
    return str(output_path)

