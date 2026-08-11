"""Gom bảng chào giá thô thành một bảng giá sạch để nạp vào CSDL giá.

Không gắn cứng dự án, nhà thầu hay đường dẫn nào: thư mục nguồn và siêu dữ liệu
đều truyền qua tham số dòng lệnh, tên nhà thầu đoán bằng core.bidder_name dùng
chung với web.

    python scripts/preprocess_raw_data.py <thư_mục_chào_giá> [-o data/cleaned_prices.xlsx]
        [--project "Tên dự án"] [--project-type "..."] [--year 2025] [--region "Miền Bắc"]
"""

import argparse
import os
import re
import sys
from pathlib import Path

import openpyxl
import pandas as pd

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from core.bidder_name import (  # noqa: E402
    guess_bidder_from_column,
    guess_bidder_name,
    strip_shared_tokens,
)

# Khoảng đơn giá hợp lệ: dưới cận dưới thường là số lượng/thứ tự bị đọc nhầm,
# trên cận trên thường là dòng tổng cộng. Chỉnh được cho ngành có đơn giá khác.
DEFAULT_MIN_PRICE = 500.0
DEFAULT_MAX_PRICE = 5_000_000_000.0

# Sheet không chứa bảng khối lượng. Bổ sung qua HSMT_SKIP_SHEETS (cách nhau dấu phẩy).
SKIP_SHEETS = {"tổng hợp", "tong hop", "cover", "trang bìa", "trang bia"} | {
    s.strip().lower() for s in os.getenv("HSMT_SKIP_SHEETS", "").split(",") if s.strip()
}

_parser = argparse.ArgumentParser(description=__doc__,
                                  formatter_class=argparse.RawDescriptionHelpFormatter)
_parser.add_argument("source", help="thư mục chứa các file chào giá (.xlsx)")
_parser.add_argument("-o", "--output", default="data/cleaned_prices.xlsx",
                     help="tệp Excel kết quả (mặc định: data/cleaned_prices.xlsx)")
_parser.add_argument("--project", default="", help="tên dự án ghi vào từng bản ghi")
_parser.add_argument("--project-type", default="", help="loại công trình")
_parser.add_argument("--year", type=int, default=None, help="năm của mặt bằng giá")
_parser.add_argument("--region", default="", help="vùng miền của mặt bằng giá")
_parser.add_argument("--min-price", type=float, default=DEFAULT_MIN_PRICE)
_parser.add_argument("--max-price", type=float, default=DEFAULT_MAX_PRICE)
_args = _parser.parse_args()

base_dir = Path(_args.source).expanduser().resolve()
output_file = Path(_args.output).expanduser().resolve()
output_dir = output_file.parent

if not base_dir.is_dir():
    _parser.error(f"Không thấy thư mục nguồn: {base_dir}")

print("=== TIỀN XỬ LÝ DỮ LIỆU BÁO GIÁ THÔ ===")
print(f"Nguồn : {base_dir}")
print(f"Kết quả: {output_file}")

os.makedirs(output_dir, exist_ok=True)

# Các từ khóa nhận diện đơn vị tính chuẩn
UNIT_MAPPING = {
    "m": "m", "mét": "m", "met": "m", "m.": "m", "md": "m", "mét dài": "m",
    "cái": "cái", "cai": "cái", "chiếc": "cái", "chiec": "cái", "pc": "cái", "pcs": "cái",
    "bộ": "bộ", "bo": "bộ", "set": "bộ",
    "lô": "lô", "lo": "lô",
    "kg": "kg", "kilogam": "kg", "kí": "kg",
    "cuộn": "cuộn", "cuon": "cuộn",
    "bình": "bình", "binh": "bình",
    "quả": "quả", "qua": "quả",
    "hộp": "hộp", "hop": "hộp",
    "mét vuông": "m2", "m2": "m2", "mv": "m2",
    "mét khối": "m3", "m3": "m3", "mk": "m3",
}

def clean_unit(unit_str):
    if not unit_str:
        return ""
    val = str(unit_str).strip().lower().replace("\n", "").replace("\r", "")
    val = re.sub(r'\s+', ' ', val)
    return UNIT_MAPPING.get(val, val)

def clean_price(price_val):
    if price_val is None:
        return None
    if isinstance(price_val, (int, float)):
        return float(price_val)
    text = str(price_val).strip().replace(".", "").replace(",", "").replace("đ", "").replace("VND", "").replace("VNĐ", "").replace(" ", "")
    try:
        return float(text)
    except ValueError:
        return None

def clean_text(text_val):
    if not text_val:
        return ""
    val = str(text_val).strip().replace("\n", " ").replace("\r", " ")
    val = re.sub(r'\s+', ' ', val)
    return val

# Danh sách chứa toàn bộ bản ghi sạch sau xử lý
cleaned_records = []

# Đọc các file thô trong thư mục
xlsx_files = sorted(list(base_dir.glob("*.xlsx")))
if not xlsx_files:
    print(f"ERROR: Không tìm thấy file Excel nào trong {base_dir}")
    exit(1)

# Tên nhà thầu theo file: đoán cho TOÀN BỘ file rồi mới bỏ phần dùng chung, nhờ
# vậy tên dự án lặp ở mọi file tự rụng — xử lý từng file riêng thì không thấy được.
# Quét cả thư mục thường lẫn file khác (bảng tổng hợp, file kết quả) nên tên dự
# án không xuất hiện ở đủ 100% tên file — dùng ngưỡng đa số thay vì tuyệt đối.
_file_bidders = dict(zip(
    xlsx_files,
    strip_shared_tokens([guess_bidder_name(f.name) for f in xlsx_files], min_ratio=0.6),
))
print("Nhà thầu nhận ra từ tên file:")
for _f, _name in _file_bidders.items():
    print(f"  {_name or '(không rõ)':<28} <- {_f.name[:58]}")

for file_path in xlsx_files:
    print(f"\nProcessing file: {file_path.name}")
    try:
        # Load workbook dùng openpyxl
        wb = openpyxl.load_workbook(str(file_path), data_only=True, read_only=True)
        
        file_records_count = 0
        
        # Duyệt qua toàn bộ các sheet trong file (Trừ các sheet Tổng hợp)
        for sheet_name in wb.sheetnames:
            if sheet_name.strip().lower() in SKIP_SHEETS:
                continue
                
            ws = wb[sheet_name]
            
            # Bỏ qua các sheet dạng Chart (Chartsheet) không có dữ liệu bảng
            if not hasattr(ws, "iter_rows"):
                continue
                
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                continue
                
            # Quét tìm tiêu đề bằng cách ghép 2 dòng kề nhau (Hỗ trợ tiêu đề phức hợp/nested)
            # Thay vì ngắt sớm ở dòng đầu tiên khớp, ta tìm dòng có điểm khớp cột tối đa (tối ưu hóa độ chính xác)
            best_row_idx = None
            best_col_map = {}
            max_score = 0
            best_is_merged = False
            
            for i in range(min(25, len(rows) - 1)):
                headers = []
                for col_idx in range(max(len(rows[i]), len(rows[i+1]))):
                    val1 = str(rows[i][col_idx] or "").strip() if col_idx < len(rows[i]) else ""
                    val2 = str(rows[i+1][col_idx] or "").strip() if col_idx < len(rows[i+1]) else ""
                    headers.append(f"{val1} {val2}".strip().lower())
                
                temp_map = {}
                composite_price_cols = []
                fallback_price_cols = []
                
                for col_idx, val in enumerate(headers):
                    if any(k in val for k in ["tên hạng mục", "tên vật tư", "diễn giải", "nội dung công việc"]):
                        temp_map["item_name"] = col_idx
                    elif any(k in val for k in ["đvt", "đơn vị"]):
                        temp_map["unit"] = col_idx
                    elif any(k in val for k in ["quy cách", "thông số", "mô tả/ quy cách", "spec"]):
                        temp_map["material_spec"] = col_idx
                    elif any(k in val for k in ["thương hiệu", "hãng", "brand"]):
                        temp_map["brand"] = col_idx
                    elif any(k in val for k in ["xuất xứ", "origin"]):
                        temp_map["origin"] = col_idx
                    
                    # Match price columns
                    if any(k in val for k in ["đơn giá tổng hợp", "đg tổng hợp", "đg thầu", "đơn giá thầu", "đơn giá", "unit price", "đg"]):
                        if not any(x in val for x in ["vl chính", "vl phụ", "nhân công", "nc", "máy", "quản lý", "lợi nhuận"]):
                            composite_price_cols.append((col_idx, val))
                    elif any(k in val for k in ["vl chính", "vật liệu chính"]):
                        fallback_price_cols.append((col_idx, val))
                
                if composite_price_cols:
                    temp_map["prices"] = composite_price_cols
                elif fallback_price_cols:
                    temp_map["prices"] = fallback_price_cols
                
                if "item_name" in temp_map and "unit" in temp_map and "prices" in temp_map:
                    score = len(temp_map)  # Number of mapped columns (max 6)
                    # We give extra priority to headers containing technical specs/brands to avoid matches on title banners
                    if "material_spec" in temp_map: score += 1
                    if "brand" in temp_map: score += 1
                    
                    if score > max_score:
                        max_score = score
                        best_row_idx = i
                        best_col_map = temp_map
                        best_is_merged = True
            
            # Nếu ghép dòng không tìm thấy, thử quét dòng đơn lẻ
            if best_row_idx is None:
                for i in range(min(25, len(rows))):
                    row_vals = [str(c or "").strip().lower() for c in rows[i]]
                    temp_map = {}
                    composite_price_cols = []
                    fallback_price_cols = []
                    
                    for col_idx, val in enumerate(row_vals):
                        if any(k in val for k in ["tên hạng mục", "tên vật tư", "diễn giải", "nội dung"]):
                            temp_map["item_name"] = col_idx
                        elif any(k in val for k in ["đvt", "đơn vị"]):
                            temp_map["unit"] = col_idx
                        elif any(k in val for k in ["quy cách", "thông số", "spec"]):
                            temp_map["material_spec"] = col_idx
                        elif any(k in val for k in ["thương hiệu", "hãng", "brand"]):
                            temp_map["brand"] = col_idx
                        elif any(k in val for k in ["xuất xứ", "origin"]):
                            temp_map["origin"] = col_idx
                        
                        # Match price columns
                        if any(k in val for k in ["đơn giá tổng hợp", "đg tổng hợp", "đg thầu", "đơn giá thầu", "đơn giá", "unit price", "đg"]):
                            if not any(x in val for x in ["vl chính", "vl phụ", "nhân công", "nc", "máy", "quản lý", "lợi nhuận"]):
                                composite_price_cols.append((col_idx, val))
                        elif any(k in val for k in ["vl chính", "vật liệu chính"]):
                            fallback_price_cols.append((col_idx, val))
                            
                    if composite_price_cols:
                        temp_map["prices"] = composite_price_cols
                    elif fallback_price_cols:
                        temp_map["prices"] = fallback_price_cols
                        
                    if "item_name" in temp_map and "prices" in temp_map:
                        score = len(temp_map)
                        if "material_spec" in temp_map: score += 1
                        if "brand" in temp_map: score += 1
                        
                        if score > max_score:
                            max_score = score
                            best_row_idx = i
                            best_col_map = temp_map
                            best_is_merged = False
            
            # Gán kết quả tốt nhất tìm được
            if best_row_idx is not None:
                header_row_idx = best_row_idx
                col_map = best_col_map
                is_merged_header = best_is_merged
            else:
                continue
                
            # Đọc dữ liệu từ sheet
            start_data_row = header_row_idx + (2 if is_merged_header else 1)
            sheet_records = 0
            
            for r_idx in range(start_data_row, len(rows)):
                row = rows[r_idx]
                if not row:
                    continue
                    
                item_name_idx = col_map.get("item_name")
                if item_name_idx is None or item_name_idx >= len(row):
                    continue
                    
                item_name = clean_text(row[item_name_idx])
                if not item_name or item_name.isdigit() or len(item_name) < 2:
                    continue
                    
                # Bỏ qua các dòng tiêu đề chương/mục thầu
                if item_name.lower().startswith(("chương", "phần", "mục", "i.", "ii.", "iii.", "iv.", "v.", "a.", "b.", "c.")):
                    has_any_price = False
                    if "prices" in col_map:
                        for p_col, _ in col_map["prices"]:
                            if p_col < len(row) and clean_price(row[p_col]) is not None:
                                has_any_price = True
                                break
                    if not has_any_price:
                        continue
                
                # Trích xuất các trường thông tin
                unit = clean_unit(row[col_map["unit"]]) if "unit" in col_map and col_map["unit"] < len(row) else ""
                spec = clean_text(row[col_map["material_spec"]]) if "material_spec" in col_map and col_map["material_spec"] < len(row) else ""
                brand = clean_text(row[col_map["brand"]]) if "brand" in col_map and col_map["brand"] < len(row) else ""
                origin = clean_text(row[col_map["origin"]]) if "origin" in col_map and col_map["origin"] < len(row) else ""
                
                # Quét các cột đơn giá
                for price_col_idx, col_title in col_map["prices"]:
                    if price_col_idx >= len(row):
                        continue
                    price_val = clean_price(row[price_col_idx])
                    
                    if price_val is not None and _args.min_price <= price_val <= _args.max_price:
                        # Tiêu đề cột nêu đích danh nhà thầu thì ưu tiên (một file
                        # tổng hợp có thể chứa nhiều cột giá của nhiều nhà thầu);
                        # không có thì lấy tên suy từ tên file.
                        contractor = (guess_bidder_from_column(col_title)
                                      or _file_bidders.get(file_path, ""))

                        rec = {
                            "item_name": item_name,
                            "item_code": "",
                            "unit": unit,
                            "unit_price": price_val,
                            "total_price": None,
                            "quantity": None,
                            "project_name": _args.project or base_dir.name,
                            "project_type": _args.project_type,
                            "year": _args.year,
                            "region": _args.region,
                            "brand": f"{brand} ({contractor})" if brand and contractor else (brand or contractor),
                            "origin": origin,
                            "material_spec": spec,
                            "source_file": f"{file_path.name} | {sheet_name}"
                        }
                        cleaned_records.append(rec)
                        sheet_records += 1
                        file_records_count += 1
            
            if sheet_records > 0:
                print(f"    - Sheet '{sheet_name}': Trích xuất được {sheet_records:,} dòng")
                
        print(f"  -> Tổng cộng file {file_path.name}: trích xuất {file_records_count:,} dòng giá")
        wb.close()
    except Exception as e:
        print(f"  ✗ Lỗi khi xử lý file: {e}")

# 4. Lưu dữ liệu sạch
df = pd.DataFrame(cleaned_records)
if df.empty:
    print("\n✗ Không trích xuất được bất kỳ dữ liệu nào!")
    exit(1)

initial_count = len(df)
df.drop_duplicates(subset=["item_name", "unit", "unit_price", "brand", "material_spec"], keep="first", inplace=True)
final_count = len(df)

print(f"\n📊 THỐNG KÊ LÀM SẠCH:")
print(f"  - Tổng số bản ghi thô thu thập: {initial_count:,}")
print(f"  - Số bản ghi sau khi loại trùng : {final_count:,} (Giảm {(initial_count - final_count):,} bản ghi trùng lặp)")

df.to_excel(output_file, index=False)
print(f"\n✓ Đã xuất tệp dữ liệu sạch thành công ra: {output_file}")
print("=== TIỀN XỬ LÝ HOÀN TẤT ===")
