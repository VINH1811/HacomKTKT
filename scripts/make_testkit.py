"""Sinh bộ file đầu vào để thử toàn bộ chức năng của hệ thống.

Bộ file cố ý cài sẵn 8 lỗi (thiếu khối lượng, sai thương hiệu, lệch thành phần
giá, sai phép tính, cùng hạng mục hai đơn giá, hạng mục phát sinh, giá bất
thường giữa hai nhà thầu, thiếu hạng mục) để đối chiếu kết quả chấm.

    python scripts/make_testkit.py [đường_dẫn_file_zip]
"""
import argparse
import io
import shutil
import sys
import zipfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

_parser = argparse.ArgumentParser(description=__doc__)
_parser.add_argument("output", nargs="?", default="reports/BO_FILE_TEST_HSMT.zip",
                     help="nơi ghi tệp zip (mặc định: reports/BO_FILE_TEST_HSMT.zip)")
ZIP_OUT = Path(_parser.parse_args().output).resolve()
ZIP_OUT.parent.mkdir(parents=True, exist_ok=True)
WORK = ZIP_OUT.with_suffix("") / "_dung"

HDR = Font(bold=True, color="FFFFFF", size=10)
FILL = PatternFill("solid", fgColor="1F4E78")
GRP = PatternFill("solid", fgColor="2F75B5")
THIN = Side(style="thin", color="B4C6E7")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

shutil.rmtree(WORK, ignore_errors=True)
WORK.mkdir(parents=True, exist_ok=True)


# ---------------------------------------------------------------- PL01 (KLMT)
def make_pl01():
    wb = Workbook(); ws = wb.active; ws.title = "KLMT PCCC"
    ws.append(["DỰ ÁN: CÔNG TRÌNH MẪU"])
    ws.append(["GÓI THẦU: PHÒNG CHÁY CHỮA CHÁY"])
    ws.append([])
    ws.append(["STT", "Mã hiệu", "Diễn giải", "Đơn vị", "Khối lượng mời thầu", "Ghi chú"])
    rows = [
        ("1", "DI-M9102", "Đầu báo khói địa chỉ", "cái", 120),
        ("2", "DI-M9103", "Đầu báo nhiệt địa chỉ", "cái", 45),
        ("3", "FRD-012", "Đầu sprinkler hướng xuống", "cái", 180),
        ("4", "C-9314P", "Đèn chỉ thị báo cháy phòng", "bộ", 80),
        ("5", "FC-10L", "Tủ trung tâm báo cháy 10 loop", "bộ", 1),
        ("6", "ST-100", "Ống thép tráng kẽm DN100", "m", 350),
        ("7", "BC-8", "Bình chữa cháy bột ABC 8kg", "bình", 40),
        ("8", "VG-50", "Van góc chữa cháy DN50", "cái", 20),
        ("9", "CB-6", "Chuông báo cháy 6 inch", "cái", 25),
        ("10", "TDK-3B", "Tủ điều khiển bơm chữa cháy", "bộ", 1),
    ]
    for r in rows:
        ws.append(list(r) + [""])
    for c in ws[4]:
        if c.value:
            c.font = HDR; c.fill = FILL; c.alignment = CENTER; c.border = BOX
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 38
    for col in "DEF":
        ws.column_dimensions[col].width = 16
    wb.save(WORK / "1_PL01_Khoi_luong_moi_thau.xlsx")


# ---------------------------------------------------------------- PL02 (VTTB)
def make_pl02():
    wb = Workbook(); ws = wb.active; ws.title = "DMVT PCCC"
    ws.append(["PHỤ LỤC 02: DANH MỤC VẬT TƯ THIẾT BỊ"])
    ws.append(["DỰ ÁN: CÔNG TRÌNH MẪU"])
    ws.append([])
    ws.append(["STT", "VẬT TƯ THIẾT BỊ", "THƯƠNG HIỆU - XUẤT XỨ", "GHI CHÚ"])
    rows = [
        ("1", "Đầu báo khói địa chỉ", "GST - China / Hochiki - Japan", "hoặc tương đương"),
        ("2", "Đầu báo nhiệt địa chỉ", "GST - China / Hochiki - Japan", "hoặc tương đương"),
        ("3", "Đầu sprinkler", "Forede - China / Tyco - USA", "hoặc tương đương"),
        ("4", "Tủ trung tâm báo cháy", "GST - China / Notifier - USA", "hoặc tương đương"),
        ("5", "Ống thép tráng kẽm", "Hòa Phát - Việt Nam", "hoặc tương đương"),
        ("6", "Bình chữa cháy", "Tomoken - Việt Nam", "hoặc tương đương"),
    ]
    for r in rows:
        ws.append(list(r))
    for c in ws[4]:
        if c.value:
            c.font = HDR; c.fill = FILL; c.alignment = CENTER; c.border = BOX
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 34
    ws.column_dimensions["D"].width = 20
    wb.save(WORK / "2_PL02_Danh_muc_vat_tu.xlsx")


# ------------------------------------------------------------- Hồ sơ chào giá
def bid_sheet(ws, title, rows):
    """Header ba tầng giống hồ sơ thật: nhóm gộp / tên cột con / đánh số."""
    ws.append(["DỰ ÁN: CÔNG TRÌNH MẪU"])
    ws.append([f"GÓI THẦU: {title}"])
    ws.append([])
    ws.append(["Stt", "Mô tả công việc mời thầu", "ĐVT", None, None,
               "Thông tin VTTB", None, None, None,
               "Đơn giá", None, None, None, None, None, "Thành tiền"])
    ws.append([None, None, None, "KLMT", "Nhà thầu chào",
               "Mô tả quy cách", "Mã hiệu", "Thương hiệu", "Xuất xứ",
               "VL chính", "VL phụ", "NC& máy TC", "CP quản lý", "Lợi nhuận",
               "Đơn giá tổng hợp", "Thành tiền NT chào"])
    ws.append([str(i) for i in range(1, 17)])
    ws.merge_cells("F4:I4"); ws.merge_cells("J4:O4")
    for cell in ws[4] + ws[5]:
        if cell.value:
            cell.font = HDR
            cell.fill = GRP if cell.row == 4 else FILL
            cell.alignment = CENTER; cell.border = BOX

    for spec in rows:
        (stt, name, unit, klmt, kl, quycach, code, brand, origin,
         vl, nc, ql, ln, force_dg, force_tt, note) = spec
        vlp = round(vl * 0.03)
        dg = vl + vlp + nc + ql + ln if force_dg is None else force_dg
        tt = (kl or 0) * dg if force_tt is None else force_tt
        ws.append([stt, name, unit, klmt, kl, quycach, code, brand, origin,
                   vl, vlp, nc, ql, ln, dg, tt])
        ws.cell(ws.max_row, 17, note)

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 7
    for col in "DEFGHIJKLMNOP":
        ws.column_dimensions[col].width = 13
    ws.column_dimensions["Q"].width = 58
    ws.freeze_panes = "A7"


def make_bidder_a():
    """Nhà thầu A — cài đủ các loại lỗi để thử."""
    wb = Workbook(); ws = wb.active; ws.title = "PCCC"
    R = [
        # (stt, tên, đvt, KLMT, KL chào, quy cách, mã, hãng, xuất xứ, VL, NC, QL, LN, épĐG, épTT, ghi chú)
        ("1", "Đầu báo khói địa chỉ", "cái", 120, 120, "Loại địa chỉ", "DI-M9102", "GST", "China",
         500_000, 200_000, 30_000, 25_240, None, None, "Bình thường"),
        ("2", "Đầu báo nhiệt địa chỉ", "cái", 45, 45, "Loại địa chỉ", "DI-M9103", "GST", "China",
         480_000, 190_000, 28_000, 24_000, None, None, "Bình thường"),
        ("3", "Đầu sprinkler hướng xuống", "cái", 180, 180, "K80, 68 độ C", "FRD-012", "Forede", "China",
         150_000, 70_000, 8_000, 7_000, None, None, "Bình thường"),
        ("4", "Đèn chỉ thị báo cháy phòng", "bộ", 80, 80, "LED 24VDC", "C-9314P", "GST", "China",
         420_000, 180_000, 25_000, 32_000, None, None, "Bình thường"),
        ("5", "Tủ trung tâm báo cháy 10 loop", "bộ", 1, 1, "10 loop", "FC-10L", "GST", "China",
         30_000_000, 8_000_000, 1_500_000, 1_200_000, None, None, "Bình thường"),
        ("6", "Ống thép tráng kẽm DN100", "m", 350, 300, "DN100 dày 3.0mm", "ST-100", "Hòa Phát", "Việt Nam",
         280_000, 90_000, 12_000, 11_600, None, None,
         "LỖI 1: khối lượng chào 300 ≠ KLMT 350 (thiếu 14%)"),
        ("7", "Bình chữa cháy bột ABC 8kg", "bình", 40, 40, "MFZ8", "BC-8", "Kidde", "USA",
         300_000, 120_000, 20_000, 15_000, None, None,
         "LỖI 2: thương hiệu Kidde không có trong Phụ lục 02 (yêu cầu Tomoken)"),
        ("8", "Van góc chữa cháy DN50", "cái", 20, 20, "DN50 đồng", "VG-50", "Shanxi", "China",
         900_000, 300_000, 40_000, 100_000, 1_400_000, None,
         "LỖI 3: tổng 5 thành phần = 1.367.000 nhưng ghi đơn giá 1.400.000"),
        ("9", "Chuông báo cháy 6 inch", "cái", 25, 25, "24VDC", "CB-6", "GST", "China",
         200_000, 80_000, 12_000, 14_000, None, 99_999_999,
         "LỖI 4: thành tiền 99.999.999 ≠ 25 × đơn giá"),
        ("10", "Tủ điều khiển bơm chữa cháy", "bộ", 1, 1, "3 bơm", "TDK-3B", "Schneider", "Pháp",
         55_000_000, 12_000_000, 2_000_000, 1_800_000, None, None, "Bình thường"),
        ("11", "Đầu báo khói địa chỉ", "cái", None, 30, "Loại địa chỉ", "DI-M9102", "GST", "China",
         560_000, 220_000, 32_000, 21_200, None, None,
         "LỖI 5: cùng hạng mục dòng 7 nhưng đơn giá cao hơn ~10% (lệch giá nội bộ)"),
        ("12", "Giá đỡ ống inox bổ sung", "bộ", None, 60, "Inox 304", "GD-INOX", "Hòa Phát", "Việt Nam",
         120_000, 40_000, 6_000, 5_000, None, None,
         "LỖI 6: hạng mục PHÁT SINH, không có trong Phụ lục 01"),
    ]
    bid_sheet(ws, "PHÒNG CHÁY CHỮA CHÁY", R)
    wb.save(WORK / "3_Nha_thau_A_chao_gia.xlsx")


def make_bidder_b():
    """Nhà thầu B — giá chênh lệch để thử so sánh ngang; thiếu 1 hạng mục."""
    wb = Workbook(); ws = wb.active; ws.title = "PCCC"
    R = [
        ("1", "Đầu báo khói địa chỉ", "cái", 120, 120, "Loại địa chỉ", "DI-M9102", "Hochiki", "Japan",
         900_000, 300_000, 45_000, 62_000, None, None,
         "Giá cao hơn nhà thầu A khoảng 60% → so sánh ngang sẽ cảnh báo"),
        ("2", "Đầu báo nhiệt địa chỉ", "cái", 45, 45, "Loại địa chỉ", "DI-M9103", "Hochiki", "Japan",
         500_000, 195_000, 29_000, 25_000, None, None, "Bình thường"),
        ("3", "Đầu sprinkler hướng xuống", "cái", 180, 180, "K80, 68 độ C", "FRD-012", "Tyco", "USA",
         160_000, 72_000, 8_500, 7_500, None, None, "Bình thường"),
        ("4", "Đèn chỉ thị báo cháy phòng", "bộ", 80, 80, "LED 24VDC", "C-9314P", "GST", "China",
         430_000, 182_000, 25_500, 32_800, None, None, "Bình thường"),
        ("5", "Tủ trung tâm báo cháy 10 loop", "bộ", 1, 1, "10 loop", "FC-10L", "Notifier", "USA",
         42_000_000, 9_000_000, 1_800_000, 1_500_000, None, None, "Bình thường"),
        ("6", "Ống thép tráng kẽm DN100", "m", 350, 350, "DN100 dày 3.0mm", "ST-100", "Hòa Phát", "Việt Nam",
         285_000, 91_000, 12_200, 11_800, None, None, "Bình thường"),
        ("7", "Bình chữa cháy bột ABC 8kg", "bình", 40, 40, "MFZ8", "BC-8", "Tomoken", "Việt Nam",
         310_000, 122_000, 20_500, 15_300, None, None, "Bình thường"),
        ("8", "Van góc chữa cháy DN50", "cái", 20, 20, "DN50 đồng", "VG-50", "Shanxi", "China",
         920_000, 305_000, 41_000, 101_000, None, None, "Bình thường"),
        ("9", "Chuông báo cháy 6 inch", "cái", 25, 25, "24VDC", "CB-6", "GST", "China",
         205_000, 81_000, 12_200, 14_300, None, None, "Bình thường"),
        # thiếu hạng mục 10 (Tủ điều khiển bơm) -> test phát hiện thiếu
    ]
    bid_sheet(ws, "PHÒNG CHÁY CHỮA CHÁY", R)
    wb.save(WORK / "4_Nha_thau_B_chao_gia.xlsx")


def make_bidder_a_v2():
    """Bản chào giá lần 2 của nhà thầu A — để thử So sánh phiên bản."""
    wb = Workbook(); ws = wb.active; ws.title = "PCCC"
    R = [
        ("1", "Đầu báo khói địa chỉ", "cái", 120, 120, "Loại địa chỉ", "DI-M9102", "GST", "China",
         520_000, 205_000, 31_000, 26_000, None, None, "TĂNG giá so với bản 1"),
        ("2", "Đầu báo nhiệt địa chỉ", "cái", 45, 45, "Loại địa chỉ", "DI-M9103", "GST", "China",
         480_000, 190_000, 28_000, 24_000, None, None, "Giữ nguyên"),
        ("3", "Đầu sprinkler hướng xuống", "cái", 180, 180, "K80, 68 độ C", "FRD-012", "Forede", "China",
         150_000, 70_000, 8_000, 7_000, None, None, "Giữ nguyên"),
        ("4", "Đèn chỉ thị báo cháy phòng", "bộ", 80, 80, "LED 24VDC", "C-9314P", "GST", "China",
         420_000, 180_000, 25_000, 32_000, None, None, "Giữ nguyên"),
        ("5", "Tủ trung tâm báo cháy 10 loop", "bộ", 1, 1, "10 loop", "FC-10L", "GST", "China",
         30_000_000, 8_000_000, 1_500_000, 1_200_000, None, None, "Giữ nguyên"),
        ("6", "Ống thép tráng kẽm DN100", "m", 350, 350, "DN100 dày 3.0mm", "ST-100", "Hòa Phát", "Việt Nam",
         280_000, 90_000, 12_000, 11_600, None, None, "ĐÃ SỬA: chào đủ 350m"),
        ("7", "Bình chữa cháy bột ABC 8kg", "bình", 40, 40, "MFZ8", "BC-8", "Tomoken", "Việt Nam",
         300_000, 120_000, 20_000, 15_000, None, None, "ĐÃ SỬA: đổi sang Tomoken đúng PL02"),
        ("8", "Van góc chữa cháy DN50", "cái", 20, 20, "DN50 đồng", "VG-50", "Shanxi", "China",
         900_000, 300_000, 40_000, 100_000, None, None, "ĐÃ SỬA: đơn giá khớp tổng thành phần"),
        ("9", "Chuông báo cháy 6 inch", "cái", 25, 25, "24VDC", "CB-6", "GST", "China",
         200_000, 80_000, 12_000, 14_000, None, None, "ĐÃ SỬA: thành tiền đúng"),
        ("10", "Tủ điều khiển bơm chữa cháy", "bộ", 1, 1, "3 bơm", "TDK-3B", "Schneider", "Pháp",
         55_000_000, 12_000_000, 2_000_000, 1_800_000, None, None, "Giữ nguyên"),
        ("11", "Van báo động DN100", "cái", None, 4, "DN100", "VBD-100", "Shanxi", "China",
         3_200_000, 900_000, 150_000, 212_500, None, None, "THÊM MỚI so với bản 1"),
    ]
    bid_sheet(ws, "PHÒNG CHÁY CHỮA CHÁY", R)
    wb.save(WORK / "5_Nha_thau_A_chao_gia_LAN_2.xlsx")


# ------------------------------------------------------------------- Làm rõ
def make_rfi():
    head = ["STT", "NỘI DUNG ĐÁNH GIÁ THEO HSYC", "YÊU CẦU",
            "Nhà thầu A kê khai", "Ý kiến CĐT", "Nhà thầu trả lời làm rõ"]
    body = [
        ["A", "ĐÁNH GIÁ TÍNH HỢP LỆ", "", "", "", ""],
        ["1", "Hiệu lực hồ sơ chào giá", "90 ngày", "60 ngày",
         "Đề nghị nhà thầu gia hạn hiệu lực lên 90 ngày", ""],
        ["B", "ĐÁNH GIÁ NĂNG LỰC", "", "", "", ""],
        ["2", "Hợp đồng tương tự", "02 hợp đồng", "Kê khai 03 hợp đồng",
         "Đề nghị bổ sung bản sao công chứng của 03 hợp đồng", ""],
        ["3", "Doanh thu bình quân 3 năm", "≥ 50 tỷ", "42 tỷ",
         "Đề nghị giải trình doanh thu chưa đạt yêu cầu", ""],
        ["C", "ĐÁNH GIÁ KỸ THUẬT", "", "", "", ""],
        ["4", "Thương hiệu bình chữa cháy", "Theo Phụ lục 02", "Kidde - USA",
         "Đề nghị làm rõ căn cứ chào thương hiệu ngoài Phụ lục 02", ""],
    ]
    for fname, answers in (
        ("6_Lam_ro_YEU_CAU_cua_CDT.xlsx", ["", "", "", "", ""]),
        ("7_Lam_ro_PHAN_HOI_cua_nha_thau_A.xlsx",
         ["Nhà thầu đồng ý gia hạn hiệu lực lên 90 ngày",
          "Đã bổ sung bản sao công chứng 03 hợp đồng (đính kèm)",
          "",   # cố ý BỎ TRỐNG -> hệ thống phải báo chưa trả lời
          "Kidde đạt tiêu chuẩn tương đương, đính kèm chứng chỉ UL"]),
    ):
        wb = Workbook(); ws = wb.active; ws.title = "PL 1"
        ws.append(["DỰ ÁN: CÔNG TRÌNH MẪU"])
        ws.append(["GÓI THẦU: PHÒNG CHÁY CHỮA CHÁY"])
        ws.append([])
        ws.append(head)
        idx = 0
        for row in body:
            r = list(row)
            if r[4] and fname.startswith("7"):
                r[5] = answers[idx] if idx < len(answers) else ""
                idx += 1
            ws.append(r)
        for c in ws[4]:
            if c.value:
                c.font = HDR; c.fill = FILL; c.alignment = CENTER; c.border = BOX
        ws.column_dimensions["A"].width = 6
        for col, w in zip("BCDEF", (34, 18, 26, 42, 42)):
            ws.column_dimensions[col].width = w
        wb.save(WORK / fname)


# ------------------------------------------------------- Hồ sơ ZIP (checklist)
def make_dossier_zip():
    files = [
        "1. Don chao gia/Don chao gia.pdf",
        "2. Bang chao gia/BOQ PCCC chi tiet.xlsx",
        "3. Danh muc vat tu/Danh muc vat tu thiet bi.pdf",
        "4. Tai chinh/Bao cao tai chinh nam 2023.pdf",
        "4. Tai chinh/Bao cao tai chinh nam 2024.pdf",
        "5. Nang luc/Giay chung nhan dang ky doanh nghiep.pdf",
        "6. Hop dong tuong tu/Bang ke khai hop dong tuong tu.pdf",
        "7. Nhan su/So do to chuc va nhan su chu chot.pdf",
        "8. Tien do/Bang tien do thi cong.pdf",
        "9. Bien phap thi cong/Bien phap thi cong tong the.pdf",
        "10. Catalogue/Catalogue GST.pdf",
        # cố ý THIẾU: an toàn lao động; báo cáo tài chính chỉ có 2/3 năm
    ]
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        for f in files:
            z.writestr(f, b"noi dung mau")
    (WORK / "8_Ho_so_nang_luc_Nha_thau_A.zip").write_bytes(buf.getvalue())


# ---------------------------------------------------------------- Hướng dẫn
def make_guide():
    wb = Workbook(); ws = wb.active; ws.title = "HƯỚNG DẪN"
    G = [
        ("BỘ FILE THỬ CHỨC NĂNG – HỆ THỐNG ĐÁNH GIÁ HỒ SƠ CHÀO GIÁ", ""),
        ("", ""),
        ("CHỨC NĂNG", "FILE CẦN TẢI LÊN"),
        ("1. Phụ lục & nhà thầu (1 nhà thầu)",
         "Phụ lục 01 = file 1  ·  Phụ lục 02 = file 2  ·  Hồ sơ nhà thầu = file 3"),
        ("2. Phụ lục & nhà thầu (2 nhà thầu)",
         "Phụ lục 01 = file 1  ·  Phụ lục 02 = file 2  ·  Hồ sơ nhà thầu = file 3 và file 4"),
        ("3. So sánh nhà thầu", "Chọn file 3 và file 4"),
        ("4. HSMT với HSDT", "HSMT = file 1  ·  HSDT = file 3"),
        ("5. So sánh phiên bản", "Bản cũ = file 3  ·  Bản mới = file 5"),
        ("6. Theo dõi làm rõ (RFI)", "File yêu cầu = file 6  ·  File phản hồi = file 7"),
        ("7. Checklist hồ sơ", "Tải file 8 (định dạng ZIP)"),
        ("", ""),
        ("LỖI ĐÃ CÀI SẴN – KẾT QUẢ ĐÚNG PHẢI BẮT ĐƯỢC", ""),
        ("Trong hồ sơ nhà thầu A (file 3)", ""),
        ("  Lỗi 1", "Ống thép DN100: chào 300m nhưng khối lượng mời thầu 350m (thiếu ~14%)"),
        ("  Lỗi 2", "Bình chữa cháy: thương hiệu Kidde không có trong Phụ lục 02"),
        ("  Lỗi 3", "Van góc DN50: tổng 5 thành phần giá lệch đơn giá tổng hợp"),
        ("  Lỗi 4", "Chuông báo cháy: thành tiền khác khối lượng × đơn giá"),
        ("  Lỗi 5", "Đầu báo khói địa chỉ xuất hiện 2 lần với 2 đơn giá khác nhau (chênh ~10%)"),
        ("  Lỗi 6", "Giá đỡ ống inox: hạng mục phát sinh, không có trong Phụ lục 01"),
        ("Trong hồ sơ nhà thầu B (file 4)", ""),
        ("  Lỗi 7", "Đầu báo khói giá cao hơn nhà thầu A ~60% (chỉ hiện khi so sánh 2 nhà thầu)"),
        ("  Lỗi 8", "Thiếu hạng mục Tủ điều khiển bơm chữa cháy so với Phụ lục 01"),
        ("Khi so sánh phiên bản (file 3 và 5)", ""),
        ("  Kỳ vọng", "Đầu báo khói TĂNG giá · Ống thép và bình chữa cháy ĐÃ SỬA · thêm mới Van báo động"),
        ("Khi theo dõi làm rõ (file 6 và 7)", ""),
        ("  Kỳ vọng", "4 yêu cầu, 3 đã trả lời, 1 CHƯA trả lời (mục doanh thu bình quân)"),
        ("Khi chạy checklist (file 8)", ""),
        ("  Kỳ vọng", "Thiếu An toàn lao động · Báo cáo tài chính chỉ có 2/3 năm"),
        ("", ""),
        ("GHI CHÚ", "Cột Q trong các file chào giá ghi rõ lý do từng dòng, dùng để đối chiếu kết quả."),
    ]
    for i, (a, b) in enumerate(G, 1):
        ws.cell(i, 1, a); ws.cell(i, 2, b)
    ws.cell(1, 1).font = Font(bold=True, size=14, color="1F4E78")
    for r in (3, 12, 13, 20, 23, 25, 27, 30):
        ws.cell(r, 1).font = Font(bold=True)
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 92
    wb.save(WORK / "0_HUONG_DAN_SU_DUNG.xlsx")


make_pl01(); make_pl02(); make_bidder_a(); make_bidder_b()
make_bidder_a_v2(); make_rfi(); make_dossier_zip(); make_guide()

with zipfile.ZipFile(ZIP_OUT, "w", zipfile.ZIP_DEFLATED) as z:
    for f in sorted(WORK.iterdir()):
        z.write(f, f.name)
shutil.rmtree(WORK, ignore_errors=True)

# Console Windows mặc định cp1252 sẽ vỡ khi in tiếng Việt.
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

print("Đã tạo:", ZIP_OUT)
with zipfile.ZipFile(ZIP_OUT) as z:
    for n in z.namelist():
        print("   ", n)
