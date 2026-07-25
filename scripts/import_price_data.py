"""Import historical price data from Excel/CSV into the PriceAdvisor database.

Usage
-----
::

    # Import from Excel
    python scripts/import_price_data.py --input data/prices.xlsx

    # Import from CSV
    python scripts/import_price_data.py --input data/prices.csv --encoding utf-8

    # Specify column mapping
    python scripts/import_price_data.py --input data/prices.xlsx \\
        --col-name "Tên vật tư" --col-unit "ĐVT" --col-price "Đơn giá"

    # Show database stats
    python scripts/import_price_data.py --stats

Expected columns (auto-detected, case-insensitive):
    - Tên hạng mục / Tên vật tư / Mô tả / item_name
    - Mã hiệu / item_code
    - ĐVT / Đơn vị / unit
    - Đơn giá / unit_price
    - Thành tiền / total_price
    - Khối lượng / quantity
    - Dự án / project_name
    - Năm / year
    - Khu vực / region
    - Thương hiệu / brand
    - Xuất xứ / origin
    - Quy cách / material_spec
"""
from __future__ import annotations

import sys
try:
    sys.stdout.reconfigure(encoding='utf-8')
    sys.stderr.reconfigure(encoding='utf-8')
except Exception:
    pass

import argparse
import csv
import os
import time
from pathlib import Path
from typing import Any, Optional

# Add project root to path
PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from dotenv import load_dotenv
load_dotenv(PROJECT_ROOT / ".env", override=True)


# Column name mappings (lowercase) → field name
_COLUMN_MAP: dict[str, str] = {
    # item_name
    "tên hạng mục": "item_name",
    "tên vật tư": "item_name",
    "mô tả": "item_name",
    "diễn giải": "item_name",
    "item_name": "item_name",
    "name": "item_name",
    # item_code
    "mã hiệu": "item_code",
    "mã": "item_code",
    "item_code": "item_code",
    "code": "item_code",
    # unit
    "đvt": "unit",
    "đơn vị": "unit",
    "đơn vị tính": "unit",
    "unit": "unit",
    # unit_price
    "đơn giá": "unit_price",
    "đơn giá tổng hợp": "unit_price",
    "unit_price": "unit_price",
    # total_price
    "thành tiền": "total_price",
    "total_price": "total_price",
    "amount": "total_price",
    # quantity
    "khối lượng": "quantity",
    "kl": "quantity",
    "quantity": "quantity",
    "qty": "quantity",
    # project_name
    "dự án": "project_name",
    "công trình": "project_name",
    "project_name": "project_name",
    "project": "project_name",
    # project_type
    "loại công trình": "project_type",
    "project_type": "project_type",
    # year
    "năm": "year",
    "year": "year",
    # region
    "khu vực": "region",
    "vùng": "region",
    "region": "region",
    # brand
    "thương hiệu": "brand",
    "brand": "brand",
    # origin
    "xuất xứ": "origin",
    "origin": "origin",
    # material_spec
    "quy cách": "material_spec",
    "quy cách kỹ thuật": "material_spec",
    "vật tư": "material_spec",
    "material_spec": "material_spec",
    "spec": "material_spec",
}


def _detect_columns(headers: list[str]) -> dict[int, str]:
    """Map column indices to field names based on header text."""
    mapping: dict[int, str] = {}
    for idx, header in enumerate(headers):
        normalized = header.strip().lower().replace("\n", " ").replace("\r", " ")
        # 1. Exact match
        if normalized in _COLUMN_MAP:
            mapping[idx] = _COLUMN_MAP[normalized]
            continue
        # 2. Substring matches with prioritization
        if any(kw in normalized for kw in ["tên hạng mục", "tên vật tư", "diễn giải"]):
            mapping[idx] = "item_name"
        elif "đơn giá" in normalized or "đơn giá" in normalized:
            mapping[idx] = "unit_price"
        elif "đơn vị" in normalized or "đvt" in normalized:
            mapping[idx] = "unit"
        elif "mã hiệu" in normalized or "mã vật tư" in normalized:
            mapping[idx] = "item_code"
        elif any(kw in normalized for kw in ["thương hiệu", "hiệu", "hãng", "brand"]):
            mapping[idx] = "brand"
        elif "xuất xứ" in normalized or "origin" in normalized:
            mapping[idx] = "origin"
        elif "quy cách" in normalized or "spec" in normalized or "vật liệu chính" in normalized:
            mapping[idx] = "material_spec"
        elif "năm" in normalized or "year" in normalized:
            mapping[idx] = "year"
        elif "dự án" in normalized or "project" in normalized:
            mapping[idx] = "project_name"
        elif "khu vực" in normalized or "region" in normalized:
            mapping[idx] = "region"
        elif "mô tả" in normalized:
            # If item_name is already mapped, map to material_spec
            if "item_name" in mapping.values():
                mapping[idx] = "material_spec"
            else:
                mapping[idx] = "item_name"
    return mapping


def _parse_float(value: Any) -> Optional[float]:
    """Parse a float from various formats (Vietnamese number conventions)."""
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value or "").strip().replace(",", "").replace(" ", "")
    if not text or text == "-":
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _parse_int(value: Any) -> Optional[int]:
    """Parse an integer from various formats."""
    f = _parse_float(value)
    return int(f) if f is not None else None


def read_excel(path: Path, sheet: Optional[str] = None) -> list[dict[str, Any]]:
    """Read an Excel file and return a list of row dicts."""
    try:
        import openpyxl
    except ImportError:
        print("ERROR: Cần cài openpyxl: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        print(f"WARNING: File {path.name} chỉ có {len(rows)} dòng, bỏ qua.")
        return []

    # 1. Tìm dòng tiêu đề tốt nhất (hoặc đơn lẻ, hoặc ghép 2 dòng kề nhau)
    best_row_idx = 0
    best_col_map = {}
    max_detected = 0
    is_merged_header = False

    # Thử quét từng dòng đơn lẻ trong 25 dòng đầu
    for i in range(min(25, len(rows))):
        headers = [str(cell or "").strip() for cell in rows[i]]
        col_map = _detect_columns(headers)
        detected_count = len(col_map)
        if "item_name" in col_map.values():
            detected_count += 2
        if "unit_price" in col_map.values():
            detected_count += 2
        if detected_count > max_detected:
            max_detected = detected_count
            best_row_idx = i
            best_col_map = col_map
            is_merged_header = False

    # Thử quét ghép 2 dòng kề nhau
    for i in range(min(24, len(rows) - 1)):
        headers = []
        for col_idx in range(max(len(rows[i]), len(rows[i+1]))):
            val1 = str(rows[i][col_idx] or "").strip() if col_idx < len(rows[i]) else ""
            val2 = str(rows[i+1][col_idx] or "").strip() if col_idx < len(rows[i+1]) else ""
            if val1 and val2:
                combined = f"{val1} {val2}"
            else:
                combined = val1 or val2
            headers.append(combined)
        
        col_map = _detect_columns(headers)
        detected_count = len(col_map)
        if "item_name" in col_map.values():
            detected_count += 2
        if "unit_price" in col_map.values():
            detected_count += 2
        if detected_count > max_detected:
            max_detected = detected_count
            best_row_idx = i
            best_col_map = col_map
            is_merged_header = True

    # Fallback nếu không phát hiện được gì
    if "item_name" not in best_col_map.values() or max_detected < 3:
        headers = [str(cell or "").strip() for cell in rows[0]]
        col_map = _detect_columns(headers)
        header_row_offset = 1
        start_row = 1
    else:
        col_map = best_col_map
        header_row_offset = 2 if is_merged_header else 1
        start_row = best_row_idx + header_row_offset

    if "item_name" not in col_map.values():
        print(f"WARNING: Không tìm thấy cột tên hạng mục trong {path.name}")
        return []

    records = []
    for row_idx, row in enumerate(rows[start_row:], start=start_row + 1):
        data: dict[str, Any] = {}
        for col_idx, field in col_map.items():
            if col_idx < len(row):
                data[field] = row[col_idx]
        name = str(data.get("item_name") or "").strip()
        # Bỏ qua các dòng trống hoặc dòng chỉ chứa số thứ tự (ví dụ: '1', '2'...)
        if name and not name.isdigit():
            records.append(data)
    return records


def read_csv(path: Path, encoding: str = "utf-8") -> list[dict[str, Any]]:
    """Read a CSV file and return a list of row dicts."""
    records = []
    with open(path, "r", encoding=encoding, newline="") as f:
        reader = csv.reader(f)
        headers = next(reader, [])
        headers = [h.strip() for h in headers]
        col_map = _detect_columns(headers)

        if "item_name" not in col_map.values():
            print(f"WARNING: Không tìm thấy cột tên hạng mục trong {path.name}")
            print(f"  Headers: {headers}")
            return []

        for row in reader:
            data: dict[str, Any] = {}
            for col_idx, field in col_map.items():
                if col_idx < len(row):
                    data[field] = row[col_idx]
            if data.get("item_name") and str(data["item_name"]).strip():
                records.append(data)
    return records


def main():
    parser = argparse.ArgumentParser(
        description="Nạp dữ liệu giá lịch sử vào PriceAdvisor database",
    )
    parser.add_argument("--input", "-i", type=str, help="Đường dẫn file Excel (.xlsx) hoặc CSV")
    parser.add_argument("--sheet", type=str, default=None, help="Tên sheet trong Excel (mặc định: sheet đầu tiên)")
    parser.add_argument("--encoding", type=str, default="utf-8", help="Encoding cho CSV (mặc định: utf-8)")
    parser.add_argument("--project", type=str, default="", help="Tên dự án (gắn cho tất cả bản ghi)")
    parser.add_argument("--year", type=int, default=None, help="Năm (gắn cho tất cả bản ghi)")
    parser.add_argument("--region", type=str, default="", help="Khu vực (gắn cho tất cả bản ghi)")
    parser.add_argument("--stats", action="store_true", help="Chỉ hiển thị thống kê database")
    parser.add_argument("--no-embed", action="store_true", help="Bỏ qua bước tạo embedding (nạp nhanh)")
    args = parser.parse_args()

    from core.price_advisor.config import PriceAdvisorConfig
    from core.price_advisor.database import PriceDatabase
    from core.price_advisor.models import PriceRecord

    config = PriceAdvisorConfig.from_env()
    db = PriceDatabase(config.db_path, config=config)

    if args.stats:
        stats = db.get_stats()
        print("\n═══ Thống kê Kho giá PriceAdvisor ═══")
        print(f"  Tổng bản ghi     : {stats['total_records']:,}")
        print(f"  Có embedding     : {stats['records_with_embedding']:,}")
        print(f"  Năm phủ          : {stats['year_min']} – {stats['year_max']}")
        print(f"  Số dự án         : {stats['project_count']}")
        print(f"  Feedback entries  : {stats['feedback_entries']}")
        print(f"  DB path          : {stats['db_path']}")
        return

    if not args.input:
        parser.print_help()
        sys.exit(1)

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"ERROR: File không tồn tại: {input_path}")
        sys.exit(1)

    # Read data
    print(f"\n📂 Đọc file: {input_path.name}")
    suffix = input_path.suffix.lower()
    if suffix in {".xlsx", ".xls"}:
        raw_records = read_excel(input_path, sheet=args.sheet)
    elif suffix == ".csv":
        raw_records = read_csv(input_path, encoding=args.encoding)
    else:
        print(f"ERROR: Định dạng không hỗ trợ: {suffix}. Chỉ nhận .xlsx hoặc .csv")
        sys.exit(1)

    print(f"  → Đọc được {len(raw_records):,} bản ghi")

    if not raw_records:
        print("WARNING: Không có bản ghi nào để nạp.")
        return

    # Convert to PriceRecord objects
    records: list[PriceRecord] = []
    skipped = 0
    for raw in raw_records:
        name = str(raw.get("item_name", "")).strip()
        if not name:
            skipped += 1
            continue
        record = PriceRecord(
            item_name=name,
            item_code=str(raw.get("item_code", "")).strip(),
            unit=str(raw.get("unit", "")).strip(),
            unit_price=_parse_float(raw.get("unit_price")),
            total_price=_parse_float(raw.get("total_price")),
            quantity=_parse_float(raw.get("quantity")),
            project_name=args.project or str(raw.get("project_name", "")).strip(),
            project_type=str(raw.get("project_type", "")).strip(),
            year=args.year or _parse_int(raw.get("year")),
            region=args.region or str(raw.get("region", "")).strip(),
            brand=str(raw.get("brand", "")).strip(),
            origin=str(raw.get("origin", "")).strip(),
            material_spec=str(raw.get("material_spec", "")).strip(),
            source_file=input_path.name,
        )
        records.append(record)

    print(f"  → {len(records):,} bản ghi hợp lệ, {skipped} bỏ qua (thiếu tên)")

    # Generate embeddings
    if not args.no_embed and config.api_key:
        try:
            from core.price_advisor.embedder import Embedder

            embedder = Embedder(config)
            descriptions = [r.description_text for r in records]
            print(f"\n🔄 Đang tạo embedding cho {len(descriptions):,} bản ghi...")

            batch_size = 100
            start_time = time.perf_counter()
            for i in range(0, len(descriptions), batch_size):
                batch = descriptions[i:i + batch_size]
                embeddings = embedder.embed_batch(batch)
                for j, emb in enumerate(embeddings):
                    records[i + j].embedding = emb
                done = min(i + batch_size, len(descriptions))
                pct = done * 100 // len(descriptions)
                print(f"  [{pct:3d}%] {done:,}/{len(descriptions):,}", end="\r")

            elapsed = time.perf_counter() - start_time
            print(f"\n  → Hoàn tất embedding trong {elapsed:.1f}s")
        except Exception as exc:
            print(f"\n⚠ Lỗi tạo embedding: {exc}")
            print("  Tiếp tục nạp dữ liệu KHÔNG có embedding. Chạy lại sau để bổ sung.")
    elif not config.api_key:
        print("\n⚠ Không có API key. Nạp dữ liệu KHÔNG có embedding.")
        print("  Đặt PRICE_ADVISOR_API_KEY trong .env rồi chạy lại để tạo embedding.")

    # Insert into database
    print(f"\n💾 Đang nạp {len(records):,} bản ghi vào database...")
    inserted = db.insert_records(records)
    print(f"  → Đã nạp {inserted:,} bản ghi")

    # Show stats
    stats = db.get_stats()
    print(f"\n═══ Thống kê sau nạp ═══")
    print(f"  Tổng bản ghi     : {stats['total_records']:,}")
    print(f"  Có embedding     : {stats['records_with_embedding']:,}")
    print(f"  Năm phủ          : {stats['year_min']} – {stats['year_max']}")
    print(f"  Số dự án         : {stats['project_count']}")
    print(f"  DB path          : {stats['db_path']}")
    print("\n✅ Hoàn tất!")
    db.close()


if __name__ == "__main__":
    main()
