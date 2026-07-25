import openpyxl
from pathlib import Path

base_dir = Path(r"c:\KHMT\HacomHolding\5. Tong hop chao gia 11.12.2025-20260628T172908Z-3-001\5. Tong hop chao gia 11.12.2025")
file_path = base_dir / "1. 2025.12.08 Chao gia ME Hacom Mall Linh Anh V2.xlsx"

print("=== DIAGNOSING LINH ANH EXCEL EXTRACTION ===")
if not file_path.exists():
    print(f"Error: File {file_path.name} not found.")
    exit(1)

wb = openpyxl.load_workbook(str(file_path), data_only=True, read_only=True)
ws = wb["1. HT điện"]
rows = list(ws.iter_rows(values_only=True))

# Print row 6 & 7 (headers)
print("\nRow 6 (1-indexed, Python index 5):")
print([str(c) for c in rows[5][:20]])
print("\nRow 7 (1-indexed, Python index 6):")
print([str(c) for c in rows[6][:20]])

# Let's search for "Giá treo trục ngang"
found = False
for idx, row in enumerate(rows):
    item_name = str(row[1] or "").strip()
    if "Giá treo trục ngang" in item_name:
        print(f"\nFound row at Python index {idx} (Row {idx+1}):")
        for col_idx, val in enumerate(row):
            if val is not None:
                print(f"  Col {col_idx} ({openpyxl.utils.get_column_letter(col_idx+1)}): {val}")
        found = True

if not found:
    print("\nCould not find 'Giá treo trục ngang' in Sheet 1. HT điện.")

wb.close()
