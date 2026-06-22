import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class ReportBuilder:
    def __init__(self):
        # Palette colors
        self.navy_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        self.light_gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        # Soft alert fills
        self.soft_red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        self.soft_red_font = Font(name="Arial", size=10, color="9C0006")
        
        self.soft_yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        self.soft_yellow_font = Font(name="Arial", size=10, color="9C6500")
        
        self.soft_green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        self.soft_green_font = Font(name="Arial", size=10, color="006100")
        
        # Standard fonts
        self.title_font = Font(name="Arial", size=16, bold=True, color="1F4E78")
        self.section_font = Font(name="Arial", size=12, bold=True, color="333333")
        self.header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        self.sub_header_font = Font(name="Arial", size=10, bold=True, color="333333")
        self.bold_font = Font(name="Arial", size=10, bold=True)
        self.regular_font = Font(name="Arial", size=10)
        
        # Borders
        thin_side = Side(border_style="thin", color="D9D9D9")
        self.thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        thick_bottom_side = Side(border_style="medium", color="1F4E78")
        self.header_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thick_bottom_side)
        
        double_bottom_side = Side(border_style="double", color="333333")
        top_thin_side = Side(border_style="thin", color="333333")
        self.total_border = Border(top=top_thin_side, bottom=double_bottom_side)

    def style_row(self, ws, row_idx, font, alignment, fill=None, border=None):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = font
            cell.alignment = alignment
            if fill:
                cell.fill = fill
            if border:
                cell.border = border

    def auto_fit_columns(self, ws):
        max_rows = min(200, ws.max_row)
        max_cols = ws.max_column
        col_widths = [0] * max_cols
        
        for row in ws.iter_rows(max_row=max_rows, max_col=max_cols, values_only=True):
            for col_idx, val in enumerate(row):
                if val is not None:
                    val_str = str(val)
                    if len(val_str) <= 50:
                        col_widths[col_idx] = max(col_widths[col_idx], len(val_str))
                        
        for col_idx, width in enumerate(col_widths, 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = max(width + 3, 12)

    def build_summary_sheet(self, wb, contractors, summary_data, comments):
        ws = wb.active
        ws.title = "Tổng hợp kết quả"
        
        # Title
        ws.cell(row=2, column=1, value="BÁO CÁO SO SÁNH TỔNG HỢP GIÁ CHÀO THẦU").font = self.title_font
        ws.cell(row=3, column=1, value="Dự án: Tòa nhà hỗn hợp Hacom Mall").font = self.regular_font
        ws.row_dimensions[2].height = 25
        
        # Headers
        headers_row1 = ["Hệ thống"]
        headers_row2 = [""]
        for c in contractors:
            headers_row1.extend([c, "", "", ""])
            headers_row2.extend(["Theo KLMT", "Nhà thầu chào", "Phát sinh ngoài", "Tổng cộng"])
            
        ws.append([]) # row 4
        
        # Append headers to worksheet
        ws.append(headers_row1) # row 5
        ws.append(headers_row2) # row 6
        ws.row_dimensions[5].height = 20
        ws.row_dimensions[6].height = 20
        
        # Merge header cells for contractors
        col_idx = 2
        for c in contractors:
            ws.merge_cells(start_row=5, start_column=col_idx, end_row=5, end_column=col_idx + 3)
            col_idx += 4
            
        # Format Headers
        for r_idx in [5, 6]:
            self.style_row(ws, r_idx, self.header_font, Alignment(horizontal="center", vertical="center", wrap_text=True), self.navy_fill, self.thin_border)
            
        # Write System rows
        start_data_row = 7
        for sys_idx, row in enumerate(summary_data):
            r_idx = start_data_row + sys_idx
            ws.cell(row=r_idx, column=1, value=row['Hệ thống'])
            
            col_idx = 2
            for c in contractors:
                ws.cell(row=r_idx, column=col_idx, value=row[f'{c}_TheoKLMT']).number_format = '#,##0'
                ws.cell(row=r_idx, column=col_idx + 1, value=row[f'{c}_Chào']).number_format = '#,##0'
                ws.cell(row=r_idx, column=col_idx + 2, value=row[f'{c}_PhátSinh']).number_format = '#,##0'
                ws.cell(row=r_idx, column=col_idx + 3, value=row[f'{c}_TổngCộng']).number_format = '#,##0'
                col_idx += 4
                
            self.style_row(ws, r_idx, self.regular_font, Alignment(horizontal="right", vertical="center"), border=self.thin_border)
            ws.cell(row=r_idx, column=1).alignment = Alignment(horizontal="left")
            ws.cell(row=r_idx, column=1).font = self.bold_font
            ws.row_dimensions[r_idx].height = 20

        # Write Total row
        total_row_idx = start_data_row + len(summary_data)
        ws.cell(row=total_row_idx, column=1, value="TỔNG CỘNG")
        
        col_idx = 2
        for c in contractors:
            theo_klmt_tot = sum(row[f'{c}_TheoKLMT'] for row in summary_data)
            chao_tot = sum(row[f'{c}_Chào'] for row in summary_data)
            ps_tot = sum(row[f'{c}_PhátSinh'] for row in summary_data)
            tong_tot = sum(row[f'{c}_TổngCộng'] for row in summary_data)
            
            ws.cell(row=total_row_idx, column=col_idx, value=theo_klmt_tot).number_format = '#,##0'
            ws.cell(row=total_row_idx, column=col_idx + 1, value=chao_tot).number_format = '#,##0'
            ws.cell(row=total_row_idx, column=col_idx + 2, value=ps_tot).number_format = '#,##0'
            ws.cell(row=total_row_idx, column=col_idx + 3, value=tong_tot).number_format = '#,##0'
            col_idx += 4
            
        self.style_row(ws, total_row_idx, self.bold_font, Alignment(horizontal="right", vertical="center"), self.light_gray_fill, self.total_border)
        ws.cell(row=total_row_idx, column=1).alignment = Alignment(horizontal="left")
        ws.row_dimensions[total_row_idx].height = 22

        # Write Automated Comments block
        comment_start_row = total_row_idx + 3
        ws.cell(row=comment_start_row, column=1, value="NHẬN XÉT TỰ ĐỘNG").font = self.section_font
        
        current_row = comment_start_row + 1
        for line in comments.split('\n'):
            if not line.strip():
                continue
            
            # Format markdown headers or bullets
            clean_line = line.replace('###', '').replace('**', '').replace('-', '•').strip()
            ws.cell(row=current_row, column=1, value=clean_line).font = self.bold_font if '•' not in line else self.regular_font
            
            # Merge comment cells across columns for readability
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(contractors)*4 + 1)
            current_row += 1
            
        self.auto_fit_columns(ws)

    def build_flags_sheet(self, wb, sheet_name, flags, flag_type_filter, title, headers, col_keys):
        """
        Generic function to build specialized reports sheets (Thiếu hạng mục, Sai khác KL, vv).
        """
        ws = wb.create_sheet(title=sheet_name)
        
        # Title
        ws.cell(row=2, column=1, value=title).font = self.title_font
        ws.row_dimensions[2].height = 25
        ws.append([]) # row 3
        
        # Headers
        ws.append(headers) # row 4
        self.style_row(ws, 4, self.header_font, Alignment(horizontal="center", vertical="center", wrap_text=True), self.navy_fill, self.thin_border)
        ws.row_dimensions[4].height = 22
        
        # Filter flags
        filtered_flags = [f for f in flags if f['type'] in flag_type_filter]
        
        # Write rows
        for flag in filtered_flags:
            row_vals = [flag.get(key, "") for key in col_keys]
            ws.append(row_vals)
            row_num = ws.max_row
            ws.row_dimensions[row_num].height = 20
            
            # Style the data row
            fill = self.soft_red_fill if flag['severity'] == 'HIGH' else self.soft_yellow_fill
            font = self.soft_red_font if flag['severity'] == 'HIGH' else self.soft_yellow_font
            
            for col_idx, cell in enumerate(ws[row_num], 1):
                cell.font = self.regular_font
                cell.border = self.thin_border
                
                key = col_keys[col_idx - 1]
                if key in ['delta_abs', 'delta_pct']:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    if key == 'delta_pct':
                        cell.value = cell.value / 100.0 if cell.value else 0.0
                        cell.number_format = '0.00%'
                    else:
                        cell.number_format = '#,##0'
                elif key in ['stt', 'contractor']:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                
                # Highlight last cell (severity/comment/value)
                if col_idx == len(col_keys):
                    cell.fill = fill
                    cell.font = font
                    
        self.auto_fit_columns(ws)

    def build_arising_sheet(self, wb, canonical_b):
        """
        Creates sheet listing all contractor arising/extra items (Section B).
        """
        ws = wb.create_sheet(title="Hạng mục phát sinh")
        
        # Title
        ws.cell(row=2, column=1, value="DANH SÁCH CÁC HẠNG MỤC PHÁT SINH NGOÀI KLMT").font = self.title_font
        ws.row_dimensions[2].height = 25
        ws.append([]) # row 3
        
        # Headers
        headers = ["Nhà thầu", "Hệ thống", "STT", "Mã hiệu", "Mô tả công việc", "Đơn vị", "Khối lượng chào", "Đơn giá chào", "Thành tiền chào"]
        ws.append(headers) # row 4
        self.style_row(ws, 4, self.header_font, Alignment(horizontal="center", vertical="center"), self.navy_fill, self.thin_border)
        ws.row_dimensions[4].height = 22
        
        # Write rows
        for c_item in canonical_b:
            if c_item['row_type'] != 'priced':
                continue
                
            for contractor, bid in c_item['bids'].items():
                qty_bid = bid.get('qty_bid')
                price = bid.get('price')
                total = bid.get('total')
                
                # If they bid it as arising
                if qty_bid or price or total:
                    tot_val = total if total else (qty_bid * price if qty_bid and price else 0)
                    row_vals = [
                        contractor,
                        c_item['system'],
                        c_item['stt'],
                        c_item['code'],
                        c_item['description'],
                        c_item['unit'],
                        qty_bid,
                        price,
                        tot_val
                    ]
                    ws.append(row_vals)
                    row_num = ws.max_row
                    ws.row_dimensions[row_num].height = 20
                    
                    for col_idx, cell in enumerate(ws[row_num], 1):
                        cell.font = self.regular_font
                        cell.border = self.thin_border
                        
                        if col_idx in [1, 3, 4, 6]:
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                        elif col_idx in [2, 5]:
                            cell.alignment = Alignment(horizontal="left", vertical="center")
                        elif col_idx == 7:
                            cell.alignment = Alignment(horizontal="right", vertical="center")
                            cell.number_format = '#,##0.00'
                        elif col_idx in [8, 9]:
                            cell.alignment = Alignment(horizontal="right", vertical="center")
                            cell.number_format = '#,##0'
                            
        self.auto_fit_columns(ws)

    def build_detailed_comparison_sheet(self, wb, contractors, canonical_a):
        """
        Creates side-by-side comparative matrix similar to target V2 Excel sheet.
        """
        ws = wb.create_sheet(title="Ma trận so sánh chi tiết")
        
        # Title
        ws.cell(row=2, column=1, value="MA TRẬN SO SÁNH GIÁ CHÀO THẦU CHI TIẾT").font = self.title_font
        ws.row_dimensions[2].height = 25
        
        # Headers Row 1
        h1 = ["Thông tin mời thầu", "", "", "", ""]
        for c in contractors:
            h1.extend([f"Nhà thầu: {c}", "", "", "", "", "", ""])
            
        # Headers Row 2
        h2 = ["STT", "Mã hiệu", "Diễn giải", "ĐVT", "KL Mời thầu"]
        for c in contractors:
            h2.extend(["KL Chào", "VL Chính", "VL Phụ", "Nhân công & Máy", "CPQL & LN", "Đơn giá tổng", "Thành tiền thầu"])
            
        ws.append([]) # row 3
        ws.append(h1) # row 4
        ws.append(h2) # row 5
        
        # Formatting headers
        ws.row_dimensions[4].height = 20
        ws.row_dimensions[5].height = 20
        
        # Merge invited header
        ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=5)
        
        col_idx = 6
        for c in contractors:
            ws.merge_cells(start_row=4, start_column=col_idx, end_row=4, end_column=col_idx + 6)
            col_idx += 7
            
        self.style_row(ws, 4, self.header_font, Alignment(horizontal="center", vertical="center"), self.navy_fill, self.thin_border)
        self.style_row(ws, 5, self.header_font, Alignment(horizontal="center", vertical="center", wrap_text=True), self.navy_fill, self.thin_border)
        
        # Write rows
        row_num = 6
        for c_item in canonical_a:
            row_type = c_item['row_type']
            is_header = (row_type == 'header')
            fill = self.light_gray_fill if is_header else None
            font = self.bold_font if is_header else self.regular_font
            
            # 1. Write STT
            cell = ws.cell(row=row_num, column=1, value=c_item['stt'])
            cell.font = font
            cell.border = self.thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if fill: cell.fill = fill
            
            # 2. Write Code
            cell = ws.cell(row=row_num, column=2, value=c_item['code'])
            cell.font = font
            cell.border = self.thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if fill: cell.fill = fill
            
            # 3. Write Desc
            cell = ws.cell(row=row_num, column=3, value=c_item['description'])
            cell.font = font
            cell.border = self.thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center")
            if fill: cell.fill = fill
            
            # 4. Write Unit
            cell = ws.cell(row=row_num, column=4, value=c_item['unit'])
            cell.font = font
            cell.border = self.thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if fill: cell.fill = fill
            
            # 5. Write Qty Invited
            cell = ws.cell(row=row_num, column=5, value=c_item['qty_invited'])
            cell.font = font
            cell.border = self.thin_border
            if cell.value is not None:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if fill: cell.fill = fill
            
            # Contractors
            col_offset = 6
            for c in contractors:
                bid = c_item['bids'].get(c)
                if bid:
                    # Qty Bid
                    cell = ws.cell(row=row_num, column=col_offset, value=bid.get('qty_bid'))
                    cell.font = font
                    cell.border = self.thin_border
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    if fill: cell.fill = fill
                    if cell.value is not None: cell.number_format = '#,##0.00'
                    
                    # VL Chinh
                    cell = ws.cell(row=row_num, column=col_offset + 1, value=bid.get('p_vl_chinh'))
                    cell.font = font
                    cell.border = self.thin_border
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    if fill: cell.fill = fill
                    if cell.value is not None: cell.number_format = '#,##0'
                    
                    # VL Phu
                    cell = ws.cell(row=row_num, column=col_offset + 2, value=bid.get('p_vl_phu'))
                    cell.font = font
                    cell.border = self.thin_border
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    if fill: cell.fill = fill
                    if cell.value is not None: cell.number_format = '#,##0'
                    
                    # NC May
                    cell = ws.cell(row=row_num, column=col_offset + 3, value=bid.get('p_nc_may'))
                    cell.font = font
                    cell.border = self.thin_border
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    if fill: cell.fill = fill
                    if cell.value is not None: cell.number_format = '#,##0'
                    
                    # CPQL & LN
                    qp = 0.0
                    if bid.get('p_quanly'): qp += bid.get('p_quanly')
                    if bid.get('p_loinhuan'): qp += bid.get('p_loinhuan')
                    cell = ws.cell(row=row_num, column=col_offset + 4, value=qp if qp > 0 else None)
                    cell.font = font
                    cell.border = self.thin_border
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    if fill: cell.fill = fill
                    if cell.value is not None: cell.number_format = '#,##0'
                    
                    # Price
                    cell = ws.cell(row=row_num, column=col_offset + 5, value=bid.get('price'))
                    cell.font = font
                    cell.border = self.thin_border
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    if fill: cell.fill = fill
                    if cell.value is not None: cell.number_format = '#,##0'
                    
                    # Total
                    cell = ws.cell(row=row_num, column=col_offset + 6, value=bid.get('total'))
                    cell.font = font
                    cell.border = self.thin_border
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    if fill: cell.fill = fill
                    if cell.value is not None: cell.number_format = '#,##0'
                else:
                    # Write N/A
                    for j in range(7):
                        cell = ws.cell(row=row_num, column=col_offset + j, value="N/A" if j == 0 else None)
                        cell.font = font
                        cell.border = self.thin_border
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        if fill: cell.fill = fill
                
                col_offset += 7
                
            ws.row_dimensions[row_num].height = 20
            row_num += 1
            
        self.auto_fit_columns(ws)

    def generate_report(self, comparison_results, file_path=None):
        """
        Builds the Excel report. Saves to path if specified, otherwise returns BytesIO stream.
        """
        wb = Workbook()
        
        contractors = comparison_results['contractors']
        summary_data = comparison_results['summary_data']
        comments = comparison_results['comments']
        flags = comparison_results['flags']
        canonical_a = comparison_results['canonical_a']
        canonical_b = comparison_results['canonical_b']
        
        # 1. Build Summary sheet
        self.build_summary_sheet(wb, contractors, summary_data, comments)
        
        # 2. Build Missing Items sheet
        self.build_flags_sheet(
            wb, 
            sheet_name="Hạng mục thiếu", 
            flags=flags, 
            flag_type_filter=['ITEM_MISSING'],
            title="DANH SÁCH CÁC HẠNG MỤC THIẾU / BỎ SÓT CHÀO GIÁ",
            headers=["Nhà thầu", "Hệ thống", "STT", "Diễn giải hạng mục", "Nội dung cảnh báo"],
            col_keys=["contractor", "system", "stt", "description", "message"]
        )
        
        # 3. Build Arising Items sheet
        self.build_arising_sheet(wb, canonical_b)
        
        # 4. Build Quantity Deviations sheet
        self.build_flags_sheet(
            wb, 
            sheet_name="Sai khác khối lượng", 
            flags=flags, 
            flag_type_filter=['QTY_OVER', 'QTY_UNDER'],
            title="DANH SÁCH CÁC HẠNG MỤC SAI KHÁC KHỐI LƯỢNG CHÀO THẦU",
            headers=["Nhà thầu", "Hệ thống", "STT", "Diễn giải hạng mục", "Nội dung cảnh báo", "Độ lệch (abs)", "Độ lệch (%)"],
            col_keys=["contractor", "system", "stt", "description", "message", "delta_abs", "delta_pct"]
        )
        
        # 5. Build Unit Price Deviations sheet
        self.build_flags_sheet(
            wb, 
            sheet_name="Sai khác đơn giá", 
            flags=flags, 
            flag_type_filter=['PRICE_HIGH', 'PRICE_LOW'],
            title="DANH SÁCH CÁC HẠNG MỤC CHÊNH LỆCH ĐƠN GIÁ LỚN (>25% so với Median)",
            headers=["Nhà thầu", "Hệ thống", "STT", "Diễn giải hạng mục", "Nội dung cảnh báo", "Lượng chênh (abs)", "Lượng chênh (%)"],
            col_keys=["contractor", "system", "stt", "description", "message", "delta_abs", "delta_pct"]
        )
        
        # 6. Build Detailed side-by-side comparison sheet
        self.build_detailed_comparison_sheet(wb, contractors, canonical_a)
        
        # Save or return
        if file_path:
            wb.save(file_path)
            return file_path
        else:
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer
