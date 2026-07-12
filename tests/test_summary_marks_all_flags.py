"""Bảng tổng hợp phải đánh dấu MỌI ô có sai lệch bị gắn cờ (khớp với danh sách
cảnh báo trên web), không chỉ ô đơn giá/thành tiền — ví dụ lệch khối lượng,
thương hiệu, xuất xứ giữa các nhà thầu.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from core.config import EnterpriseConfig
from core.pipeline import compare_bidder_files
from core.reporter import export_consolidated_summary

SHEET = "BOQ"


def _cfg() -> EnterpriseConfig:
    cfg = EnterpriseConfig()
    cfg.enable_semantic_matching = False
    cfg.enable_reranker = False
    return cfg


def _bidder(path: Path, kl: float, price: float, brand: str, origin: str) -> None:
    wb = Workbook(); ws = wb.active; ws.title = SHEET
    ws.append(["STT", "Mã hiệu", "Tên hạng mục", "ĐVT", "Khối lượng nhà thầu chào",
               "Đơn giá tổng hợp", "Thành tiền", "Thương hiệu", "Xuất xứ"])
    ws.append(["1", "M-01", "Tủ điện tổng", "Cái", kl, price, kl * price, brand, origin])
    wb.save(path)


def _leaf_columns(ws) -> dict[str, list[int]]:
    result: dict[str, list[int]] = {}
    for cell in ws[4]:
        if cell.value:
            result.setdefault(str(cell.value).replace("\n", " ").strip(), []).append(cell.column)
    return result


def _data_row(ws, cols: list[int]) -> int:
    for r in range(5, ws.max_row + 1):
        if any(ws.cell(r, c).value not in (None, "") for c in cols):
            return r
    raise AssertionError("Không tìm thấy dòng dữ liệu")


def _marked(cell) -> bool:
    has_fill = cell.fill is not None and cell.fill.patternType and cell.fill.fgColor.rgb not in (None, "00000000")
    return bool(cell.comment is not None and has_fill)


def _build(tmp_path: Path):
    data = {
        "NT A": (1, 1_000_000, "Schneider", "Pháp"),
        "NT B": (1, 1_000_000, "Schneider", "Pháp"),
        "NT C": (5, 1_000_000, "ABB", "Đức"),   # lệch khối lượng + thương hiệu + xuất xứ
    }
    files = []
    for name, (kl, price, brand, origin) in data.items():
        p = tmp_path / f"{name}.xlsx"
        _bidder(p, kl, price, brand, origin)
        files.append((name, p))
    result = compare_bidder_files(files, config=_cfg())
    out = tmp_path / "tong_hop.xlsx"
    export_consolidated_summary(result, out)
    return load_workbook(out)[SHEET]


def test_quantity_deviation_cell_is_marked(tmp_path: Path):
    ws = _build(tmp_path)
    cols = _leaf_columns(ws)
    kl_cols = sorted(cols["KL NT chào"])          # 3 nhà thầu -> 3 cột KL
    ntc_kl = kl_cols[2]                            # NT C là block thứ 3
    row = _data_row(ws, kl_cols)
    assert _marked(ws.cell(row, ntc_kl)), "Ô khối lượng lệch của NT C phải được tô màu + chú thích"


def test_brand_and_origin_deviation_cells_are_marked(tmp_path: Path):
    ws = _build(tmp_path)
    cols = _leaf_columns(ws)
    brand_cols = sorted(cols["Thương hiệu"])
    origin_cols = sorted(cols["Xuất xứ"])
    row = _data_row(ws, brand_cols)
    assert _marked(ws.cell(row, brand_cols[2])), "Ô thương hiệu khác của NT C phải được đánh dấu"
    assert _marked(ws.cell(row, origin_cols[2])), "Ô xuất xứ khác của NT C phải được đánh dấu"


def test_equal_price_cell_not_marked(tmp_path: Path):
    ws = _build(tmp_path)
    cols = _leaf_columns(ws)
    dg_cols = sorted(cols["ĐG tổng hợp"])
    row = _data_row(ws, dg_cols)
    # Đơn giá bằng nhau -> không đánh dấu (không dương tính giả).
    for c in dg_cols:
        assert ws.cell(row, c).comment is None
