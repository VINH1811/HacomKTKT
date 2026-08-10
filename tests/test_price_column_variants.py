"""Cột đơn giá viết tắt hoặc viết sai vẫn phải nhận diện được; và khi thật sự
không nhận ra thì phải CẢNH BÁO chứ không im lặng bỏ qua.

Thất bại thầm lặng là trường hợp nguy hiểm nhất: người dùng thấy "không có lỗi"
và tưởng hồ sơ sạch, trong khi thực chất hệ thống không kiểm tra được gì.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from core.excel_reader import _flatten_header, load_workbook_items, map_columns
from core.models import DocumentRole

RECOGNISED = [
    "Đơn giá tổng hợp", "ĐG tổng hợp", "ĐGTH", "ĐG", "Đơn giá", "Đơn giá TH", "Giá",
    "Đơn gía tổng hợp",          # sai dấu
    "Don gia tong hop",          # không dấu
    "DON GIA",                   # viết hoa
    "Đơn  giá  tổng  hợp",       # thừa khoảng trắng
    "Đơn giá chào", "Đơn giá dự thầu", "Đơn giá (VNĐ)", "Đơn giá sau thuế",
    "Giá chào", "Đơn giá tổng hợp (chưa VAT)",
]


def _map_price_column(label: str):
    top = ["STT", "Hạng mục", "ĐVT", "KL nhà thầu chào", label, "Thành tiền"]
    fixed, _ = map_columns(_flatten_header([top], len(top)), DocumentRole.HSDT)
    return fixed.get(4)


def test_price_column_spelling_variants_are_recognised():
    failed = [v for v in RECOGNISED if _map_price_column(v) != "unit_price_total"]
    assert not failed, f"Không nhận diện được cột đơn giá viết dạng: {failed}"


def test_price_component_columns_not_swallowed():
    # Các cột thành phần giá không được nhận nhầm thành đơn giá tổng hợp.
    for label, expected in (("Đơn giá VL chính", "price_main"),
                            ("Đơn giá NC&M", "price_labor"),
                            ("CP quản lý", "price_management"),
                            ("Lợi nhuận", "price_profit")):
        assert _map_price_column(label) == expected, f"{label!r} bị nhận nhầm"


def _workbook(path: Path, headers: list[str]) -> Path:
    wb = Workbook(); ws = wb.active; ws.title = "S1"
    ws.append(headers)
    ws.append(["1", "DI-M9102", "Đầu báo khói địa chỉ", "cái", 71, 770_440, 54_701_240])
    wb.save(path)
    return path


def test_unrecognised_price_column_warns_instead_of_silence(tmp_path: Path):
    path = _workbook(tmp_path / "la.xlsx",
                     ["STT", "Mã hiệu", "Tên hạng mục", "ĐVT", "Số chào", "Tiền hàng", "Cộng"])
    data = load_workbook_items(path, DocumentRole.HSDT, bidder="NT")
    warned = [w for w in data.warnings if "không nhận ra cột đơn giá" in w]
    assert warned, "Không nhận ra cột đơn giá thì phải cảnh báo, không được im lặng"
    # Cảnh báo phải kèm tiêu đề đọc được để người dùng đối chiếu.
    assert "Tiêu đề đọc được" in warned[0]


def test_recognised_columns_do_not_warn(tmp_path: Path):
    path = _workbook(tmp_path / "chuan.xlsx",
                     ["STT", "Mã hiệu", "Tên hạng mục", "ĐVT",
                      "KL nhà thầu chào", "Đơn giá tổng hợp", "Thành tiền"])
    data = load_workbook_items(path, DocumentRole.HSDT, bidder="NT")
    assert not [w for w in data.warnings if "không nhận ra cột" in w]


def test_reference_appendix_never_warns_about_price_column(tmp_path: Path):
    # Bảng khối lượng mời thầu vốn không có cột đơn giá — thiếu là đúng.
    path = _workbook(tmp_path / "pl01.xlsx",
                     ["STT", "Mã hiệu", "Tên hạng mục", "ĐVT", "Khối lượng mời thầu", "Ghi chú", "x"])
    data = load_workbook_items(path, DocumentRole.HSMT, bidder="")
    assert not [w for w in data.warnings if "không nhận ra cột" in w]
