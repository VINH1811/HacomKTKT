"""Giảm nhiễu: cảnh báo chỉ nên xuất hiện khi người chấm cần làm gì đó.

Đo trên hồ sơ thật (gói NOXH01, 1.713 dòng kết quả):
- 1.348 ghi chú "Mã hiệu: hồ sơ đối chiếu để trống" — phụ lục KHÔNG dùng mã hiệu
  ở bất kỳ dòng nào, nhà thầu tự điền vào. Báo trên từng dòng là vô nghĩa và
  làm chìm các cảnh báo thật;
- 465 ghi chú "Mã hiệu trùng nhưng mô tả khác nhau (0-0)" — "0-0" là ô điền cho
  có, không phải mã hiệu.
"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from core.comparison import build_bidder_rows
from core.excel_reader import _is_real_code, load_workbook_items
from core.models import DocumentRole

HDR_REF = ["STT", "Mã hiệu", "Tên hạng mục", "ĐVT", "Khối lượng mời thầu"]
HDR_BID = ["STT", "Mã hiệu", "Tên hạng mục", "ĐVT", "KL nhà thầu chào",
           "Đơn giá tổng hợp", "Thành tiền"]


@pytest.mark.parametrize("code", ["DI-M9102", "EI45-SM", "TDC", "403816", "FIG-015"])
def test_real_codes_are_kept(code: str):
    assert _is_real_code(code) is True


@pytest.mark.parametrize("code", ["0-0", "0", "00", "-", "--", "0.0", "", " ", "A", "1"])
def test_placeholder_codes_are_ignored(code: str):
    assert _is_real_code(code) is False


def _book(path: Path, header: list[str], rows: list[list]) -> Path:
    wb = Workbook(); ws = wb.active; ws.title = "PCCC"
    ws.append(header)
    for row in rows:
        ws.append(row)
    wb.save(path)
    return path


def test_placeholder_code_does_not_raise_duplicate_warning(tmp_path: Path):
    path = _book(tmp_path / "nt.xlsx", HDR_BID, [
        ["1", "0-0", "Đầu báo khói địa chỉ", "cái", 10, 500_000, 5_000_000],
        ["2", "0-0", "Van góc chữa cháy DN50", "cái", 4, 900_000, 3_600_000],
        ["3", "0-0", "Ống thép tráng kẽm DN100", "m", 50, 280_000, 14_000_000],
    ])
    data = load_workbook_items(path, DocumentRole.HSDT, bidder="NT")
    assert not [w for w in data.warnings if "Mã hiệu trùng" in w]


def test_real_duplicate_code_still_warns(tmp_path: Path):
    path = _book(tmp_path / "nt.xlsx", HDR_BID, [
        ["1", "DI-M9102", "Đầu báo khói địa chỉ", "cái", 10, 500_000, 5_000_000],
        ["2", "DI-M9102", "Van góc chữa cháy DN50", "cái", 4, 900_000, 3_600_000],
    ])
    data = load_workbook_items(path, DocumentRole.HSDT, bidder="NT")
    assert [w for w in data.warnings if "Mã hiệu trùng" in w]


def _rows(tmp_path: Path, ref_code: str, bid_code: str):
    ref = _book(tmp_path / "pl01.xlsx", HDR_REF,
                [["1", ref_code, "Đầu báo khói địa chỉ", "cái", 10],
                 ["2", ref_code, "Van góc chữa cháy DN50", "cái", 4]])
    bid = _book(tmp_path / "nt.xlsx", HDR_BID,
                [["1", bid_code, "Đầu báo khói địa chỉ", "cái", 10, 500_000, 5_000_000],
                 ["2", bid_code, "Van góc chữa cháy DN50", "cái", 4, 900_000, 3_600_000]])
    from core.config import EnterpriseConfig
    from core.matcher import match_items_cached
    cfg = EnterpriseConfig(); cfg.enable_semantic_matching = False; cfg.enable_reranker = False
    a = load_workbook_items(ref, DocumentRole.HSMT, bidder="")
    b = load_workbook_items(bid, DocumentRole.HSDT, bidder="NT")
    return build_bidder_rows(a.items, b.items, "NT", match_items_cached(a, b, cfg), cfg)


def test_no_note_when_the_tender_never_gives_a_code(tmp_path: Path):
    # Phụ lục bỏ trống mã hiệu, nhà thầu điền thêm -> không phải phát hiện.
    rows = _rows(tmp_path, "", "DI-M9102")
    notes = [x.message for r in rows for x in r.differences if "Mã hiệu" in str(x.field)]
    assert not notes, f"Không được ghi chú khi phụ lục để trống: {notes}"


def test_bidder_leaving_a_required_code_blank_is_still_reported(tmp_path: Path):
    # Ngược lại: phụ lục có mã, nhà thầu bỏ trống -> vẫn phải báo.
    rows = _rows(tmp_path, "DI-M9102", "")
    notes = [str(x.message) for r in rows for x in r.differences if "Mã hiệu" in str(x.field)]
    assert notes and any("nhà thầu để trống" in n.lower() for n in notes)
