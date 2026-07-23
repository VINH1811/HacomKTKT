"""Sheet "Tổng hợp 03.1" trong file tổng hợp: đứng đầu file, mỗi sheet hệ thống
một dòng, mỗi nhà thầu hai cột (Theo KLMT / Nhà thầu chào), số liệu là CÔNG THỨC
SỐNG =SUM(...) trỏ về dải dòng mục A của từng sheet chi tiết; dòng B cộng dải
phát sinh; C = A + B."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from core.config import EnterpriseConfig
from core.pipeline import compare_bidder_files
from core.reporter import export_consolidated_summary


def _cfg() -> EnterpriseConfig:
    cfg = EnterpriseConfig()
    cfg.enable_semantic_matching = False
    cfg.enable_reranker = False
    return cfg


def _bidder(path: Path, price: float) -> None:
    wb = Workbook()
    ws1 = wb.active; ws1.title = "1. HT dien"
    ws1.append(["STT", "Mã hiệu", "Tên hạng mục", "ĐVT", "Khối lượng nhà thầu chào",
                "Đơn giá tổng hợp", "Thành tiền"])
    ws1.append(["1", "TD-01", "Tủ điện tổng", "Cái", 2, price, 2 * price])
    ws1.append(["2", "CAP-01", "Cáp Cu/XLPE 3x95", "m", 100, 850_000, 85_000_000])
    ws2 = wb.create_sheet("2. HT CTN")
    ws2.append(["STT", "Mã hiệu", "Tên hạng mục", "ĐVT", "Khối lượng nhà thầu chào",
                "Đơn giá tổng hợp", "Thành tiền"])
    ws2.append(["1", "ONG-01", "Ống PPR D32", "m", 500, 120_000, 60_000_000])
    wb.save(path)


def _build(tmp_path: Path):
    files = []
    for name, price in {"NT A": 500_000_000, "NT B": 520_000_000}.items():
        p = tmp_path / f"{name}.xlsx"
        _bidder(p, price)
        files.append((name, p))
    result = compare_bidder_files(files, config=_cfg())
    out = tmp_path / "tong_hop.xlsx"
    export_consolidated_summary(result, out)
    return load_workbook(out)


def test_rollup_sheet_first_and_structured(tmp_path: Path):
    wb = _build(tmp_path)
    assert wb.sheetnames[0] == "Tổng hợp 03.1"
    ws = wb["Tổng hợp 03.1"]
    assert ws["A4"].value == "A"
    labels = [ws.cell(r, 1).value for r in range(4, 10)]
    assert "B" in labels and "C" in labels


def test_rollup_formulas_reference_detail_sheets(tmp_path: Path):
    wb = _build(tmp_path)
    ws = wb["Tổng hợp 03.1"]
    detail_sheets = [s for s in wb.sheetnames if s != "Tổng hợp 03.1"]
    # Các dòng hạng mục (sau dòng A) phải là công thức =SUM('sheet'!...)
    formulas = []
    for r in range(5, 5 + len(detail_sheets)):
        v = ws.cell(r, 3).value
        assert isinstance(v, str) and v.startswith("=SUM("), f"Ô C{r} phải là công thức SUM, được {v!r}"
        formulas.append(v)
    for sheet in detail_sheets:
        assert any(sheet in v for v in formulas), f"Thiếu công thức trỏ tới sheet {sheet!r}"


def test_rollup_c_equals_a_plus_b(tmp_path: Path):
    wb = _build(tmp_path)
    ws = wb["Tổng hợp 03.1"]
    rows = {ws.cell(r, 1).value: r for r in range(4, ws.max_row + 1)}
    c_row = rows["C"]
    v = ws.cell(c_row, 3).value
    assert isinstance(v, str) and v.startswith("=") and "+" in v


def test_rollup_two_columns_per_bidder(tmp_path: Path):
    wb = _build(tmp_path)
    ws = wb["Tổng hợp 03.1"]
    subheads = [ws.cell(3, c).value for c in range(3, 7)]
    assert subheads == ["Theo KLMT", "Nhà thầu chào", "Theo KLMT", "Nhà thầu chào"]
