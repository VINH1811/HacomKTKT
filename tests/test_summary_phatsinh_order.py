"""Bảng tổng hợp: đẩy khối hạng mục PHÁT SINH xuống cuối, giữ vật tư con dưới cha.

Kịch bản: nhà thầu tự chèn một hạng mục ngoài PL01 vào GIỮA sheet (kèm vật tư
con của nó). Hệ thống phải:
  - Giữ các hạng mục khớp PL01 theo đúng thứ tự chuẩn, vật tư con ngay dưới cha.
  - Dồn nguyên khối hạng mục phát sinh (cha + con) xuống CUỐI trang.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from core.models import (
    ComparedItem,
    ComparisonResult,
    ComparisonSummary,
    DocumentRole,
    ItemRecord,
    MatchKind,
    MatchResult,
    RowType,
)
from core.reporter import export_consolidated_summary

SHEET = "1. HT điện"


def _rec(*, role, sheet, row, stt, name, row_type, ref_qty=None, bid_qty=None, price=None) -> ItemRecord:
    return ItemRecord(
        source_id=f"{role.value}:{sheet}:{row}",
        role=role,
        bidder="PL01" if role is DocumentRole.HSMT else "NT A",
        workbook="wb.xlsx",
        sheet=sheet,
        row_number=row,
        stt=stt,
        item_name=name,
        unit="Cái",
        reference_quantity=ref_qty,
        bid_quantity=bid_qty,
        unit_price_total=price,
        bid_amount=(bid_qty * price) if (bid_qty and price) else None,
        row_type=row_type,
    )


def _matched(cid, pl01_row, stt, name, bidder_row, price) -> ComparedItem:
    ref = _rec(role=DocumentRole.HSMT, sheet=SHEET, row=pl01_row, stt=stt, name=name, row_type=RowType.DETAIL, ref_qty=1)
    cand = _rec(role=DocumentRole.HSDT, sheet=SHEET, row=bidder_row, stt=stt, name=name, row_type=RowType.DETAIL, bid_qty=1, price=price)
    return ComparedItem(cid, "NT A", ref, cand, MatchResult(0, 0, MatchKind.EXACT_NAME, 1.0))


def _extra(cid, bidder_row, stt, name, row_type, price) -> ComparedItem:
    cand = _rec(role=DocumentRole.HSDT, sheet=SHEET, row=bidder_row, stt=stt, name=name, row_type=row_type, bid_qty=1, price=price)
    return ComparedItem(cid, "NT A", None, cand, MatchResult(None, 0, MatchKind.EXTRA, 0.0))


def _diengiai_order(ws) -> list[str]:
    names = []
    for row in range(5, ws.max_row + 1):
        val = ws.cell(row, 3).value  # cột 'Diễn giải' trong khối KLMT
        if val is not None and str(val).strip():
            names.append(str(val).strip())
    return names


def test_phatsinh_block_pushed_to_bottom_with_components(tmp_path: Path):
    rows = [
        _matched("HM1", pl01_row=5, stt="1", name="Tủ điện tổng", bidder_row=5, price=100),
        _extra("C1a", bidder_row=6, stt="", name="Vỏ tủ form 3B", row_type=RowType.COMPONENT, price=10),
        _extra("C1b", bidder_row=7, stt="", name="ACB 4P 1600AT", row_type=RowType.COMPONENT, price=20),
        # Hạng mục nhà thầu TỰ CHÈN giữa sheet (không có trong PL01) + vật tư con.
        _extra("PS", bidder_row=8, stt="1B", name="Bổ sung ngoài thầu", row_type=RowType.DETAIL, price=99),
        _extra("PSc", bidder_row=9, stt="", name="Phụ kiện bổ sung", row_type=RowType.COMPONENT, price=5),
        _matched("HM2", pl01_row=6, stt="2", name="Cáp điện CU XLPE", bidder_row=10, price=200),
    ]
    summary = ComparisonSummary("PL01", 1, 2, len(rows), 0, 0, 0, 0, 0, 0, 0, 0.0, {}, "")
    result = ComparisonResult(rows=rows, summary=summary, warnings=[], audit={"bidder_sha256": {"NT A": ""}, "thresholds": {}})

    out = tmp_path / "tong_hop.xlsx"
    export_consolidated_summary(result, out)

    ws = load_workbook(out)[SHEET]
    names = _diengiai_order(ws)

    def idx(part):
        return next(i for i, n in enumerate(names) if part in n)

    # Hạng mục khớp giữ thứ tự chuẩn, vật tư con ngay dưới cha.
    assert idx("Tủ điện tổng") < idx("Vỏ tủ form 3B") < idx("ACB 4P 1600AT")
    assert idx("Tủ điện tổng") < idx("Cáp điện CU XLPE")
    # Khối phát sinh (cha + con) nằm SAU toàn bộ hạng mục khớp.
    assert idx("Cáp điện CU XLPE") < idx("Bổ sung ngoài thầu")
    assert idx("Bổ sung ngoài thầu") < idx("Phụ kiện bổ sung")
    # Vật tư con của phát sinh vẫn bám ngay dưới cha phát sinh.
    assert idx("Phụ kiện bổ sung") == idx("Bổ sung ngoài thầu") + 1


def _row_containing(ws, text: str) -> int:
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v is not None and text in str(v):
                return r
    raise AssertionError(f"Không thấy dòng chứa '{text}'")


def test_summary_has_section_b_header_before_phatsinh(tmp_path: Path):
    rows = [
        _matched("HM1", pl01_row=5, stt="1", name="Tủ điện tổng", bidder_row=5, price=100),
        _matched("HM2", pl01_row=6, stt="2", name="Cáp điện CU XLPE", bidder_row=6, price=200),
        _extra("PS", bidder_row=7, stt="1B", name="Bổ sung ngoài thầu", row_type=RowType.DETAIL, price=99),
    ]
    summary = ComparisonSummary("PL01", 1, 2, len(rows), 0, 0, 0, 0, 0, 0, 0, 0.0, {}, "")
    result = ComparisonResult(rows=rows, summary=summary, warnings=[], audit={"bidder_sha256": {"NT A": ""}, "thresholds": {}})

    out = tmp_path / "tong_hop.xlsx"
    export_consolidated_summary(result, out)
    ws = load_workbook(out)[SHEET]

    r_matched = _row_containing(ws, "Cáp điện CU XLPE")
    r_b_header = _row_containing(ws, "PHÁT SINH NGOÀI KLMT")
    r_phatsinh = _row_containing(ws, "Bổ sung ngoài thầu")

    # Tiêu đề mục B nằm SAU hạng mục khớp và NGAY TRƯỚC hạng mục phát sinh.
    assert r_matched < r_b_header < r_phatsinh


def test_matched_items_normalized_to_stt_when_bidder_shifted(tmp_path: Path):
    # Nhà thầu chào LỆCH thứ tự (theo dòng là 3, 1, 2) -> bảng tổng hợp phải xếp
    # lại theo SỐ THỨ TỰ của bản chuẩn: 1, 2, 3.
    rows = [
        _matched("HM3", pl01_row=7, stt="3", name="Đèn LED", bidder_row=5, price=50),
        _matched("HM1", pl01_row=5, stt="1", name="Tủ điện", bidder_row=6, price=100),
        _matched("HM2", pl01_row=6, stt="2", name="Cáp đồng", bidder_row=7, price=200),
    ]
    summary = ComparisonSummary("PL01", 1, 3, len(rows), 0, 0, 0, 0, 0, 0, 0, 0.0, {}, "")
    result = ComparisonResult(rows=rows, summary=summary, warnings=[], audit={"bidder_sha256": {"NT A": ""}, "thresholds": {}})

    out = tmp_path / "tong_hop.xlsx"
    export_consolidated_summary(result, out)
    names = _diengiai_order(load_workbook(out)[SHEET])

    assert names == ["Tủ điện", "Cáp đồng", "Đèn LED"]


PL01_SHEET = "2 - PHAN TU HA THE"
BIDDER_SHEET = "1. HT điện"


def _matched_x(cid, pl01_row, name, bidder, bidder_row, price) -> ComparedItem:
    """Hạng mục khớp: PL01 và nhà thầu đặt tên sheet KHÁC nhau."""
    ref = _rec(role=DocumentRole.HSMT, sheet=PL01_SHEET, row=pl01_row, stt=str(pl01_row), name=name, row_type=RowType.DETAIL, ref_qty=1)
    cand = ItemRecord(
        source_id=f"{bidder}:{BIDDER_SHEET}:{bidder_row}", role=DocumentRole.HSDT, bidder=bidder,
        workbook="b.xlsx", sheet=BIDDER_SHEET, row_number=bidder_row, stt=str(pl01_row), item_name=name,
        unit="Cái", bid_quantity=1, unit_price_total=price, bid_amount=price, row_type=RowType.DETAIL,
    )
    return ComparedItem(cid, bidder, ref, cand, MatchResult(0, 0, MatchKind.EXACT_NAME, 1.0))


def _extra_x(cid, bidder, bidder_row, name, price) -> ComparedItem:
    """Hạng mục phát sinh: chỉ có ở file nhà thầu, mang tên sheet nhà thầu."""
    cand = ItemRecord(
        source_id=f"{bidder}:{BIDDER_SHEET}:{bidder_row}", role=DocumentRole.HSDT, bidder=bidder,
        workbook="b.xlsx", sheet=BIDDER_SHEET, row_number=bidder_row, stt="PS", item_name=name,
        unit="Cái", bid_quantity=1, unit_price_total=price, bid_amount=price, row_type=RowType.DETAIL,
    )
    return ComparedItem(cid, bidder, None, cand, MatchResult(None, 0, MatchKind.EXTRA, 0.0))


def test_summary_uses_bidder_sheet_names_multi_bidder(tmp_path: Path):
    # 2 nhà thầu; PL01 đặt tên sheet KHÁC nhà thầu. Bảng tổng hợp phải giữ theo
    # sheet của FILE NHÀ THẦU (file gốc), không đổi sang tên sheet của PL01; phát
    # sinh của từng nhà thầu nằm ở CUỐI trang đó.
    rows = [
        _matched_x("HM1", pl01_row=5, name="Tủ điện tổng", bidder="NT A", bidder_row=5, price=100),
        _matched_x("HM2", pl01_row=6, name="Cáp điện CU XLPE", bidder="NT A", bidder_row=6, price=200),
        _extra_x("PS_A", bidder="NT A", bidder_row=7, name="Bổ sung của NT A", price=50),
        _matched_x("HM1", pl01_row=5, name="Tủ điện tổng", bidder="NT B", bidder_row=5, price=110),
        _matched_x("HM2", pl01_row=6, name="Cáp điện CU XLPE", bidder="NT B", bidder_row=6, price=210),
        _extra_x("PS_B", bidder="NT B", bidder_row=8, name="Bổ sung của NT B", price=60),
    ]
    summary = ComparisonSummary("PL01", 2, 2, len(rows), 0, 0, 0, 0, 0, 0, 0, 0.0, {}, "")
    result = ComparisonResult(rows=rows, summary=summary, warnings=[],
                              audit={"bidder_sha256": {"NT A": "", "NT B": ""}, "thresholds": {}})

    out = tmp_path / "tong_hop.xlsx"
    export_consolidated_summary(result, out)

    wb = load_workbook(out)
    # Giữ tên sheet của nhà thầu; KHÔNG dùng tên sheet của PL01.
    assert BIDDER_SHEET in wb.sheetnames
    assert PL01_SHEET not in wb.sheetnames

    names = _diengiai_order(wb[BIDDER_SHEET])

    def idx(part):
        return next(i for i, n in enumerate(names) if part in n)

    # Hạng mục khớp giữ thứ tự, rồi phát sinh của cả hai nhà thầu ở cuối.
    assert idx("Tủ điện tổng") < idx("Cáp điện CU XLPE")
    assert idx("Cáp điện CU XLPE") < idx("Bổ sung của NT A")
    assert idx("Cáp điện CU XLPE") < idx("Bổ sung của NT B")
