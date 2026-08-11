"""Bảng khối lượng đọc theo TIÊU ĐỀ CỘT, không theo vị trí cột.

Hai câu hỏi thực tế khi nhận hồ sơ từ nhiều nhà thầu:
- đảo thứ tự cột thì có sai không;
- tên cột viết khác đi thì còn nhận ra không.

Riêng "Tên vật tư"/"Danh mục vật tư" từng làm cả sheet đọc ra 0 hạng mục vì bị
nhận là cột quy cách vật tư thay vì cột tên hạng mục.
"""

from __future__ import annotations

import itertools
from pathlib import Path

import pytest
from openpyxl import Workbook

from core.excel_reader import _flatten_header, load_workbook_items, map_columns
from core.models import DocumentRole

HEADER = ["STT", "Mã hiệu", "Tên hạng mục", "ĐVT", "KL nhà thầu chào",
          "Đơn giá tổng hợp", "Thành tiền", "Thương hiệu", "Xuất xứ"]
ROWS = [
    ["1", "DI-M9102", "Đầu báo khói địa chỉ", "cái", 120, 500_000, 60_000_000, "GST", "China"],
    ["2", "VG-50", "Van góc chữa cháy DN50", "cái", 20, 900_000, 18_000_000, "Shanxi", "China"],
]


def _book(tmp_path: Path, header: list[str], rows: list[list], name: str) -> Path:
    wb = Workbook(); ws = wb.active; ws.title = "PCCC"
    ws.append(header)
    for row in rows:
        ws.append(row)
    path = tmp_path / name
    wb.save(path)
    return path


def _read(path: Path) -> list[tuple]:
    data = load_workbook_items(path, DocumentRole.HSDT, bidder="NT")
    return [
        (i.item_name, i.unit, i.bid_quantity, i.unit_price_total, i.bid_amount, i.brand)
        for i in data.items if i.is_comparable
    ]


def _field(label: str, slot: int, role: DocumentRole = DocumentRole.HSDT):
    header = list(HEADER[:7])
    header[slot] = label
    fixed, _ = map_columns(_flatten_header([header], len(header)), role)
    return fixed.get(slot)


# --------------------------------------------------------------- thứ tự cột

@pytest.mark.parametrize("order", [
    (2, 0, 1, 3, 4, 5, 6, 7, 8),          # tên hạng mục lên đầu
    (8, 7, 6, 5, 4, 3, 2, 1, 0),          # đảo ngược hoàn toàn
    (5, 4, 8, 0, 6, 2, 1, 7, 3),          # xáo trộn
    (3, 2, 5, 6, 0, 1, 4, 8, 7),
])
def test_column_order_does_not_change_the_result(tmp_path: Path, order: tuple[int, ...]):
    base = _read(_book(tmp_path, HEADER, ROWS, "goc.xlsx"))
    shuffled = _book(
        tmp_path,
        [HEADER[i] for i in order],
        [[row[i] for i in order] for row in ROWS],
        f"tron_{''.join(map(str, order))}.xlsx",
    )
    assert _read(shuffled) == base
    assert base and base[0][0] == "Đầu báo khói địa chỉ"


def test_every_permutation_of_the_core_columns(tmp_path: Path):
    # Vét cạn 24 hoán vị của bốn cột quan trọng nhất.
    core = ["Tên hạng mục", "ĐVT", "KL nhà thầu chào", "Đơn giá tổng hợp"]
    values = ["Đầu báo khói địa chỉ", "cái", 120, 500_000]
    for index, order in enumerate(itertools.permutations(range(4))):
        path = _book(tmp_path, [core[i] for i in order],
                     [[values[i] for i in order]], f"hv{index}.xlsx")
        got = _read(path)
        assert got and got[0][:4] == ("Đầu báo khói địa chỉ", "cái", 120.0, 500_000.0), \
            f"Sai với thứ tự {[core[i] for i in order]}"


# ------------------------------------------------------------- tên cột khác

@pytest.mark.parametrize("label", [
    "Tên hạng mục", "Diễn giải", "Nội dung công việc", "Mô tả công việc",
    "Danh mục vật tư", "Tên vật tư", "MÔ TẢ", "Hạng mục", "Description",
])
def test_item_name_variants(label: str):
    assert _field(label, 2) == "item_name"


@pytest.mark.parametrize("label", ["ĐVT", "Đơn vị", "Đơn vị tính", "DVT", "Đ.V.T", "ĐV tính", "Unit"])
def test_unit_variants(label: str):
    assert _field(label, 3) == "unit"


@pytest.mark.parametrize("label", [
    "KL nhà thầu chào", "Khối lượng nhà thầu chào", "KL chào", "NT chào",
    "KLNT", "Số lượng nhà thầu", "Quantity",
])
def test_bid_quantity_variants(label: str):
    assert _field(label, 4) == "bid_quantity"


@pytest.mark.parametrize("label", [
    "Đơn giá tổng hợp", "Đơn giá", "ĐGTH", "ĐG", "Giá", "Đơn gía tổng hợp",
    "Don gia tong hop", "DON GIA", "Đơn giá (VNĐ)", "Unit Price",
])
def test_unit_price_variants(label: str):
    assert _field(label, 5) == "unit_price_total"


@pytest.mark.parametrize("label", [
    "Thành tiền", "Thành tiền NT chào", "THÀNH TIỀN", "Giá trị", "Tổng tiền", "Amount",
])
def test_amount_variants(label: str):
    assert _field(label, 6) == "bid_amount"


def test_material_column_is_borrowed_only_when_no_name_column(tmp_path: Path):
    # Có cả hai thì "Tên hạng mục" vẫn là cột tên, "Mô tả/Quy cách" vẫn là quy cách.
    header = ["STT", "Tên hạng mục", "Mô tả/Quy cách", "ĐVT", "KL nhà thầu chào",
              "Đơn giá tổng hợp", "Thành tiền"]
    fixed, _ = map_columns(_flatten_header([header], len(header)), DocumentRole.HSDT)
    assert fixed.get(1) == "item_name"
    assert fixed.get(2) == "material"


def test_english_header_sheet_is_read(tmp_path: Path):
    header = ["No.", "Code", "Description", "Unit", "Quantity", "Unit Price", "Amount"]
    rows = [["1", "DI-M9102", "Addressable smoke detector", "pcs", 120, 500_000, 60_000_000]]
    got = _read(_book(tmp_path, header, rows, "en.xlsx"))
    assert got and got[0][:5] == ("Addressable smoke detector", "pcs", 120.0, 500_000.0, 60_000_000.0)


def test_vietnamese_wins_over_english_when_both_present():
    # Tiêu đề tiếng Anh chỉ là phương án cuối, không được lấn tiếng Việt.
    header = ["STT", "Mã hiệu", "Tên hạng mục", "ĐVT", "KL nhà thầu chào",
              "Đơn giá tổng hợp", "Amount"]
    fixed, _ = map_columns(_flatten_header([header], len(header)), DocumentRole.HSDT)
    assert fixed.get(2) == "item_name"
    assert fixed.get(5) == "unit_price_total"
    assert fixed.get(6) == "bid_amount"
