"""RFI Tracker: đọc file làm rõ HSCG (cột Ý kiến CĐT = yêu cầu), ghép với file
phản hồi của nhà thầu theo nội dung (chịu chèn/sửa dòng) và chấm trạng thái."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from core.rfi_tracker import (
    STATUS_ANSWERED,
    STATUS_NOT_FOUND,
    STATUS_UNANSWERED,
    export_rfi_report,
    parse_rfi_file,
    track_rfi,
)

HEAD = ["STT", "NỘI DUNG ĐÁNH GIÁ THEO HSYC", "YÊU CẦU", "Nhà thầu kê khai",
        "Ý kiến CĐT", "Nhà thầu trả lời làm rõ"]


def _write(path: Path, rows: list[list], sheet: str = "PL 1") -> None:
    wb = Workbook(); ws = wb.active; ws.title = sheet
    ws.append(["DỰ ÁN: TEST"]); ws.append(["GÓI THẦU: TEST"]); ws.append([])
    ws.append(HEAD)
    for row in rows:
        ws.append(row)
    wb.save(path)


def _request_rows() -> list[list]:
    return [
        ["A", "ĐÁNH GIÁ TÍNH HỢP LỆ", "", "", "", ""],
        ["", "Hiệu lực của HSCG: 60 ngày", "60 ngày", "58 ngày", "Đề nghị làm rõ hiệu lực", ""],
        ["1.3", "Hợp đồng tương tự", "02 hợp đồng", "Kê khai 3 HĐ", "Bổ sung bản sao công chứng", ""],
        ["2", "Doanh thu bình quân", ">= 100 tỷ", "80 tỷ", "Giải trình doanh thu", ""],
    ]


def test_parse_detects_requests_only(tmp_path: Path):
    p = tmp_path / "req.xlsx"
    _write(p, _request_rows())
    items = parse_rfi_file(p)
    assert len(items) == 3  # dòng A (không có Ý kiến CĐT) không tính
    assert items[0].cdt_request.startswith("Đề nghị làm rõ")
    # Dòng không STT thừa kế nội dung gần nhất
    assert "Hiệu lực" in items[0].content


def test_answered_and_unanswered(tmp_path: Path):
    req = tmp_path / "req.xlsx"
    _write(req, _request_rows())
    rows = _request_rows()
    rows[1][5] = "Nhà thầu xác nhận gia hạn hiệu lực 60 ngày"
    rows[2][5] = "Đã bổ sung bản sao công chứng 3 HĐ"
    # dòng doanh thu KHÔNG trả lời
    resp = tmp_path / "resp.xlsx"
    _write(resp, rows)
    res = track_rfi(req, resp, "NT A")
    assert res.count(STATUS_ANSWERED) == 2
    assert res.count(STATUS_UNANSWERED) == 1
    un = next(it for it in res.items if it.status == STATUS_UNANSWERED)
    assert "doanh thu" in un.request.cdt_request.lower()


def test_response_with_inserted_rows_still_matches(tmp_path: Path):
    req = tmp_path / "req.xlsx"
    _write(req, _request_rows())
    rows = _request_rows()
    for row in rows:
        if row[4]:
            row[5] = "Đã phản hồi"
    rows.insert(1, ["", "Ghi chú thêm của nhà thầu", "", "", "", ""])  # chèn dòng
    resp = tmp_path / "resp.xlsx"
    _write(resp, rows)
    res = track_rfi(req, resp, "NT B")
    assert res.count(STATUS_ANSWERED) == 3
    assert res.count(STATUS_NOT_FOUND) == 0


def test_missing_item_in_response_flagged(tmp_path: Path):
    req = tmp_path / "req.xlsx"
    _write(req, _request_rows())
    rows = [r for r in _request_rows() if "Doanh thu" not in r[1]]  # nhà thầu xoá dòng
    for row in rows:
        if row[4]:
            row[5] = "OK"
    resp = tmp_path / "resp.xlsx"
    _write(resp, rows)
    res = track_rfi(req, resp, "NT C")
    assert res.count(STATUS_NOT_FOUND) == 1


def test_export_report(tmp_path: Path):
    req = tmp_path / "req.xlsx"; _write(req, _request_rows())
    resp = tmp_path / "resp.xlsx"
    rows = _request_rows()
    rows[1][5] = "Trả lời"
    _write(resp, rows)
    res = track_rfi(req, resp, "NT D")
    out = tmp_path / "rfi.xlsx"
    export_rfi_report([res], out)
    wb = load_workbook(out)
    assert "Tổng quan RFI" in wb.sheetnames and "RFI NT D" in wb.sheetnames
    ws = wb["RFI NT D"]
    statuses = {ws.cell(r, 6).value for r in range(2, ws.max_row + 1)}
    assert STATUS_ANSWERED in statuses and STATUS_UNANSWERED in statuses
