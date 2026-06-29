"""Phần B — Lỗi parse số / dữ liệu (kịch bản B-01 → B-06 trong kế hoạch test).

Tự động hóa các tình huống nhập liệu khó: định dạng số VN/quốc tế, ô #REF!,
thiếu cột đơn giá, ô STT bị merge, đơn vị viết biến thể.
"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from core.config import EnterpriseConfig
from core.comparison import build_bidder_rows
from core.excel_reader import load_workbook_items
from core.matcher import match_items
from core.models import DocumentRole, ItemRecord, RowType
from core.number_parser import parse_number
from core.text_normalizer import normalize_code, normalize_name, normalize_unit


def _cfg() -> EnterpriseConfig:
    cfg = EnterpriseConfig()
    cfg.enable_semantic_matching = False
    cfg.enable_reranker = False
    return cfg


def _item(name: str, unit: str, role: DocumentRole, qty: float = 10, price: float = 100) -> ItemRecord:
    return ItemRecord(
        source_id="x", role=role, bidder="b", workbook="w.xlsx", sheet="S", row_number=2,
        item_code="M01", item_name=name, unit=unit, reference_quantity=qty, bid_quantity=qty,
        unit_price_total=price, row_type=RowType.DETAIL,
        normalized_code=normalize_code("M01"), normalized_name=normalize_name(name),
        normalized_unit=normalize_unit(unit),
    )


# --- B-01: số kiểu Việt Nam (dấu chấm phân cách nghìn) ------------------------

@pytest.mark.parametrize("raw,expected", [
    ("1.234", 1234),
    ("1.234.567", 1234567),
    ("1.234.567,89", 1234567.89),
    ("12.000", 12000),
])
def test_b01_vietnamese_number_format(raw, expected):
    assert parse_number(raw) == expected


# --- B-02: số kiểu quốc tế (dấu phẩy phân cách nghìn) -------------------------

@pytest.mark.parametrize("raw,expected", [
    ("1,234", 1234),
    ("1,234,567", 1234567),
    ("1,234,567.89", 1234567.89),
])
def test_b02_international_number_format(raw, expected):
    assert parse_number(raw) == expected


def test_b01_b02_decimal_values_preserved():
    # Giá trị thập phân thật vẫn giữ đúng (không bị hiểu thành phân cách nghìn).
    assert parse_number("0,5") == 0.5
    assert parse_number("12,75") == 12.75
    assert parse_number("0.5") == 0.5


# --- B-03: ô thành tiền chứa lỗi #REF! ---------------------------------------

def test_b03_ref_error_is_detected_file_still_read(tmp_path: Path):
    path = tmp_path / "ref_error.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "BOQ"
    ws.append(["STT", "Mã hiệu", "Tên hạng mục", "ĐVT", "Khối lượng", "Đơn giá tổng hợp", "Thành tiền"])
    ws.append(["1", "M-01", "Tủ điện tổng", "Cái", 1, 1_000_000, 1_000_000])
    ws["G2"] = "=#REF!"
    wb.save(path)

    workbook = load_workbook_items(path, DocumentRole.HSDT, bidder="NT A")
    # File vẫn đọc được, và lỗi công thức được phát hiện + cảnh báo.
    assert workbook.items, "Phải vẫn đọc được hạng mục"
    assert any(issue["cell"] == "G2" for issue in workbook.formula_issues)
    assert any("#REF" in w.upper() for w in workbook.warnings)


# --- B-04: thiếu toàn bộ cột đơn giá -----------------------------------------

def test_b04_missing_price_column_is_handled(tmp_path: Path):
    path = tmp_path / "no_price.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "BOQ"
    # Không có cột đơn giá / thành tiền — chỉ có khối lượng.
    ws.append(["STT", "Mã hiệu", "Tên hạng mục", "ĐVT", "Khối lượng"])
    ws.append(["1", "M-01", "Tủ điện tổng", "Cái", 5])
    ws.append(["2", "M-02", "Cáp đồng XLPE", "m", 100])
    wb.save(path)

    # Không crash, đọc được hạng mục, và đơn giá để trống (None).
    workbook = load_workbook_items(path, DocumentRole.HSDT, bidder="NT A")
    comparable = [it for it in workbook.items if it.is_comparable]
    assert len(comparable) == 2
    assert all(it.unit_price_total is None for it in comparable)


# --- B-05: ô STT bị merge với dòng dưới --------------------------------------

def test_b05_merged_stt_cell_does_not_crash(tmp_path: Path):
    path = tmp_path / "merged.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "BOQ"
    ws.append(["STT", "Mã hiệu", "Tên hạng mục", "ĐVT", "Khối lượng", "Đơn giá tổng hợp", "Thành tiền"])
    ws.append(["I.1", "M-01", "Tủ điện tổng", "Cái", 1, 1_000_000, 1_000_000])
    ws.append([None, "M-02", "Cáp đồng XLPE", "m", 100, 50_000, 5_000_000])
    # Merge ô STT của 2 dòng dữ liệu (A2:A3).
    ws.merge_cells("A2:A3")
    wb.save(path)

    # Không được crash vì merged cell; đọc được ít nhất một dòng dữ liệu.
    workbook = load_workbook_items(path, DocumentRole.HSDT, bidder="NT A")
    comparable = [it for it in workbook.items if it.is_comparable]
    assert len(comparable) >= 1
    names = {it.item_name for it in comparable}
    assert "Cáp đồng XLPE" in names or "Tủ điện tổng" in names


# --- B-06: đơn vị viết biến thể "mét" vs "m" ---------------------------------

def test_b06_unit_variant_met_equals_m_no_mismatch():
    ref = [_item("Cáp đồng XLPE 4x10", "m", DocumentRole.HSMT)]
    cand = [_item("Cáp đồng XLPE 4x10", "mét", DocumentRole.HSDT)]
    cfg = _cfg()
    matches = match_items(ref, cand, cfg)
    rows = build_bidder_rows(ref, cand, "NT A", matches, cfg, reference_is_boq=True)

    unit_diffs = [d for r in rows for d in r.differences if "đơn vị" in str(d.field).lower()]
    assert unit_diffs == [], f"'mét' và 'm' phải coi là cùng đơn vị, không flag: {unit_diffs}"


@pytest.mark.parametrize("u1,u2", [("m", "mét"), ("m2", "m²"), ("Bộ", "bộ"), ("m³", "mét khối")])
def test_b06_more_unit_variants_are_equivalent(u1, u2):
    ref = [_item("Hạng mục X", u1, DocumentRole.HSMT)]
    cand = [_item("Hạng mục X", u2, DocumentRole.HSDT)]
    cfg = _cfg()
    matches = match_items(ref, cand, cfg)
    rows = build_bidder_rows(ref, cand, "NT A", matches, cfg, reference_is_boq=True)
    unit_diffs = [d for r in rows for d in r.differences if "đơn vị" in str(d.field).lower()]
    assert unit_diffs == []


def test_b06_genuinely_different_units_still_flagged():
    # Đơn vị thật sự khác nhau (Bộ vs m) vẫn phải bị cảnh báo.
    ref = [_item("Hạng mục Y", "Bộ", DocumentRole.HSMT)]
    cand = [_item("Hạng mục Y", "m", DocumentRole.HSDT)]
    cfg = _cfg()
    matches = match_items(ref, cand, cfg)
    rows = build_bidder_rows(ref, cand, "NT A", matches, cfg, reference_is_boq=True)
    unit_diffs = [d for r in rows for d in r.differences if "đơn vị" in str(d.field).lower()]
    assert unit_diffs, "Đơn vị khác nhau thật sự phải được cảnh báo"
