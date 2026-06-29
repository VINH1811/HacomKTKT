"""Cảnh báo khi nghi đặt nhầm vị trí file (không tự sửa, chỉ cảnh báo).

Tín hiệu: file mời thầu (HSMT/PL01) thường chưa có đơn giá; hồ sơ dự thầu thì
đã điền đơn giá. Khi tín hiệu lệch với ô tải lên -> cảnh báo để người dùng soi lại.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from core.comparison import detail_price_fill_ratio, misplacement_warnings
from core.config import EnterpriseConfig
from core.pipeline import compare_tender_files
from core.tender_package import compare_appendices_with_bidders


def _cfg() -> EnterpriseConfig:
    cfg = EnterpriseConfig()
    cfg.enable_semantic_matching = False
    cfg.enable_reranker = False
    return cfg


def _invitation(path: Path) -> None:
    """Bảng mời thầu/KLMT: có khối lượng, KHÔNG có đơn giá."""
    wb = Workbook()
    ws = wb.active
    ws.title = "KLMT"
    ws.append(["STT", "Mã hiệu", "Tên hạng mục", "ĐVT", "Khối lượng mời thầu"])
    ws.append(["1", "M-01", "Tủ điện tổng", "Cái", 2])
    ws.append(["2", "M-02", "Cáp đồng XLPE", "m", 100])
    wb.save(path)


def _bid(path: Path, name: str = "BOQ") -> None:
    """Hồ sơ dự thầu: đã điền đơn giá + thành tiền."""
    wb = Workbook()
    ws = wb.active
    ws.title = name
    ws.append(["STT", "Mã hiệu", "Tên hạng mục", "ĐVT", "Khối lượng mời thầu",
               "Khối lượng nhà thầu chào", "Đơn giá tổng hợp", "Thành tiền"])
    ws.append(["1", "M-01", "Tủ điện tổng", "Cái", 2, 2, 1_000_000, 2_000_000])
    ws.append(["2", "M-02", "Cáp đồng XLPE", "m", 100, 100, 50_000, 5_000_000])
    wb.save(path)


# --- Hàm đo tỷ lệ điền giá -----------------------------------------------------

def test_fill_ratio_distinguishes_invitation_from_bid(tmp_path: Path):
    from core.excel_reader import load_workbook_items
    from core.models import DocumentRole

    inv = tmp_path / "inv.xlsx"; _invitation(inv)
    bid = tmp_path / "bid.xlsx"; _bid(bid)
    inv_wb = load_workbook_items(inv, DocumentRole.HSMT, bidder="KLMT")
    bid_wb = load_workbook_items(bid, DocumentRole.HSDT, bidder="NT A")

    assert detail_price_fill_ratio(inv_wb) <= 0.05    # mời thầu: ~0
    assert detail_price_fill_ratio(bid_wb) >= 0.5      # dự thầu: cao


# --- Package: hồ sơ dự thầu bị bỏ vào ô Phụ lục 01 -----------------------------

def test_package_warns_when_bid_uploaded_as_pl1(tmp_path: Path):
    # PL01 thực ra là một hồ sơ dự thầu (đã chào giá) -> phải cảnh báo.
    fake_pl1 = tmp_path / "fake_pl1.xlsx"; _bid(fake_pl1)
    bidder = tmp_path / "bidder.xlsx"; _bid(bidder)

    out = compare_appendices_with_bidders(
        bidder_files=[("NT A", bidder)], output_dir=tmp_path / "out",
        pl1_path=fake_pl1, pl2_path=None, config=_cfg(),
    )
    assert any("phụ lục 01" in w.lower() and "dự thầu" in w.lower() for w in out.result.warnings)


def test_package_warns_when_invitation_in_bidder_slot(tmp_path: Path):
    # PL01 đúng, nhưng "nhà thầu" lại là một bảng mời thầu chưa có giá -> cảnh báo.
    pl1 = tmp_path / "pl1.xlsx"; _invitation(pl1)
    fake_bidder = tmp_path / "fake_bidder.xlsx"; _invitation(fake_bidder)

    out = compare_appendices_with_bidders(
        bidder_files=[("NT A", fake_bidder)], output_dir=tmp_path / "out",
        pl1_path=pl1, pl2_path=None, config=_cfg(),
    )
    assert any("nhà thầu" in w.lower() and ("chưa có đơn giá" in w.lower() or "mời thầu" in w.lower())
               for w in out.result.warnings)


def test_package_no_warning_when_files_placed_correctly(tmp_path: Path):
    pl1 = tmp_path / "pl1.xlsx"; _invitation(pl1)
    bidder = tmp_path / "bidder.xlsx"; _bid(bidder)

    out = compare_appendices_with_bidders(
        bidder_files=[("NT A", bidder)], output_dir=tmp_path / "out",
        pl1_path=pl1, pl2_path=None, config=_cfg(),
    )
    assert not any("kiểm tra lại vị trí tải lên" in w.lower() for w in out.result.warnings)


# --- Tender: HSMT ↔ HSDT -------------------------------------------------------

def test_tender_warns_when_hsmt_is_actually_a_bid(tmp_path: Path):
    fake_hsmt = tmp_path / "fake_hsmt.xlsx"; _bid(fake_hsmt)
    bidder = tmp_path / "bidder.xlsx"; _bid(bidder)

    result = compare_tender_files(fake_hsmt, [("NT A", bidder)], config=_cfg())
    assert any("hsmt" in w.lower() and "dự thầu" in w.lower() for w in result.warnings)
