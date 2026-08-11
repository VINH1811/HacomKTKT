"""Hệ thống phải dùng được cho MỌI chuyên ngành, không riêng cơ điện/PCCC.

Từ vựng phân biệt chủng loại vật tư được xây từ hồ sơ cơ điện và phòng cháy, nên
dễ chỉ đúng với ngành đó. Hai chỗ từng chặn oan ở hồ sơ xây dựng/hạ tầng:

- "ống nhựa uPVC" và "ống nhựa PVC" bị coi là hai vật liệu khác nhau;
- "ống HDPE" và "ống PE" cũng vậy.

Lưu ý về cách sửa: đã thử hạ điều kiện chặn (rời nhau / tập con) và đo trên hồ
sơ thật — cả hai đều sinh thêm hơn 50 cặp ghép sai. Cách đúng là quy đồng cách
viết, giữ nguyên điều kiện chặn.
"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from core.config import EnterpriseConfig
from core.excel_reader import load_workbook_items
from core.internal_consistency import find_price_inconsistencies
from core.matcher import _has_type_conflict
from core.models import DocumentRole, ItemRecord, RowType
from core.text_normalizer import normalize_name


def _item(name: str, unit: str = "m") -> ItemRecord:
    return ItemRecord(
        source_id="x", role=DocumentRole.HSDT, bidder="b", workbook="w", sheet="s",
        row_number=1, row_type=RowType.DETAIL, stt="1", item_code="", item_name=name,
        unit=unit, material="", brand="", origin="", note="",
        normalized_name=normalize_name(name), normalized_code="",
        normalized_unit=unit, normalized_path="", structural_key="",
    )


@pytest.mark.parametrize("left, right", [
    ("Ống nhựa PVC D110 thoát nước", "Ống nhựa uPVC D110 thoát nước"),
    ("Ống nhựa uPVC D160", "Ống uPVC D160 thoát nước"),
    ("Ống HDPE D110 PN10", "Ống PE D110 PN10"),
    ("Ống nhựa PPR D25", "Ống PPR D25 nước nóng"),
])
def test_same_material_written_differently_is_not_blocked(left: str, right: str):
    assert _has_type_conflict(_item(left), _item(right)) is False


@pytest.mark.parametrize("left, right", [
    ("Ống thép tráng kẽm DN100 chữa cháy", "Ống nhựa luồn dây điện DN100"),
    ("Ống thép DN100", "Ống inox DN100"),
    ("Ống nhựa PVC D110", "Ống thép mạ kẽm D110"),
    ("Đầu báo khói địa chỉ", "Đầu báo nhiệt địa chỉ"),
    ("Cáp CXV/XLPE 4x50", "Cáp CVV/PVC 4x50"),
    ("Bình chữa cháy bột ABC", "Bình chữa cháy khí CO2"),
])
def test_genuinely_different_materials_stay_blocked(left: str, right: str):
    assert _has_type_conflict(_item(left), _item(right)) is True


# --------------------------------------------------------- đọc bảng đa ngành

CIVIL = [
    ("AF.31110", "Bê tông móng đá 1x2 M300", "m3", 860.0, 1_680_000),
    ("AF.61510", "Cốt thép cột D<=10mm", "kg", 45_800.0, 21_500),
    ("AE.22110", "Xây tường gạch ống 10x10x20 vữa M75", "m3", 620.4, 1_450_000),
    ("AK.51220", "Lát gạch granite 60x60 nền", "m2", 4_680.0, 320_000),
    ("AD.23214", "Bê tông nhựa chặt C19 dày 7cm", "m2", 18_600.0, 168_000),
    ("AB.11110", "Đào đất hữu cơ bằng máy đào <=0.8m3", "m3", 12_400.0, 32_000),
    ("AH.11120", "Ống HDPE D110 PN10 chôn ngầm", "m", 1_840.0, 285_000),
    ("NT.03", "Tủ bếp trên gỗ công nghiệp MFC", "md", 62.0, 4_200_000),
]


def _civil_book(tmp_path: Path, extra_rows: list[list] | None = None) -> Path:
    wb = Workbook(); ws = wb.active; ws.title = "Xay dung"
    ws.append(["STT", "Mã hiệu", "Nội dung công việc", "Đơn vị",
               "KL nhà thầu chào", "Đơn giá tổng hợp", "Thành tiền"])
    for index, (code, name, unit, qty, price) in enumerate(CIVIL, 1):
        ws.append([str(index), code, name, unit, qty, price, round(qty * price)])
    for row in extra_rows or []:
        ws.append(row)
    path = tmp_path / "xd.xlsx"
    wb.save(path)
    return path


def test_civil_bill_of_quantities_is_read(tmp_path: Path):
    data = load_workbook_items(_civil_book(tmp_path), DocumentRole.HSDT, bidder="NT")
    items = [i for i in data.items if i.is_comparable]
    assert len(items) == len(CIVIL)
    assert {i.unit for i in items} == {"m3", "kg", "m2", "m", "md"}
    assert sum(i.bid_amount or 0 for i in items) > 0


def test_price_inconsistency_works_on_civil_units(tmp_path: Path):
    # Cùng mã, cùng tên, cùng đơn vị m3 nhưng hai đơn giá lệch 12%.
    path = _civil_book(tmp_path, [
        ["9", "AF.31110", "Bê tông móng đá 1x2 M300", "m3", 100.0, 1_880_000, 188_000_000],
    ])
    data = load_workbook_items(path, DocumentRole.HSDT, bidder="NT")
    issues = find_price_inconsistencies(data.items)
    assert len(issues) == 1
    assert "AF.31110" in issues[0].key_label
    assert issues[0].spread_pct > 0.10


def test_volume_and_weight_units_are_not_treated_as_lump_sum(tmp_path: Path):
    # m3/kg/m2 phải được so giá; chỉ "lô", "gói"... mới bỏ qua.
    path = _civil_book(tmp_path, [
        ["9", "AF.61510", "Cốt thép cột D<=10mm", "kg", 1_000.0, 25_370, 25_370_000],
    ])
    data = load_workbook_items(path, DocumentRole.HSDT, bidder="NT")
    assert find_price_inconsistencies(data.items), "Đơn vị kg vẫn phải so được giá"


def test_lump_sum_units_are_still_skipped(tmp_path: Path):
    path = _civil_book(tmp_path, [
        ["9", "TG.01", "Trọn gói phần sân vườn", "lô", 1, 500_000_000, 500_000_000],
        ["10", "TG.01", "Trọn gói phần sân vườn", "lô", 1, 900_000_000, 900_000_000],
    ])
    data = load_workbook_items(path, DocumentRole.HSDT, bidder="NT")
    codes = [i.key_label for i in find_price_inconsistencies(data.items)]
    assert not any("TG.01" in c for c in codes), "Đơn vị trọn gói không so giá được"


def test_end_to_end_comparison_across_trades(tmp_path: Path):
    cfg = EnterpriseConfig()
    cfg.enable_semantic_matching = False
    cfg.enable_reranker = False

    reference = tmp_path / "pl01.xlsx"
    wb = Workbook(); ws = wb.active; ws.title = "Xay dung"
    ws.append(["STT", "Mã hiệu", "Nội dung công việc", "Đơn vị", "Khối lượng mời thầu"])
    for index, (code, name, unit, qty, _) in enumerate(CIVIL, 1):
        ws.append([str(index), code, name, unit, qty])
    wb.save(reference)

    from core.pipeline import compare_tender_files

    result = compare_tender_files(reference, [("NT", _civil_book(tmp_path))],
                                  output_path=tmp_path / "kq.xlsx", config=cfg)
    matched = [r for r in result.rows if r.reference is not None and r.candidate is not None]
    assert len(matched) == len(CIVIL), "Mọi hạng mục xây dựng phải ghép được với phụ lục"
