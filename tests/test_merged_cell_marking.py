"""Đánh dấu vào ô nằm trong vùng gộp phải rơi đúng ô neo.

Bảng khối lượng thật dùng rất nhiều ô gộp. Ô không phải góc trên trái của vùng
gộp là ``MergedCell``:

- gán ghi chú -> lỗi cứng, hỏng cả file kết quả;
- gán màu -> chạy được nhưng Excel hiển thị theo ô neo nên MÀU KHÔNG HIỆN, cảnh
  báo mất hút mà không ai biết.

Cả hai trường hợp đều phải chuyển về ô neo.
"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import PatternFill

from core.annotator import _append_comment, _markable_cell


def _sheet():
    wb = Workbook()
    ws = wb.active
    ws.merge_cells("B2:E2")     # ô neo là B2
    ws.merge_cells("C5:C8")     # gộp theo chiều dọc, ô neo là C5
    return wb, ws


def test_comment_on_merged_cell_would_crash_without_redirect():
    _, ws = _sheet()
    assert isinstance(ws.cell(2, 4), MergedCell)
    with pytest.raises(AttributeError):
        ws.cell(2, 4).comment = object()


def test_markable_cell_redirects_to_anchor():
    _, ws = _sheet()
    assert (_markable_cell(ws, 2, 4).row, _markable_cell(ws, 2, 4).column) == (2, 2)
    assert (_markable_cell(ws, 7, 3).row, _markable_cell(ws, 7, 3).column) == (5, 3)
    # Ô thường giữ nguyên vị trí.
    assert (_markable_cell(ws, 9, 1).row, _markable_cell(ws, 9, 1).column) == (9, 1)


def test_marking_through_anchor_survives_save(tmp_path: Path):
    wb, ws = _sheet()
    cell = _markable_cell(ws, 2, 4)
    cell.fill = PatternFill("solid", fgColor="FCE4D6")
    _append_comment(cell, "Lệch khối lượng")
    out = tmp_path / "danh_dau.xlsx"
    wb.save(out)

    reread = load_workbook(out).active
    anchor = reread.cell(2, 2)
    assert anchor.comment is not None and "Lệch khối lượng" in anchor.comment.text
    assert anchor.fill.fgColor.rgb.endswith("FCE4D6"), "Màu phải nằm ở ô neo mới hiện ra"


def test_second_comment_is_appended_not_replaced():
    _, ws = _sheet()
    cell = _markable_cell(ws, 2, 3)
    _append_comment(cell, "Lỗi thứ nhất")
    _append_comment(cell, "Lỗi thứ hai")
    assert "Lỗi thứ nhất" in cell.comment.text
    assert "Lỗi thứ hai" in cell.comment.text


def test_anchor_map_is_per_worksheet(tmp_path: Path):
    # Đệm phải gắn theo worksheet: hai file khác nhau không được dùng chung bản đồ.
    _, a = _sheet()
    wb_b = Workbook(); b = wb_b.active
    b.merge_cells("A1:A4")
    assert (_markable_cell(a, 2, 4).row, _markable_cell(a, 2, 4).column) == (2, 2)
    assert (_markable_cell(b, 3, 1).row, _markable_cell(b, 3, 1).column) == (1, 1)
    assert (_markable_cell(b, 2, 4).row, _markable_cell(b, 2, 4).column) == (2, 4)
