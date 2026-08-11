"""Dòng tổng phụ của mục không phải hạng mục để đối chiếu.

Dòng có thành tiền nhưng KHÔNG có đơn vị, khối lượng lẫn đơn giá thì không
đối chiếu được gì. Giữ lại chỉ sinh ra "hạng mục phát sinh ngoài Phụ lục 01"
giả với giá trị bằng cả một hệ thống.

Trước đây luật còn đòi STT phải dạng "A", "I", "II"..., nên tổng phụ đánh số
thường hoặc bỏ trống STT vẫn lọt qua.

Ngược lại, hạng mục trọn gói thật — có ĐVT "Lô", khối lượng 1 và đơn giá —
phải được giữ để còn so sánh giá.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from core.excel_reader import load_workbook_items
from core.models import DocumentRole, RowType

HEADER = ["STT", "Mã hiệu", "Tên hạng mục", "ĐVT", "KL nhà thầu chào",
          "Đơn giá tổng hợp", "Thành tiền"]
_BASE = [
    ["1", "DI-M9102", "Đầu báo khói địa chỉ", "cái", 10, 500_000, 5_000_000],
    ["2", "VG-50", "Van góc chữa cháy DN50", "cái", 4, 900_000, 3_600_000],
]


def _book(tmp_path: Path, rows: list[list], name: str = "hs.xlsx") -> Path:
    wb = Workbook(); ws = wb.active; ws.title = "BoQ chi tiet"
    ws.append(HEADER)
    for r in _BASE + rows:
        ws.append(r)
    path = tmp_path / name
    wb.save(path)
    return path


def _row(tmp_path: Path, row: list, name: str):
    data = load_workbook_items(_book(tmp_path, [row], name), DocumentRole.HSDT, bidder="NT")
    label = str(row[2])
    found = [i for i in data.items if (i.item_name or "").strip() == label]
    assert found, f"Không đọc được dòng {label!r}"
    return found[0]


def test_numeric_stt_subtotal_is_not_an_item(tmp_path: Path):
    # STT thường ("1"), không ĐVT, không KL, không đơn giá — chỉ có thành tiền.
    row = ["1", "", "HẠNG MỤC: ĐIỆN", "", None, None, 85_852_544_382]
    assert _row(tmp_path, row, "so.xlsx").row_type is RowType.SUMMARY


def test_blank_stt_subtotal_is_not_an_item(tmp_path: Path):
    row = ["", "", "ĐẦU MỤC CÔNG VIỆC THEO KLMT", "", None, None, 176_354_076_899]
    assert _row(tmp_path, row, "trong.xlsx").row_type is RowType.SUMMARY


def test_subtotal_does_not_inflate_total(tmp_path: Path):
    path = _book(tmp_path, [["", "", "HẠNG MỤC: PCCC", "", None, None, 8_600_000]])
    data = load_workbook_items(path, DocumentRole.HSDT, bidder="NT")
    total = sum(i.bid_amount or 0 for i in data.items if i.is_comparable)
    assert total == 8_600_000, "Tổng thành tiền không được cộng thêm dòng tổng phụ"


def test_lump_sum_item_is_kept(tmp_path: Path):
    # Hạng mục trọn gói thật: có ĐVT, khối lượng và đơn giá -> vẫn phải đối chiếu.
    row = ["3", "", "HỆ THỐNG ĐIỆN", "Lô", 1, 85_852_544_382, 85_852_544_382]
    item = _row(tmp_path, row, "lo.xlsx")
    assert item.row_type is not RowType.SUMMARY
    assert item.is_comparable


def test_row_with_quantity_but_no_price_is_kept(tmp_path: Path):
    # Thiếu đơn giá là lỗi cần cảnh báo, không phải lý do loại khỏi đối chiếu.
    row = ["4", "", "Ống thép tráng kẽm DN100", "m", 350, None, None]
    assert _row(tmp_path, row, "kl.xlsx").is_comparable
