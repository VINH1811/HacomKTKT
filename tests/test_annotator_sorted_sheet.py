"""File đánh dấu: bản SẮP XẾP mang tên gốc, theo đúng cấu trúc A / B / C.

- Hạng mục phát sinh (không có trong PL01) đang nằm lẫn trong phần A được DỜI
  xuống phần B (phát sinh ngoài KLMT); giữ nguyên các mục B sẵn có.
- Ba dòng tổng A / B / C được viết lại theo dải dòng mới (A, B là SUBTOTAL;
  C = A + B) nên số tiền đúng.
- Công thức thành tiền trong-dòng vẫn "sống"; bản gốc giữ ở sheet '— gốc'.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from core.annotator import _phatsinh_block_rows
from core.config import EnterpriseConfig
from core.models import ComparedItem, DocumentRole, ItemRecord, MatchKind, MatchResult, RowType
from core.tender_package import compare_appendices_with_bidders


def _cmp(row: int, row_type: RowType, matched: bool) -> ComparedItem:
    cand = ItemRecord(source_id="c", role=DocumentRole.HSDT, bidder="NT A", workbook="w",
                      sheet="S", row_number=row, row_type=row_type)
    ref = (ItemRecord(source_id="r", role=DocumentRole.HSMT, bidder="PL01", workbook="w",
                      sheet="S", row_number=row, row_type=row_type) if matched else None)
    kind = MatchKind.EXACT_NAME if matched else MatchKind.EXTRA
    return ComparedItem("cid", "NT A", ref, cand, MatchResult(0 if matched else None, row, kind, 1.0))


def test_only_individually_marked_phatsinh_rows_are_moved():
    # Dòng ĐÃ KHỚP đứng ngay sau một hạng mục phát sinh KHÔNG bị coi là phát sinh
    # (không bị dời theo khối).
    rows = [
        _cmp(1, RowType.DETAIL, True),       # khớp
        _cmp(2, RowType.DETAIL, False),      # phát sinh
        _cmp(3, RowType.COMPONENT, True),    # ĐÃ KHỚP, ngay sau phát sinh -> KHÔNG dời
        _cmp(4, RowType.COMPONENT, False),   # phát sinh (con của mục phát sinh) -> dời
    ]
    assert _phatsinh_block_rows(rows)["S"] == {2, 4}

SHEET = "1. HT điện"
GOC = "1. HT điện — gốc"
NAME_COL = 3
AMOUNT_COL = 7  # cột Thành tiền (G)


def _cfg(**kw) -> EnterpriseConfig:
    cfg = EnterpriseConfig()
    cfg.enable_semantic_matching = False
    cfg.enable_reranker = False
    for k, v in kw.items():
        setattr(cfg, k, v)
    return cfg


def _annotate(tmp_path: Path) -> Path:
    pl1 = tmp_path / "pl1.xlsx"
    wb = Workbook(); ws = wb.active; ws.title = "KLMT"
    ws.append(["STT", "Mã hiệu", "Tên hạng mục", "ĐVT", "Khối lượng"])
    ws.append(["1", "M-01", "Tủ điện tổng", "Cái", 2])
    wb.save(pl1)

    bidder = tmp_path / "b.xlsx"
    wb = Workbook(); ws = wb.active; ws.title = SHEET
    ws.append(["STT", "Mã hiệu", "Tên hạng mục", "ĐVT", "Khối lượng nhà thầu chào", "Đơn giá tổng hợp", "Thành tiền"])
    ws.append(["A", "", "ĐẦU MỤC CÔNG VIỆC THEO KLMT", "", "", "", "=SUBTOTAL(9,G3:G4)"])          # r2 - tổng A
    ws.append(["1", "M-01", "Tủ điện tổng", "Cái", 2, 1_000_000, "=E3*F3"])                          # r3 - khớp
    ws.append(["2", "M-99", "Phát sinh trong A", "Cái", 1, 500_000, "=E4*F4"])                       # r4 - phát sinh lẫn trong A
    ws.append(["B", "", "Phát sinh ngoài KLMT Nhà thầu bổ sung", "", "", "", "=SUBTOTAL(9,G6:G6)"])  # r5 - tổng B
    ws.append(["", "", "Ống mềm bổ sung", "Cái", 1, 300_000, "=E6*F6"])                              # r6 - mục B sẵn có
    ws.append(["C", "", "TỔNG CỘNG TRƯỚC VAT: C = A + B", "", "", "", "=G2+G5"])                     # r7 - tổng cộng
    wb.save(bidder)

    out = compare_appendices_with_bidders(
        [("NT A", bidder)], tmp_path / "out", pl1_path=pl1,
        config=_cfg(annotate_workers=1, parse_cache_size=0, match_cache_size=0),
    )
    return Path(out.annotated_files["NT A"])


def _row_of(ws, part: str) -> int:
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, NAME_COL).value
        if v is not None and part in str(v):
            return r
    raise AssertionError(f"Không thấy dòng chứa '{part}'")


def test_sorted_sheet_takes_original_name_goc_kept(tmp_path: Path):
    wb = load_workbook(_annotate(tmp_path))
    assert SHEET in wb.sheetnames
    assert GOC in wb.sheetnames


def test_phatsinh_in_A_moved_into_section_B(tmp_path: Path):
    ws = load_workbook(_annotate(tmp_path))[SHEET]
    r_a = _row_of(ws, "ĐẦU MỤC CÔNG VIỆC")
    r_matched = _row_of(ws, "Tủ điện tổng")
    r_b = _row_of(ws, "Phát sinh ngoài KLMT")
    r_existing_b = _row_of(ws, "Ống mềm bổ sung")
    r_moved = _row_of(ws, "Phát sinh trong A")
    r_c = _row_of(ws, "TỔNG CỘNG")

    # Phần A: chỉ còn hạng mục khớp; phát sinh KHÔNG còn ở phần A.
    assert r_a < r_matched < r_b
    assert r_matched < r_moved
    # Phát sinh đã nằm trong phần B (sau tiêu đề B), cùng mục B sẵn có.
    assert r_b < r_existing_b
    assert r_b < r_moved
    # Trình tự: ... A ... B ... (mục B sẵn có + phát sinh dời xuống) ... C.
    assert r_moved < r_c and r_existing_b < r_c


def test_totals_rewritten_live(tmp_path: Path):
    ws = load_workbook(_annotate(tmp_path))[SHEET]
    r_a = _row_of(ws, "ĐẦU MỤC CÔNG VIỆC")
    r_b = _row_of(ws, "Phát sinh ngoài KLMT")
    r_c = _row_of(ws, "TỔNG CỘNG")

    # Tổng A và B là SUBTOTAL sống theo dải dòng mới.
    assert str(ws.cell(r_a, AMOUNT_COL).value or "").upper().startswith("=SUBTOTAL(9,")
    assert str(ws.cell(r_b, AMOUNT_COL).value or "").upper().startswith("=SUBTOTAL(9,")
    # Tổng C = A + B, trỏ đúng dòng tiêu đề A và B mới.
    assert ws.cell(r_c, AMOUNT_COL).value == f"=G{r_a}+G{r_b}"


def test_row_formula_translated_after_move(tmp_path: Path):
    ws = load_workbook(_annotate(tmp_path))[SHEET]
    r_moved = _row_of(ws, "Phát sinh trong A")
    # Công thức thành tiền trong-dòng của mục đã dời phải trỏ đúng dòng mới.
    assert ws.cell(r_moved, AMOUNT_COL).value == f"=E{r_moved}*F{r_moved}"


def test_goc_sheet_unchanged(tmp_path: Path):
    goc = load_workbook(_annotate(tmp_path))[GOC]
    # Bản gốc giữ nguyên vị trí: phát sinh vẫn ở dòng 4 (trong phần A), công thức gốc.
    assert goc.cell(4, NAME_COL).value == "Phát sinh trong A"
    assert goc.cell(4, AMOUNT_COL).value == "=E4*F4"
