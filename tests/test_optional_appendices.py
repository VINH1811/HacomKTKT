from pathlib import Path

from openpyxl import Workbook

from core.config import EnterpriseConfig
from core.tender_package import compare_pl1_pl2_with_bidders


def _pl1(path: Path):
    wb = Workbook(); ws = wb.active; ws.title = "Điện"
    ws.append(["Mã hiệu", "Tên hạng mục", "ĐVT", "Khối lượng mời thầu"])
    ws.append(["CAP-01", "Hệ thống dây cáp hạ thế", "m", 100])
    wb.save(path)


def _pl2(path: Path):
    wb = Workbook(); ws = wb.active; ws.title = "DMVT"
    ws.append(["STT", "VẬT TƯ THIẾT BỊ", "THƯƠNG HIỆU - XUẤT XỨ", "GHI CHÚ"])
    ws.append([1, "Hệ thống dây cáp hạ thế", "Cadivi/ Thipha/ Taisin - Asia", ""])
    wb.save(path)


def _bid(path: Path, price: float, brand: str = "Cadivi"):
    wb = Workbook(); ws = wb.active; ws.title = "Điện"
    ws.append([
        "Mã hiệu", "Tên hạng mục", "ĐVT", "Khối lượng mời thầu", "Khối lượng nhà thầu",
        "Thương hiệu", "Xuất xứ", "Đơn giá tổng hợp", "Thành tiền theo KLMT", "Thành tiền nhà thầu",
    ])
    ws.append(["CAP-01", "Hệ thống dây cáp hạ thế", "m", 100, 100, brand, "VN", price, 100 * price, 100 * price])
    wb.save(path)


def _cfg() -> EnterpriseConfig:
    cfg = EnterpriseConfig()
    cfg.enable_semantic_matching = False
    cfg.enable_reranker = False
    return cfg


def test_pl1_only_is_supported(tmp_path: Path):
    pl1 = tmp_path / "pl1.xlsx"; a = tmp_path / "a.xlsx"; b = tmp_path / "b.xlsx"
    _pl1(pl1); _bid(a, 100_000); _bid(b, 130_000)
    outputs = compare_pl1_pl2_with_bidders(pl1, None, [("A", a), ("B", b)], tmp_path / "out1", _cfg())
    assert outputs.report_path.exists()
    assert outputs.result.audit["mode"] == "PL01_ONLY_VS_MULTI_HSDT_PRICE_PEER"
    assert outputs.result.audit["peer_price_comparison_enabled"] is True
    assert outputs.result.audit["peer_comparison_scope"] == "price_only"
    assert outputs.result.audit["pl2_sha256"] == "NOT_PROVIDED"
    assert any("Không có PL02" in warning for warning in outputs.result.warnings)


def test_pl2_only_uses_multiway_peer_consensus(tmp_path: Path):
    pl2 = tmp_path / "pl2.xlsx"; a = tmp_path / "a.xlsx"; b = tmp_path / "b.xlsx"
    _pl2(pl2); _bid(a, 100_000); _bid(b, 130_000, "Ngoài danh sách")
    outputs = compare_pl1_pl2_with_bidders(None, pl2, [("A", a), ("B", b)], tmp_path / "out2", _cfg())
    assert outputs.report_path.exists()
    assert outputs.result.audit["mode"] == "PL02_ONLY_MULTIWAY_HSDT_PRICE_PEER"
    assert outputs.result.audit["catalogue_mode"] == "MULTIWAY_PEER_CONSENSUS"
    assert outputs.result.audit["peer_cluster_stats"]["clusters"] >= 1
    assert "baseline_bidder" not in outputs.result.audit
    assert all("không có baseline" in str(row.match.reason).lower() or row.match.kind.value == "missing" for row in outputs.result.rows)
    assert any(row.pl2_status for row in outputs.result.rows if row.candidate)


def test_single_bidder_pl1_compares_appendix_without_peer_price(tmp_path: Path):
    pl1 = tmp_path / "pl1.xlsx"; a = tmp_path / "a.xlsx"
    _pl1(pl1); _bid(a, 100_000)

    outputs = compare_pl1_pl2_with_bidders(
        pl1,
        None,
        [("A", a)],
        tmp_path / "out_single",
        _cfg(),
    )

    assert outputs.report_path.exists()
    assert outputs.package_zip.exists()
    assert outputs.result.audit["mode"] == "PL01_ONLY_VS_SINGLE_HSDT"
    assert outputs.result.audit["bidder_count"] == 1
    assert outputs.result.audit["peer_price_comparison_enabled"] is False
    assert outputs.result.audit["peer_comparison_scope"] == "disabled"
    assert outputs.result.audit["peer_stats"]["enabled"] is False
    assert not any(
        difference.field.startswith("So sánh ngang hàng")
        for row in outputs.result.rows
        for difference in row.differences
    )


def test_package_multi_bidder_peer_stage_is_price_only(tmp_path: Path):
    pl1 = tmp_path / "pl1.xlsx"; a = tmp_path / "a.xlsx"; b = tmp_path / "b.xlsx"
    _pl1(pl1); _bid(a, 100_000); _bid(b, 300_000)

    outputs = compare_pl1_pl2_with_bidders(
        pl1,
        None,
        [("A", a), ("B", b)],
        tmp_path / "out_price_only",
        _cfg(),
    )

    peer_fields = [
        difference.field
        for row in outputs.result.rows
        for difference in row.differences
        if difference.field.startswith("So sánh ngang hàng")
    ]
    assert peer_fields
    assert any("Đơn giá tổng hợp" in field for field in peer_fields)
    assert not any("Khối lượng" in field for field in peer_fields)
    assert not any("Đơn vị" in field for field in peer_fields)

