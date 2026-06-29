"""Phụ lục 02 (thương hiệu/xuất xứ) phải được so sánh VÀ đánh dấu trực tiếp lên
ô tương ứng trong bảng tổng hợp; đồng thời 'VN' và 'Việt Nam' phải coi là cùng
một xuất xứ (không báo lệch giả).
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from core.config import EnterpriseConfig
from core.pl2_reader import _origin_allowed
from core.reporter import export_consolidated_summary
from core.tender_package import compare_appendices_with_bidders


def _cfg() -> EnterpriseConfig:
    cfg = EnterpriseConfig()
    cfg.enable_semantic_matching = False
    cfg.enable_reranker = False
    return cfg


# --- Xuất xứ: VN ↔ Việt Nam ---------------------------------------------------

def test_vn_equals_vietnam():
    assert _origin_allowed("Việt Nam", ("VN",)) is True
    assert _origin_allowed("VN", ("Việt Nam",)) is True
    assert _origin_allowed("Trung Quốc", ("China",)) is True


def test_different_origins_still_flagged():
    assert _origin_allowed("Việt Nam", ("China",)) is False
    assert _origin_allowed("Trung Quốc", ("VN",)) is False


# --- Đánh dấu PL02 trong bảng tổng hợp ----------------------------------------

def _make_pl1(path: Path) -> None:
    wb = Workbook(); ws = wb.active; ws.title = "1. HT điện"
    ws.append(["STT", "Mã hiệu", "Tên hạng mục", "ĐVT", "Khối lượng mời thầu", "Đơn giá tổng hợp", "Thành tiền"])
    ws.append(["1", "C-01", "Hệ thống dây cáp hạ thế", "m", 100, 50_000, 5_000_000])
    wb.save(path)


def _make_pl2(path: Path) -> None:
    wb = Workbook(); ws = wb.active; ws.title = "DMVT"
    ws.append(["STT", "VẬT TƯ THIẾT BỊ", "THƯƠNG HIỆU - XUẤT XỨ", "GHI CHÚ"])
    ws.append(["1", "Hệ thống dây cáp hạ thế", "Cadivi / Thipha - VN", ""])
    wb.save(path)


def _make_bidder(path: Path, brand: str, origin: str) -> None:
    wb = Workbook(); ws = wb.active; ws.title = "1. HT điện"
    ws.append(["STT", "Mã hiệu", "Tên hạng mục", "ĐVT", "Khối lượng nhà thầu",
               "Đơn giá tổng hợp", "Thành tiền", "Thương hiệu", "Xuất xứ"])
    ws.append(["1", "C-01", "Hệ thống dây cáp hạ thế", "m", 100, 52_000, 5_200_000, brand, origin])
    wb.save(path)


def _leaf_cols(ws) -> dict[str, list[int]]:
    result: dict[str, list[int]] = {}
    for cell in ws[4]:
        if cell.value:
            result.setdefault(str(cell.value).replace("\n", " ").strip(), []).append(cell.column)
    return result


def test_summary_marks_pl2_brand_violation_on_cell(tmp_path: Path):
    pl1 = tmp_path / "pl1.xlsx"; pl2 = tmp_path / "pl2.xlsx"; bidder = tmp_path / "b.xlsx"
    _make_pl1(pl1); _make_pl2(pl2)
    # Thương hiệu 'Taisin' KHÔNG có trong PL02 (Cadivi/Thipha) -> phải bị đánh dấu.
    # Xuất xứ 'Việt Nam' == 'VN' trong PL02 -> KHÔNG bị đánh dấu.
    _make_bidder(bidder, brand="Taisin", origin="Việt Nam")

    out = compare_appendices_with_bidders(
        [("NT A", bidder)], tmp_path / "out", pl1_path=pl1, pl2_path=pl2, config=_cfg()
    )
    summary = tmp_path / "summary.xlsx"
    export_consolidated_summary(out.result, summary)

    wb = load_workbook(summary)
    ws = wb["1. HT điện"]
    cols = _leaf_cols(ws)
    brand_col = cols["Thương hiệu"][0]
    origin_col = cols["Xuất xứ"][0]

    # Tìm dòng dữ liệu (có giá trị thương hiệu).
    data_row = next(r for r in range(5, ws.max_row + 1) if ws.cell(r, brand_col).value)
    brand_cell = ws.cell(data_row, brand_col)
    origin_cell = ws.cell(data_row, origin_col)

    # Ô thương hiệu sai phải có ghi chú PL02 + tô màu.
    assert brand_cell.comment is not None and "Phụ lục 02" in brand_cell.comment.text
    assert brand_cell.fill is not None and brand_cell.fill.fgColor.rgb not in (None, "00000000")

    # Ô xuất xứ 'Việt Nam' khớp 'VN' -> KHÔNG bị đánh dấu (không dương tính giả).
    assert origin_cell.comment is None
