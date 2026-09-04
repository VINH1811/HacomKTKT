"""So sánh hai phiên bản chào giá của cùng nhà thầu (V1 → V2): phát hiện đúng
hạng mục đổi giá/khối lượng/thương hiệu, thêm mới, đã xoá; tổng tiền hai bản và
mức chênh; chịu được xáo thứ tự dòng giữa hai bản."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from core.config import EnterpriseConfig
from core.version_compare import (
    STATUS_ADDED,
    STATUS_CHANGED,
    STATUS_REMOVED,
    STATUS_UNCHANGED,
    compare_quote_versions,
    export_version_report,
)

SHEET = "1. HT dien"
HEADERS = ["STT", "Mã hiệu", "Tên hạng mục", "ĐVT", "Khối lượng nhà thầu chào",
           "Đơn giá tổng hợp", "Thành tiền", "Thương hiệu", "Xuất xứ"]


def _cfg() -> EnterpriseConfig:
    cfg = EnterpriseConfig()
    cfg.enable_semantic_matching = False
    cfg.enable_reranker = False
    return cfg


def _write(path: Path, rows: list[list]) -> None:
    wb = Workbook(); ws = wb.active; ws.title = SHEET
    ws.append(HEADERS)
    for row in rows:
        ws.append(row)
    wb.save(path)


def _row(stt, code, name, kl, dg, brand="Schneider", origin="Pháp"):
    return [stt, code, name, "Cái", kl, dg, kl * dg, brand, origin]


def _base_rows() -> list[list]:
    return [
        _row("1", "TD-01", "Tủ điện tổng MSB", 2, 500_000_000),
        _row("2", "TD-02", "Tủ điện tầng T1", 5, 80_000_000),
        _row("3", "CAP-01", "Cáp Cu/XLPE/PVC 3x95", 1200, 850_000),
        _row("4", "ATS-01", "Bộ chuyển nguồn ATS 4P 800A", 1, 250_000_000),
    ]


def _compare(tmp_path: Path, old_rows, new_rows):
    p1, p2 = tmp_path / "v1.xlsx", tmp_path / "v2.xlsx"
    _write(p1, old_rows)
    _write(p2, new_rows)
    return compare_quote_versions(p1, p2, "NT Test", config=_cfg())


def test_price_change_detected(tmp_path: Path):
    new = _base_rows()
    new[1] = _row("2", "TD-02", "Tủ điện tầng T1", 5, 90_000_000)  # tăng đơn giá
    res = _compare(tmp_path, _base_rows(), new)
    changed = [r for r in res.rows if r.status == STATUS_CHANGED]
    assert len(changed) == 1 and "TD-02" in (changed[0].new.item_code or "")
    fields = {c.field for c in changed[0].changes}
    assert "Đơn giá tổng hợp" in fields and "Thành tiền" in fields
    dg = next(c for c in changed[0].changes if c.field == "Đơn giá tổng hợp")
    assert dg.old_value == 80_000_000 and dg.new_value == 90_000_000
    assert abs(dg.delta_pct - 0.125) < 1e-9


def test_added_and_removed_items(tmp_path: Path):
    new = _base_rows()[:3]  # bỏ ATS-01
    new.append(_row("5", "DEN-01", "Đèn LED âm trần 12W", 300, 150_000))  # thêm mới
    res = _compare(tmp_path, _base_rows(), new)
    assert res.count(STATUS_REMOVED) == 1
    assert res.count(STATUS_ADDED) == 1
    removed = next(r for r in res.rows if r.status == STATUS_REMOVED)
    assert "ATS" in (removed.old.item_code or "")


def test_reordered_rows_still_match(tmp_path: Path):
    # V2 đảo lộn thứ tự nhưng nội dung y hệt -> tất cả GIỮ NGUYÊN.
    new = list(reversed(_base_rows()))
    res = _compare(tmp_path, _base_rows(), new)
    assert res.count(STATUS_UNCHANGED) == 4
    assert res.count(STATUS_CHANGED) == 0
    assert res.total_delta == 0


def test_brand_change_detected(tmp_path: Path):
    new = _base_rows()
    new[2] = _row("3", "CAP-01", "Cáp Cu/XLPE/PVC 3x95", 1200, 850_000, brand="LS Vina", origin="Việt Nam")
    res = _compare(tmp_path, _base_rows(), new)
    changed = next(r for r in res.rows if r.status == STATUS_CHANGED)
    fields = {c.field for c in changed.changes}
    assert {"Thương hiệu", "Xuất xứ"} <= fields
    # Giá không đổi -> không báo trường giá.
    assert "Đơn giá tổng hợp" not in fields


def test_totals_and_report_export(tmp_path: Path):
    new = _base_rows()
    new[0] = _row("1", "TD-01", "Tủ điện tổng MSB", 2, 550_000_000)  # +100tr
    res = _compare(tmp_path, _base_rows(), new)
    assert res.total_delta == 100_000_000
    out = tmp_path / "bao_cao.xlsx"
    export_version_report(res, out)
    wb = load_workbook(out)
    assert set(wb.sheetnames) == {"Tổng quan", "Theo sheet", "Thay đổi chi tiết"}
    det = wb["Thay đổi chi tiết"]
    statuses = {det.cell(r, 1).value for r in range(2, det.max_row + 1)}
    assert STATUS_CHANGED in statuses
    # Dòng GIỮ NGUYÊN không được liệt kê trong sheet chi tiết.
    assert STATUS_UNCHANGED not in statuses
