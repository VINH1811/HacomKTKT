"""So sánh phiên bản phải cho biết lỗi tự mâu thuẫn giá ĐÃ SỬA hay CÒN.

Sau vòng làm rõ, chuyên viên cần trả lời đúng một câu: nhà thầu đã xử lý lỗi
chưa. Chỉ liệt kê lỗi của từng bản riêng lẻ thì người dùng phải tự đối chiếu.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl import Workbook

from core.config import EnterpriseConfig
from core.version_compare import (
    PRICE_ISSUE_FIXED,
    PRICE_ISSUE_NEW,
    PRICE_ISSUE_REMAINS,
    compare_quote_versions,
    export_version_report,
)

HEADER = ["STT", "Mã hiệu", "Tên hạng mục", "ĐVT",
          "KL nhà thầu chào", "Đơn giá tổng hợp", "Thành tiền"]


def _quote(path: Path, rows: list[tuple]) -> Path:
    wb = Workbook(); ws = wb.active; ws.title = "PCCC"
    ws.append(HEADER)
    for stt, code, name, unit, qty, price in rows:
        ws.append([stt, code, name, unit, qty, price, qty * price])
    wb.save(path)
    return path


def _config() -> EnterpriseConfig:
    cfg = EnterpriseConfig()
    cfg.enable_semantic_matching = False
    cfg.enable_reranker = False
    return cfg


def _result(tmp_path: Path):
    # Bản cũ: đầu báo khói chào 2 giá (lỗi), van góc chào nhất quán.
    old = _quote(tmp_path / "v1.xlsx", [
        ("1", "DI-M9102", "Đầu báo khói địa chỉ loại thường", "cái", 120, 500_000),
        ("2", "VG-50", "Van góc chữa cháy DN50 bằng đồng", "cái", 20, 900_000),
        ("3", "DI-M9102", "Đầu báo khói địa chỉ loại thường", "cái", 30, 560_000),
        ("4", "VG-50", "Van góc chữa cháy DN50 bằng đồng", "cái", 5, 900_000),
    ])
    # Bản mới: đầu báo khói ĐÃ SỬA về cùng giá; van góc lại MỚI PHÁT SINH lệch.
    new = _quote(tmp_path / "v2.xlsx", [
        ("1", "DI-M9102", "Đầu báo khói địa chỉ loại thường", "cái", 120, 500_000),
        ("2", "VG-50", "Van góc chữa cháy DN50 bằng đồng", "cái", 20, 900_000),
        ("3", "DI-M9102", "Đầu báo khói địa chỉ loại thường", "cái", 30, 500_000),
        ("4", "VG-50", "Van góc chữa cháy DN50 bằng đồng", "cái", 5, 1_150_000),
    ])
    return compare_quote_versions(old, new, "Nhà thầu A", config=_config())


def _by_status(result) -> dict[str, list]:
    out: dict[str, list] = {}
    for issue in result.price_issues:
        out.setdefault(issue.status, []).append(issue)
    return out


def test_fixed_and_new_issues_are_classified(tmp_path: Path):
    grouped = _by_status(_result(tmp_path))
    fixed = grouped.get(PRICE_ISSUE_FIXED, [])
    added = grouped.get(PRICE_ISSUE_NEW, [])
    assert len(fixed) == 1 and "DI-M9102" in fixed[0].key_label, "Lỗi đã sửa phải nhận ra"
    assert len(added) == 1 and "VG-50" in added[0].key_label, "Lỗi mới phát sinh phải nhận ra"
    assert not grouped.get(PRICE_ISSUE_REMAINS)


def test_issue_still_present_is_marked_remaining(tmp_path: Path):
    # Cả hai bản đều chào lệch giá cùng một hạng mục -> CÒN LỖI.
    rows = [("1", "DI-M9102", "Đầu báo khói địa chỉ loại thường", "cái", 120, 500_000),
            ("2", "DI-M9102", "Đầu báo khói địa chỉ loại thường", "cái", 30, 560_000)]
    result = compare_quote_versions(
        _quote(tmp_path / "a.xlsx", rows), _quote(tmp_path / "b.xlsx", rows),
        "Nhà thầu A", config=_config())
    assert [i.status for i in result.price_issues] == [PRICE_ISSUE_REMAINS]


def test_clean_versions_report_no_issue(tmp_path: Path):
    rows = [("1", "DI-M9102", "Đầu báo khói địa chỉ loại thường", "cái", 120, 500_000),
            ("2", "VG-50", "Van góc chữa cháy DN50 bằng đồng", "cái", 20, 900_000)]
    result = compare_quote_versions(
        _quote(tmp_path / "a.xlsx", rows), _quote(tmp_path / "b.xlsx", rows),
        "Nhà thầu A", config=_config())
    assert result.price_issues == []


def test_pending_issues_are_sorted_before_fixed(tmp_path: Path):
    statuses = [i.status for i in _result(tmp_path).price_issues]
    assert statuses.index(PRICE_ISSUE_NEW) < statuses.index(PRICE_ISSUE_FIXED)


def test_report_has_price_issue_sheet(tmp_path: Path):
    out = export_version_report(_result(tmp_path), tmp_path / "bc.xlsx")
    wb = openpyxl.load_workbook(out)
    assert "Lệch đơn giá nội bộ" in wb.sheetnames
    ws = wb["Lệch đơn giá nội bộ"]
    text = "\n".join(str(c.value or "") for row in ws.iter_rows() for c in row)
    assert PRICE_ISSUE_NEW in text and PRICE_ISSUE_FIXED in text
    # Phải chỉ rõ dòng để mở file kiểm lại được.
    assert "dòng" in text

    overview = "\n".join(str(c.value or "") for row in wb["Tổng quan"].iter_rows() for c in row)
    assert "TỰ MÂU THUẪN ĐƠN GIÁ" in overview


def test_report_omits_sheet_when_no_issue(tmp_path: Path):
    rows = [("1", "DI-M9102", "Đầu báo khói địa chỉ loại thường", "cái", 120, 500_000)]
    result = compare_quote_versions(
        _quote(tmp_path / "a.xlsx", rows), _quote(tmp_path / "b.xlsx", rows),
        "Nhà thầu A", config=_config())
    out = export_version_report(result, tmp_path / "bc.xlsx")
    assert "Lệch đơn giá nội bộ" not in openpyxl.load_workbook(out).sheetnames
