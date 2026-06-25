from pathlib import Path

from openpyxl import Workbook, load_workbook

from core.config import EnterpriseConfig
from core.tender_package import compare_pl1_pl2_with_bidders


def _pl1(path: Path):
    wb = Workbook(); ws = wb.active; ws.title = "Điện"
    ws.append(["Mã hiệu", "Tên hạng mục", "ĐVT", "Khối lượng mời thầu"])
    ws.append(["CAP-01", "Hệ thống dây cáp hạ thế", "m", 100])
    wb.save(path)


def _pl2(path: Path):
    wb = Workbook(); ws = wb.active; ws.title = "DMVT CƠ ĐIỆN"
    ws.append(["PHỤ LỤC 02: BẢNG DANH MỤC VẬT TƯ THIẾT BỊ"])
    ws.append(["Dự án: HACOM MALL"])
    ws.append([])
    ws.append(["STT", "VẬT TƯ THIẾT BỊ", "THƯƠNG HIỆU - XUẤT XỨ", "GHI CHÚ"])
    ws.append(["I", "HỆ THỐNG ĐIỆN", None, None])
    ws.append([1, "Hệ thống dây cáp hạ thế", "Cadivi/ Thipha/ Taisin - Asia", None])
    wb.save(path)


def _bid(path: Path, brand: str, price: float):
    wb = Workbook(); ws = wb.active; ws.title = "Điện"
    ws.append([
        "Mã hiệu", "Tên hạng mục", "ĐVT", "Khối lượng mời thầu", "Khối lượng nhà thầu",
        "Thương hiệu", "Xuất xứ", "Đơn giá tổng hợp", "Thành tiền theo KLMT", "Thành tiền nhà thầu",
    ])
    ws.append(["CAP-01", "Hệ thống dây cáp hạ thế", "m", 100, 100, brand, "VN", price, 100 * price, 100 * price])
    ws["L2"] = "=H2*D2"  # unrelated formula must remain intact after annotation
    wb.save(path)


def test_package_mode_uses_pl1_pl2_and_no_bidder_baseline(tmp_path: Path):
    pl1, pl2 = tmp_path / "pl1.xlsx", tmp_path / "pl2.xlsx"
    a, b = tmp_path / "a.xlsx", tmp_path / "b.xlsx"
    _pl1(pl1); _pl2(pl2); _bid(a, "Cadivi", 100_000); _bid(b, "Ngoài danh sách", 300_000)

    cfg = EnterpriseConfig(); cfg.enable_semantic_matching = False; cfg.enable_reranker = False
    outputs = compare_pl1_pl2_with_bidders(pl1, pl2, [("A", a), ("B", b)], tmp_path / "out", cfg)

    assert outputs.report_path.exists()
    assert outputs.package_zip.exists()
    assert set(outputs.annotated_files) == {"A", "B"}
    assert outputs.result.audit["peer_price_comparison_enabled"] is True
    assert outputs.result.audit["peer_comparison_scope"] == "price_only"
    assert "horizontal comparison is limited to price fields" in outputs.result.audit["comparison_principle"]

    rows = [r for r in outputs.result.rows if r.candidate is not None]
    assert {r.bidder for r in rows} == {"A", "B"}
    assert all(r.consensus_price == 200_000 for r in rows)
    assert any("chênh" in flag.lower() and "nhà thầu" in flag.lower() for r in rows for flag in r.flags)
    bidder_b = next(r for r in rows if r.bidder == "B")
    assert bidder_b.pl2_status == "CẦN THẨM ĐỊNH TƯƠNG ĐƯƠNG"

    annotated_b = load_workbook(outputs.annotated_files["B"], data_only=False)
    try:
        assert annotated_b.sheetnames[:2] == ["AI_TONG_QUAN", "AI_KIEM_TRA"]
        assert annotated_b["Điện"]["L2"].value == "=H2*D2"
        headers = [annotated_b["Điện"].cell(1, c).value for c in range(1, annotated_b["Điện"].max_column + 1)]
        assert "AI MỨC ĐỘ" in headers
        assert "AI LÝ DO" in headers
    finally:
        annotated_b.close()
