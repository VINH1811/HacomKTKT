"""Header hai tầng có ô gộp: nhãn nhóm được phép lan sang các cột con, nhưng
nhãn của MỘT cột đơn lẻ thì không.

Bối cảnh: hồ sơ PCCC/Thông gió thật có hàng nhóm gộp ở trên, tên cột con ở dưới.
Cột "ĐVT" đứng ngay trước hai cột khối lượng và KHÔNG được gộp. Trước bản vá,
chữ "ĐVT" bị lan sang hai cột đó, khiến chúng khớp luật "đơn vị tính" trước và
bị bỏ qua luật khối lượng — kết quả là cả sheet bị gắn cờ "Thiếu khối lượng"
(báo cáo thực tế: ~1.380 cảnh báo giả trên một gói thầu).
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from core.excel_reader import _flatten_header, load_workbook_items, map_columns
from core.models import DocumentRole

# Hàng nhóm: "ĐVT" là nhãn cột đơn; "THÔNG TIN VỀ VẬT LIỆU CHÍNH" là nhóm thật.
TOP = ["STT", "Hạng mục", "ĐVT", None, None,
       "THÔNG TIN VỀ VẬT LIỆU CHÍNH", None, None, "ĐƠN GIÁ", "THÀNH TIỀN"]
SUB = [None, None, None, "KL mời thầu", "KL nhà thầu chào",
       "Mô tả", "Mã hiệu", "Xuất xứ", None, None]

ROWS = [
    ("1", "Đầu báo khói địa chỉ", "Cái", 71, 71, "DI-M9102"),
    ("2", "Đầu báo khói thường", "Cái", 800, 800, "DC-M9102"),
    ("3", "Module giám sát ngõ vào", "Cái", 108, 108, "DI-M9300"),
]


def _build(path: Path, top=TOP, sub=SUB) -> Path:
    wb = Workbook(); ws = wb.active; ws.title = "PCCC A1"
    ws.append(["DỰ ÁN: TEST"]); ws.append(["GÓI THẦU: PCCC"]); ws.append([])
    ws.append(list(top))
    if sub is not None:
        ws.append(list(sub))
    ws.append([str(i) for i in range(1, len(top) + 1)])
    for stt, name, unit, klmt, klnt, code in ROWS:
        ws.append([stt, name, unit, klmt, klnt, "Quy cách", code, "VN", 100_000, klnt * 100_000])
    wb.save(path)
    return path


def _cols(top=TOP, sub=SUB) -> dict[str, list[int]]:
    flat = _flatten_header([list(top), list(sub)], len(top))
    fixed, _ = map_columns(flat, DocumentRole.HSDT)
    out: dict[str, list[int]] = {}
    for col, field in fixed.items():
        out.setdefault(field, []).append(col)
    return out


def test_standalone_label_does_not_bleed_right():
    flat = _flatten_header([list(TOP), list(SUB)], len(TOP))
    # Cột D (index 3) chỉ mang nhãn riêng của nó, KHÔNG dính "ĐVT" của cột C.
    assert "ĐVT" not in flat[3], f"Nhãn cột đơn bị lan sang cột khối lượng: {flat[3]!r}"
    assert "KL mời thầu" in flat[3]


def test_group_label_still_bleeds_to_its_sub_columns():
    flat = _flatten_header([list(TOP), list(SUB)], len(TOP))
    # Nhóm thật vẫn phải lan xuống các cột con — mapping mã hiệu/thương hiệu dựa vào nó.
    assert "THÔNG TIN VỀ VẬT LIỆU CHÍNH" in flat[6]
    assert "Mã hiệu" in flat[6]


def test_quantity_columns_not_swallowed_by_unit_rule():
    cols = _cols()
    assert cols.get("reference_quantity") == [3], "Cột KL mời thầu phải là cột D"
    assert cols.get("bid_quantity") == [4], "Cột KL nhà thầu chào phải là cột E"
    assert cols.get("unit") == [2], "Chỉ cột C mới là đơn vị tính"


def test_rows_have_quantity_end_to_end(tmp_path: Path):
    wb = load_workbook_items(_build(tmp_path / "pccc.xlsx"), DocumentRole.HSDT, bidder="NT")
    items = [i for i in wb.items if i.is_comparable]
    assert len(items) == len(ROWS)
    missing = [i.item_name for i in items if not i.bid_quantity]
    assert not missing, f"Bị coi là thiếu khối lượng dù dữ liệu có đủ: {missing}"
    assert [i.bid_quantity for i in items] == [71, 800, 108]


def test_abbreviated_quantity_labels(tmp_path: Path):
    # Header một tầng dùng viết tắt phổ biến KLMT / NT chào.
    top = ["STT", "Hạng mục", "ĐVT", "KLMT", "NT chào",
           "Mô tả", "Mã hiệu", "Xuất xứ", "Đơn giá", "Thành tiền"]
    wb = load_workbook_items(_build(tmp_path / "abbr.xlsx", top=top, sub=None),
                             DocumentRole.HSDT, bidder="NT")
    items = [i for i in wb.items if i.is_comparable]
    assert [i.bid_quantity for i in items] == [71, 800, 108]


def test_numbering_legend_row_does_not_enable_bleeding():
    """Hàng chú giải đánh số cột (1, 2, 3, "13=5%*(10+11+12)") không phải nhãn.

    Cấu trúc lấy từ hồ sơ thật gói PCCC: hàng nhóm — hàng tên cột con — hàng
    đánh số. Nếu tính hàng đánh số là "nhãn con bên dưới" thì cột nào cũng có,
    mọi nhãn đều lan sang phải và cột khối lượng lại bị mất.
    """
    top = ["Stt", "Mô tả công việc", "ĐVT", None, None, "Thông tin VTTB", None, "Đơn giá", None]
    sub = [None, None, None, " KLMT", "Nhà thầu chào", "Mô tả quy cách", "Mã hiệu", "VL chính", "CP quản lý"]
    num = ["1", "2", "3", "4", "5", "6", "7", "10", "13=5%*(10+11+12)"]
    flat = _flatten_header([top, sub, num], len(top))
    assert "ĐVT" not in flat[3], f"Hàng đánh số làm nhãn ĐVT lan sang cột KLMT: {flat[3]!r}"
    fixed, _ = map_columns(flat, DocumentRole.HSDT)
    assert fixed.get(3) == "reference_quantity"
    assert fixed.get(4) == "bid_quantity"
    assert fixed.get(8) == "price_management"


def test_bid_quantity_label_without_kl_prefix():
    # Hồ sơ thật ghi "Nhà thầu chào" trơ trọi, không có chữ "KL" phía trước.
    top = ["Stt", "Hạng mục", "ĐVT", None, None]
    sub = [None, None, None, "KLMT", "Nhà thầu chào"]
    fixed, _ = map_columns(_flatten_header([top, sub], len(top)), DocumentRole.HSDT)
    assert fixed.get(3) == "reference_quantity"
    assert fixed.get(4) == "bid_quantity"


def test_thanh_tien_klmt_not_read_as_quantity():
    # "Thành tiền KLMT" là cột tiền, không được nuốt thành cột khối lượng.
    top = ["STT", "Hạng mục", "ĐVT", None, None, "THÀNH TIỀN", None]
    sub = [None, None, None, "KL mời thầu", "KL nhà thầu chào", "Thành tiền KLMT", "Thành tiền NT chào"]
    cols = _cols(top=top, sub=sub)
    assert cols.get("reference_quantity") == [3]
    assert cols.get("bid_quantity") == [4]
    assert 5 not in cols.get("reference_quantity", []) and 5 not in cols.get("bid_quantity", [])
