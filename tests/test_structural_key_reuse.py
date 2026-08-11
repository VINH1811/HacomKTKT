"""Trùng số thứ tự không có nghĩa là cùng hạng mục.

Giữa hai phiên bản chào giá, nhà thầu chèn/xoá dòng làm STT bị dùng lại cho một
hạng mục khác hẳn. Trước đây tầng ghép theo khóa cấu trúc tin tuyệt đối vào
sheet + STT nên gán bừa, rồi báo "đổi tên, đổi mã, đổi khối lượng" thay vì
"thêm mới" — chuyên viên đọc báo cáo sẽ hiểu sai bản chất thay đổi.

Ngược lại, mã hiệu giống nhau là bằng chứng mạnh hơn tên, nên vẫn phải giữ cặp
ghép dù nhà thầu có viết lại tên hạng mục.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from core.config import EnterpriseConfig
from core.excel_reader import load_workbook_items
from core.matcher import match_items_cached
from core.models import DocumentRole, MatchKind
from core.version_compare import (
    STATUS_ADDED,
    STATUS_REMOVED,
    compare_quote_versions,
)

HEADER = ["STT", "Mã hiệu", "Tên hạng mục", "ĐVT",
          "KL nhà thầu chào", "Đơn giá tổng hợp", "Thành tiền"]


def _quote(path: Path, rows: list[tuple]) -> Path:
    wb = Workbook(); ws = wb.active; ws.title = "PCCC"
    ws.append(HEADER)
    for stt, code, name, unit, qty, price in rows:
        ws.append([stt, code, name, unit, qty, price, qty * price])
    wb.save(path)
    return path


def _config() -> EnterpriseConfig:
    cfg = EnterpriseConfig()
    cfg.enable_semantic_matching = False
    cfg.enable_reranker = False
    return cfg


def _kinds(old: Path, new: Path) -> list[tuple[str, str, MatchKind]]:
    cfg = _config()
    a = load_workbook_items(old, DocumentRole.HSDT, bidder="NT")
    b = load_workbook_items(new, DocumentRole.HSDT, bidder="NT")
    out = []
    for m in match_items_cached(a, b, cfg):
        ref = a.items[m.reference_index] if m.reference_index is not None else None
        cand = b.items[m.candidate_index] if m.candidate_index is not None else None
        out.append(((ref.item_name if ref else ""), (cand.item_name if cand else ""), m.kind))
    return out


def test_reused_stt_does_not_force_a_match(tmp_path: Path):
    old = _quote(tmp_path / "v1.xlsx", [
        ("1", "DI-M9102", "Đầu báo khói địa chỉ", "cái", 120, 500_000),
        ("2", "DI-M9102", "Đầu báo khói địa chỉ", "cái", 30, 500_000),
    ])
    # STT 2 nay là một hạng mục hoàn toàn khác.
    new = _quote(tmp_path / "v2.xlsx", [
        ("1", "DI-M9102", "Đầu báo khói địa chỉ", "cái", 120, 500_000),
        ("2", "VBD-100", "Van báo động DN100", "cái", 4, 4_462_500),
    ])
    pairs = _kinds(old, new)
    bad = [p for p in pairs if p[0] and p[1] and p[0] != p[1]]
    assert not bad, f"Không được ghép hai hạng mục khác nhau chỉ vì trùng STT: {bad}"


def test_reused_stt_reported_as_added_and_removed(tmp_path: Path):
    old = _quote(tmp_path / "v1.xlsx", [
        ("1", "DI-M9102", "Đầu báo khói địa chỉ", "cái", 120, 500_000),
        ("2", "DI-M9102", "Đầu báo khói địa chỉ", "cái", 30, 500_000),
    ])
    new = _quote(tmp_path / "v2.xlsx", [
        ("1", "DI-M9102", "Đầu báo khói địa chỉ", "cái", 120, 500_000),
        ("2", "VBD-100", "Van báo động DN100", "cái", 4, 4_462_500),
    ])
    result = compare_quote_versions(old, new, "NT", config=_config())
    by_status = {r.status: [x.item_name for x in result.rows if x.status == r.status]
                 for r in result.rows}
    assert "Van báo động DN100" in by_status.get(STATUS_ADDED, [])
    assert "Đầu báo khói địa chỉ" in by_status.get(STATUS_REMOVED, [])


def test_same_code_still_matches_despite_renaming(tmp_path: Path):
    # Mã hiệu là bằng chứng mạnh hơn tên: đổi cách viết tên vẫn phải ghép được.
    old = _quote(tmp_path / "v1.xlsx", [
        ("1", "FC-10L", "Tủ trung tâm báo cháy 10 loop", "bộ", 1, 40_000_000),
    ])
    new = _quote(tmp_path / "v2.xlsx", [
        ("1", "FC-10L", "Trung tâm BC địa chỉ 10 vòng lặp GST", "bộ", 1, 42_000_000),
    ])
    pairs = _kinds(old, new)
    matched = [p for p in pairs if p[0] and p[1]]
    assert len(matched) == 1, "Cùng mã hiệu thì đổi tên vẫn phải ghép được"


def test_identical_names_still_match_on_structure(tmp_path: Path):
    rows = [("1", "", "Nhân công thi công và vật tư phụ (đường ống)", "m", 100, 50_000)]
    pairs = _kinds(_quote(tmp_path / "v1.xlsx", rows), _quote(tmp_path / "v2.xlsx", rows))
    assert [p for p in pairs if p[0] and p[1]], "Tên giống hệt phải giữ nguyên cặp ghép"


def test_type_conflict_blocks_structural_match(tmp_path: Path):
    # Cùng STT, tên na ná nhưng khác chủng loại vật tư — đúng lỗi ban KTKT nêu.
    old = _quote(tmp_path / "v1.xlsx", [
        ("1", "ST-100", "Ống thép tráng kẽm DN100 hệ chữa cháy", "m", 350, 393_600),
    ])
    new = _quote(tmp_path / "v2.xlsx", [
        ("1", "OLD-100", "Ống nhựa luồn dây điện DN100 hệ điện", "m", 350, 80_000),
    ])
    bad = [p for p in _kinds(old, new) if p[0] and p[1]]
    assert not bad, f"Khác chủng loại thì không được ghép: {bad}"
