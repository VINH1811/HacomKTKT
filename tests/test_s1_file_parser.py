"""S1 - Test đọc file Excel cơ bản.

Các test dùng workbook nhỏ tạo tại tmp_path nên chạy nhanh, độc lập và không
làm thay đổi dữ liệu thật của dự án.
"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from core.config import EnterpriseConfig
from core.excel_io import read_workbook_matrices, scan_xlsx_issues
from core.excel_reader import detect_header, load_workbook_items, map_columns
from core.models import DocumentRole, RowType
from core.parallel import WorkbookLoadSpec, load_workbooks_parallel


def _cfg() -> EnterpriseConfig:
    cfg = EnterpriseConfig()
    cfg.enable_semantic_matching = False
    cfg.enable_reranker = False
    cfg.excel_read_engine = "calamine"
    cfg.excel_read_workers = 4
    cfg.excel_scan_formulas = True
    cfg.excel_scan_external_links = True
    return cfg


def _simple_book(path: Path, *, price: float = 100.0, formula_error: bool = False) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Điện"
    ws.append(["Mã hiệu", "Tên hạng mục", "ĐVT", "Khối lượng nhà thầu", "Đơn giá tổng hợp", "Thành tiền nhà thầu"])
    ws.append(["M-01", "Cáp điện Cu/XLPE 4x10", "m", 10, price, 10 * price])
    if formula_error:
        ws["G2"] = "=#REF!"
    wb.save(path)


def test_s1_fi01_parse_valid_xlsx_and_preserve_source_row(tmp_path: Path):
    path = tmp_path / "valid.xlsx"
    _simple_book(path)

    workbook = load_workbook_items(path, DocumentRole.HSDT, bidder="NT A", read_engine="calamine")

    assert workbook.read_engine in {"calamine", "openpyxl"}
    assert len(workbook.items) == 1
    item = workbook.items[0]
    assert item.sheet == "Điện"
    assert item.row_number == 2
    assert item.item_code == "M-01"
    assert item.item_name == "Cáp điện Cu/XLPE 4x10"
    assert item.bid_quantity == 10
    assert item.unit_price_total == 100
    assert item.bid_amount == 1_000


def test_s1_fi02_detect_multi_level_header():
    rows = [
        ["BÁO GIÁ", None, None, None, None, None],
        ["Thông tin công việc", None, None, "Khối lượng", "Đơn giá", "Thành tiền"],
        ["Mã hiệu", "Tên hạng mục", "ĐVT", "Nhà thầu", "Tổng hợp", "Nhà thầu"],
        ["M-01", "Tủ điện tổng", "Tủ", 1, 1_000_000, 1_000_000],
    ]

    start, end, flat = detect_header(rows, max_header_depth=3)
    fixed, _technical = map_columns(flat, DocumentRole.HSDT)

    assert start == 1
    assert end == 2
    assert "item_code" in fixed.values()
    assert "item_name" in fixed.values()
    assert "unit" in fixed.values()
    assert "bid_quantity" in fixed.values()
    assert "unit_price_total" in fixed.values()
    assert "bid_amount" in fixed.values()


def test_s1_fi03_calamine_matrix_preserves_raw_rows(tmp_path: Path):
    path = tmp_path / "matrix.xlsx"
    _simple_book(path)

    matrices = read_workbook_matrices(path, engine="calamine")

    assert matrices.engine in {"calamine", "openpyxl"}
    assert len(matrices.sheets) == 1
    assert matrices.sheets[0].name == "Điện"
    assert matrices.sheets[0].rows[1][0] == "M-01"


def test_s1_fi04_formula_ref_is_detected(tmp_path: Path):
    path = tmp_path / "broken.xlsx"
    _simple_book(path, formula_error=True)

    scan = scan_xlsx_issues(path)
    workbook = load_workbook_items(path, DocumentRole.HSDT, bidder="NT A", read_engine="calamine")

    assert any(issue.kind == "FORMULA_ERROR" and issue.cell == "G2" for issue in scan.issues)
    assert any(issue["cell"] == "G2" for issue in workbook.formula_issues)
    assert any("#REF" in warning.upper() for warning in workbook.warnings)


def test_s1_fi05_four_workbooks_are_loaded_in_parallel(tmp_path: Path):
    specs = []
    for index in range(4):
        path = tmp_path / f"bidder_{index}.xlsx"
        _simple_book(path, price=100 + index)
        specs.append(WorkbookLoadSpec(
            key=str(index),
            path=path,
            role=DocumentRole.HSDT,
            bidder=f"NT{index}",
        ))

    result = load_workbooks_parallel(specs, _cfg())

    assert set(result) == {"0", "1", "2", "3"}
    assert all(book.read_engine in {"calamine", "openpyxl"} for book in result.values())
    assert [result[str(i)].items[0].unit_price_total for i in range(4)] == [100, 101, 102, 103]


def test_s1_fi06_invalid_extension_is_rejected(tmp_path: Path):
    path = tmp_path / "invalid.xls"
    path.write_bytes(b"not an xlsx workbook")

    with pytest.raises(ValueError, match=r"\.xlsx"):
        load_workbook_items(path, DocumentRole.HSDT, bidder="NT A")


def test_s1_fi07_corrupt_xlsx_returns_clear_error(tmp_path: Path):
    path = tmp_path / "corrupt.xlsx"
    path.write_bytes(b"this is not a zip based excel workbook")

    with pytest.raises(Exception):
        load_workbook_items(path, DocumentRole.HSDT, bidder="NT A", read_engine="calamine")


def test_s1_fi08_duplicate_code_rows_are_preserved_and_flagged(tmp_path: Path):
    path = tmp_path / "duplicate.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Điện"
    ws.append(["Mã hiệu", "Tên hạng mục", "ĐVT", "Khối lượng", "Đơn giá", "Thành tiền"])
    ws.append(["M-01", "Tủ điện tổng", "Tủ", 1, 1_000, 1_000])
    ws.append(["M-01", "Tủ điện nhánh", "Tủ", 1, 900, 900])
    wb.save(path)

    parsed = load_workbook_items(path, DocumentRole.HSDT, bidder="NT A")
    comparable = [item for item in parsed.items if item.is_comparable]

    assert len(comparable) == 2
    assert all(item.normalized_code == "M-01" for item in comparable)
    assert any("mã hiệu trùng" in flag.lower() for item in comparable for flag in item.data_quality_flags)


def test_s1_fi09_component_without_price_is_not_false_error(tmp_path: Path):
    path = tmp_path / "components.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Điện"
    ws.append(["STT", "Mã hiệu", "Tên hạng mục", "ĐVT", "Khối lượng", "Đơn giá", "Thành tiền"])
    ws.append(["1", "TU-01", "Tủ điện tổng", "Tủ", 1, 1_000_000, 1_000_000])
    ws.append([None, "ACB-01", "ACB 4P 3200A", "Cái", 2, None, None])
    wb.save(path)

    parsed = load_workbook_items(path, DocumentRole.HSDT, bidder="NT A")
    component = next(item for item in parsed.items if item.row_type is RowType.COMPONENT)

    assert "Thiếu đơn giá tổng hợp" not in component.data_quality_flags
    assert "Thiếu thành tiền" not in component.data_quality_flags
