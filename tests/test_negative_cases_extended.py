"""Negative test cases bổ sung - kiểm tra toàn diện các đường lỗi của hệ thống.

Mục tiêu: với mỗi input "xấu" hợp lý (file rỗng, sai định dạng, thiếu cột,
thiếu phụ lục, upload vượt giới hạn, tên file độc hại...), hệ thống phải:
  1. Không crash với traceback lộ ra ngoài (AttributeError/TypeError không kiểm soát),
  2. Trả lỗi rõ ràng bằng tiếng Việt HOẶC log warning có thể truy vết,
  3. Không tạo ra báo cáo "trông như đúng" từ dữ liệu rác.

Các test dùng workbook nhỏ tạo tại tmp_path, không phụ thuộc dữ liệu thật.
"""

from __future__ import annotations

import asyncio
import io
import zipfile
from pathlib import Path
from typing import Any

import pytest
from openpyxl import Workbook

from app import _sanitize, _save_upload, _validate_bidder_uploads, format_job_error_message
from core.config import EnterpriseConfig
from core.excel_io import read_workbook_matrices
from core.excel_reader import load_workbook_items
from core.models import DocumentRole
from core.number_parser import math_error, parse_number
from core.pl2_reader import load_pl2_requirements
from core.tender_package import compare_appendices_with_bidders


def _cfg() -> EnterpriseConfig:
    cfg = EnterpriseConfig()
    cfg.enable_semantic_matching = False
    cfg.enable_reranker = False
    cfg.excel_read_engine = "calamine"
    cfg.excel_read_workers = 2
    return cfg


def _valid_boq(path: Path, *, rows: int = 1) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "BOQ"
    ws.append(["STT", "Mã hiệu", "Tên hạng mục", "ĐVT", "Khối lượng", "Đơn giá tổng hợp", "Thành tiền"])
    for i in range(rows):
        ws.append([str(i + 1), f"M-{i:02d}", f"Hạng mục {i}", "Cái", 1, 1_000, 1_000])
    wb.save(path)


# ---------------------------------------------------------------------------
# A. File giả mạo / sai định dạng ở tầng đọc Excel
# ---------------------------------------------------------------------------

class TestFakeAndCorruptFiles:
    def test_txt_renamed_to_xlsx_is_rejected_not_silently_parsed(self, tmp_path: Path):
        path = tmp_path / "fake.xlsx"
        path.write_text("nội dung linh tinh không liên quan gì đến excel cả", encoding="utf-8")

        with pytest.raises(Exception):
            load_workbook_items(path, DocumentRole.HSDT, bidder="NT giả mạo")

    def test_zero_byte_file_with_xlsx_extension_is_rejected(self, tmp_path: Path):
        path = tmp_path / "empty.xlsx"
        path.write_bytes(b"")

        with pytest.raises(Exception):
            load_workbook_items(path, DocumentRole.HSDT, bidder="NT rỗng")

    def test_uppercase_extension_is_still_accepted(self, tmp_path: Path):
        # Hệ thống dùng .suffix.lower() nên .XLSX phải được nhận diện như .xlsx
        path = tmp_path / "valid.XLSX"
        _valid_boq(path)

        workbook = load_workbook_items(path, DocumentRole.HSDT, bidder="NT hoa")
        assert len(workbook.items) == 1

    def test_xls_extension_gives_clear_vietnamese_error(self, tmp_path: Path):
        path = tmp_path / "old.xls"
        path.write_bytes(b"not a real xls file")

        with pytest.raises(ValueError, match=r"\.xlsx"):
            load_workbook_items(path, DocumentRole.HSDT, bidder="NT")

    def test_zip_bomb_style_wrong_internal_structure_does_not_crash_silently(self, tmp_path: Path):
        # File .xlsx hợp lệ về mặt ZIP nhưng KHÔNG có cấu trúc OOXML bên trong.
        path = tmp_path / "wrong_zip.xlsx"
        with zipfile.ZipFile(path, "w") as zf:
            zf.writestr("hello.txt", "đây không phải là một workbook excel")

        with pytest.raises(Exception):
            load_workbook_items(path, DocumentRole.HSDT, bidder="NT zip lạ")

    def test_truncated_xlsx_raises_clear_error_not_garbage_data(self, tmp_path: Path):
        # Tạo file hợp lệ rồi cắt cụt để giả lập upload bị đứt giữa đường.
        good = tmp_path / "good.xlsx"
        _valid_boq(good)
        truncated = tmp_path / "truncated.xlsx"
        original_bytes = good.read_bytes()
        truncated.write_bytes(original_bytes[: len(original_bytes) // 2])

        with pytest.raises(Exception):
            load_workbook_items(truncated, DocumentRole.HSDT, bidder="NT đứt file")


# ---------------------------------------------------------------------------
# B. File hợp lệ về định dạng nhưng vô nghĩa về nội dung
# ---------------------------------------------------------------------------

class TestMeaninglessButValidWorkbooks:
    def test_completely_empty_sheet_does_not_crash(self, tmp_path: Path):
        path = tmp_path / "blank.xlsx"
        wb = Workbook()
        wb.active.title = "Sheet1"
        wb.save(path)

        workbook = load_workbook_items(path, DocumentRole.HSDT, bidder="NT trống")
        assert workbook.items == []
        assert any("không đọc được hạng mục" in w.lower() for w in workbook.warnings)

    def test_random_text_without_any_header_keyword_yields_no_items_not_garbage(self, tmp_path: Path):
        path = tmp_path / "random.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["con mèo", "con chó", "hôm nay trời đẹp", 42, "lung tung beng"])
        ws.append(["abc", "def", "ghi", 7, "xyz"])
        wb.save(path)

        workbook = load_workbook_items(path, DocumentRole.HSDT, bidder="NT lung tung")
        # Không nhận diện được header hợp lệ -> không được bịa ra hạng mục giả
        assert workbook.items == []

    def test_header_only_no_data_rows(self, tmp_path: Path):
        path = tmp_path / "header_only.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "BOQ"
        ws.append(["STT", "Mã hiệu", "Tên hạng mục", "ĐVT", "Khối lượng", "Đơn giá tổng hợp", "Thành tiền"])
        wb.save(path)

        workbook = load_workbook_items(path, DocumentRole.HSDT, bidder="NT chỉ có header")
        assert workbook.items == []


# ---------------------------------------------------------------------------
# C. number_parser - dữ liệu số bất thường
# ---------------------------------------------------------------------------

class TestNumberParserNegative:
    @pytest.mark.parametrize("raw", ["lung tung beng", "abc123xyz", "====", "N/A", "-", None, ""])
    def test_garbage_returns_none_not_exception(self, raw):
        assert parse_number(raw) is None

    def test_math_error_detects_klxdg_mismatch(self):
        # 10 x 1000 = 10000, nhưng cột "thành tiền" ghi 5000 -> sai lệch phải bị phát hiện
        err = math_error(10, 1000, 5000)
        assert err is not None and err > 0

    def test_math_error_none_when_any_input_missing(self):
        assert math_error(None, 1000, 5000) is None
        assert math_error(10, None, 5000) is None
        assert math_error(10, 1000, None) is None


# ---------------------------------------------------------------------------
# D. tender_package - luồng so sánh phụ lục / nhà thầu
# ---------------------------------------------------------------------------

class TestTenderPackageNegative:
    def test_no_bidder_files_raises_value_error(self, tmp_path: Path):
        pl1 = tmp_path / "pl1.xlsx"
        _valid_boq(pl1)

        with pytest.raises(ValueError, match=r"ít nhất 1 hồ sơ"):
            compare_appendices_with_bidders(
                bidder_files=[],
                output_dir=tmp_path / "out",
                pl1_path=pl1,
                pl2_path=None,
                config=_cfg(),
            )

    def test_no_appendix_at_all_raises_value_error(self, tmp_path: Path):
        bidder = tmp_path / "bidder.xlsx"
        _valid_boq(bidder)

        with pytest.raises(ValueError, match=r"ít nhất một phụ lục"):
            compare_appendices_with_bidders(
                bidder_files=[("NT A", bidder)],
                output_dir=tmp_path / "out",
                pl1_path=None,
                pl2_path=None,
                config=_cfg(),
            )

    def test_missing_pl2_file_on_disk_raises_clear_error(self, tmp_path: Path):
        bidder = tmp_path / "bidder.xlsx"
        _valid_boq(bidder)
        missing_pl2 = tmp_path / "khong_ton_tai.xlsx"

        with pytest.raises(RuntimeError, match=r"PHỤ LỤC 02"):
            compare_appendices_with_bidders(
                bidder_files=[("NT A", bidder)],
                output_dir=tmp_path / "out",
                pl1_path=None,
                pl2_path=missing_pl2,
                config=_cfg(),
            )

    def test_pl2_without_recognizable_headers_does_not_crash_but_warns(self, tmp_path: Path):
        # PL02 "giả" không có cột Vật tư/Thương hiệu/Xuất xứ.
        pl2 = tmp_path / "pl2_fake.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["Cột A", "Cột B", "Cột C"])
        ws.append(["1", "2", "3"])
        wb.save(pl2)

        bidder = tmp_path / "bidder.xlsx"
        _valid_boq(bidder)

        out = compare_appendices_with_bidders(
            bidder_files=[("NT A", bidder)],
            output_dir=tmp_path / "out",
            pl1_path=None,
            pl2_path=pl2,
            config=_cfg(),
        )
        # Không crash, báo cáo vẫn được tạo, nhưng phải có cảnh báo rõ ràng
        # rằng PL02 không đọc được yêu cầu nào - không được "im lặng" coi như đạt.
        assert out.report_path.exists()
        assert any("phụ lục 02" in w.lower() for w in out.result.warnings)

    def test_single_bidder_with_zero_items_raises_clear_no_boq_error(self, tmp_path: Path):
        # A-05: hồ sơ nhà thầu không có bảng khối lượng nào (sheet trống) phải báo
        # lỗi rõ ràng, không tạo báo cáo rỗng "trông như đúng".
        from core.models import UserFacingError

        pl1 = tmp_path / "pl1.xlsx"
        _valid_boq(pl1, rows=2)

        empty_bidder = tmp_path / "bidder_empty.xlsx"
        wb = Workbook()
        wb.active.title = "Sheet1"
        wb.save(empty_bidder)

        with pytest.raises(UserFacingError) as excinfo:
            compare_appendices_with_bidders(
                bidder_files=[("NT trống", empty_bidder)],
                output_dir=tmp_path / "out",
                pl1_path=pl1,
                pl2_path=None,
                config=_cfg(),
            )
        msg = str(excinfo.value)
        assert "NT trống" in msg
        assert "BOQ" in msg or "khối lượng" in msg

    def test_bidder_file_without_boq_sheet_raises_clear_error(self, tmp_path: Path):
        # A-05: file chỉ có sheet "Bìa", "Phụ lục" (không có bảng khối lượng) ->
        # phải báo lỗi rõ ràng nêu đúng tên nhà thầu.
        from core.models import UserFacingError

        pl1 = tmp_path / "pl1.xlsx"
        _valid_boq(pl1, rows=2)

        no_boq = tmp_path / "no_boq.xlsx"
        wb = Workbook()
        wb.active.title = "Bìa"
        wb["Bìa"]["A1"] = "TRANG BÌA HỒ SƠ"
        ws2 = wb.create_sheet("Phụ lục")
        ws2["A1"] = "Một vài ghi chú không phải bảng khối lượng"
        wb.save(no_boq)

        with pytest.raises(UserFacingError) as excinfo:
            compare_appendices_with_bidders(
                bidder_files=[("Nhà thầu X", no_boq)],
                output_dir=tmp_path / "out",
                pl1_path=pl1,
                pl2_path=None,
                config=_cfg(),
            )
        assert "Nhà thầu X" in str(excinfo.value)

    def test_single_bidder_disables_peer_price_comparison(self, tmp_path: Path):
        pl1 = tmp_path / "pl1.xlsx"
        _valid_boq(pl1, rows=1)
        bidder = tmp_path / "bidder.xlsx"
        _valid_boq(bidder, rows=1)

        out = compare_appendices_with_bidders(
            bidder_files=[("NT duy nhất", bidder)],
            output_dir=tmp_path / "out",
            pl1_path=pl1,
            pl2_path=None,
            config=_cfg(),
        )
        assert out.result.audit["peer_price_comparison_enabled"] is False
        assert out.result.audit["peer_stats"]["reason"]


# ---------------------------------------------------------------------------
# E. pl2_reader trực tiếp
# ---------------------------------------------------------------------------

class TestPL2ReaderNegative:
    def test_pl2_wrong_extension_raises(self, tmp_path: Path):
        path = tmp_path / "pl2.xls"
        path.write_bytes(b"not real")
        with pytest.raises(ValueError, match=r"\.xlsx"):
            load_pl2_requirements(path)

    def test_pl2_empty_workbook_returns_empty_with_warning(self, tmp_path: Path):
        path = tmp_path / "pl2_empty.xlsx"
        wb = Workbook()
        wb.active.title = "Sheet1"
        wb.save(path)

        requirements, warnings = load_pl2_requirements(path, config=_cfg())
        assert requirements == []
        assert any("không đọc được yêu cầu" in w.lower() for w in warnings)


# ---------------------------------------------------------------------------
# F. app.py - tầng upload / xử lý lỗi cho người dùng
# ---------------------------------------------------------------------------

class _FakeUploadFile:
    """Giả lập tối thiểu giao diện UploadFile của FastAPI cho test thuần asyncio."""

    def __init__(self, content: bytes, filename: str = "test.xlsx"):
        self._buffer = io.BytesIO(content)
        self.filename = filename
        self._closed = False

    async def read(self, size: int = -1) -> bytes:
        return self._buffer.read(size)

    async def close(self) -> None:
        self._closed = True


class TestUploadValidationNegative:
    def test_sanitize_strips_path_traversal(self):
        result = _sanitize("../../../etc/passwd.xlsx", "fallback.xlsx")
        assert ".." not in result
        assert "/" not in result and "\\" not in result

    def test_sanitize_strips_path_traversal_windows_style(self):
        # Trên Linux, "\\" không phải path separator nên Path(...).name không tự
        # cắt được phần thư mục kiểu Windows; "_sanitize" chỉ cần đảm bảo kết quả
        # không còn ký tự separator nào (không thể tách thành nhiều segment khi
        # ghép vào một path), không cần xoá literal ".." vì khi không còn dấu
        # gạch chéo, ".." chỉ là dấu chấm vô hại nằm trong một tên file duy nhất.
        result = _sanitize("..\\..\\Windows\\System32\\evil.xlsx", "fallback.xlsx")
        assert "\\" not in result
        assert "/" not in result
        # Ghép vào một thư mục gốc rồi resolve phải vẫn nằm trong thư mục đó.
        base = Path("/safe/uploads").resolve()
        joined = (base / result).resolve()
        assert joined.parent == base

    def test_sanitize_empty_name_falls_back(self):
        assert _sanitize("", "fallback.xlsx") == "fallback.xlsx"

    def test_save_upload_rejects_disallowed_extension(self, tmp_path: Path):
        upload = _FakeUploadFile(b"hello", filename="evil.exe")
        target = tmp_path / "evil.exe"

        from fastapi import HTTPException

        with pytest.raises(HTTPException) as excinfo:
            asyncio.run(_save_upload(upload, target, limit_bytes=1024, allowed_suffixes={".xlsx"}))
        assert excinfo.value.status_code == 400

    def test_save_upload_rejects_empty_file(self, tmp_path: Path):
        upload = _FakeUploadFile(b"", filename="empty.xlsx")
        target = tmp_path / "empty.xlsx"

        from fastapi import HTTPException

        with pytest.raises(HTTPException) as excinfo:
            asyncio.run(_save_upload(upload, target, limit_bytes=1024, allowed_suffixes={".xlsx"}))
        assert excinfo.value.status_code == 400
        assert not target.exists()

    def test_save_upload_rejects_file_over_limit(self, tmp_path: Path):
        big_content = b"x" * 2048
        upload = _FakeUploadFile(big_content, filename="big.xlsx")
        target = tmp_path / "big.xlsx"

        from fastapi import HTTPException

        with pytest.raises(HTTPException) as excinfo:
            asyncio.run(_save_upload(upload, target, limit_bytes=1024, allowed_suffixes={".xlsx"}))
        assert excinfo.value.status_code == 413
        # File tạm phải bị xoá, không để lại rác chiếm dung lượng đĩa.
        assert not target.exists()

    def test_a06_start_without_any_bidder_file_is_rejected(self):
        # A-06: bấm Bắt đầu khi chưa chọn file nhà thầu nào -> backend từ chối rõ ràng.
        from fastapi import HTTPException

        with pytest.raises(HTTPException) as excinfo:
            _validate_bidder_uploads(files=[], bidder_names=[], minimum=1)
        assert excinfo.value.status_code == 400
        assert "ít nhất" in str(excinfo.value.detail).lower()

    def test_a06_mismatched_files_and_names_is_rejected(self):
        from fastapi import HTTPException

        # Có 1 file nhưng 0 tên nhà thầu (lệch nhau) -> cũng phải bị từ chối.
        with pytest.raises(HTTPException) as excinfo:
            _validate_bidder_uploads(files=[object()], bidder_names=[], minimum=1)
        assert excinfo.value.status_code == 400

    def test_save_upload_accepts_valid_small_file(self, tmp_path: Path):
        upload = _FakeUploadFile(b"valid content", filename="ok.xlsx")
        target = tmp_path / "ok.xlsx"

        asyncio.run(_save_upload(upload, target, limit_bytes=1024, allowed_suffixes={".xlsx"}))
        assert target.exists()
        assert target.read_bytes() == b"valid content"

    def test_save_upload_rejects_xlsb_even_when_target_name_forced_xlsx(self, tmp_path: Path):
        # Ô Phụ lục 01/02 và HSMT ép tên target thành .xlsx; phải kiểm tra theo
        # ĐUÔI FILE GỐC (.xlsb) chứ không phải tên target.
        from fastapi import HTTPException

        upload = _FakeUploadFile(b"binary xlsb content", filename="Bao_gia_NT.xlsb")
        target = tmp_path / "000_PHU_LUC_01.xlsx"  # target bị ép .xlsx

        with pytest.raises(HTTPException) as excinfo:
            asyncio.run(_save_upload(upload, target, limit_bytes=10_000, allowed_suffixes={".xlsx"}))
        assert excinfo.value.status_code == 400
        assert not target.exists()

    def test_save_upload_rejects_xls_into_appendix_slot(self, tmp_path: Path):
        from fastapi import HTTPException

        upload = _FakeUploadFile(b"old binary xls", filename="KLMT.xls")
        target = tmp_path / "001_PHU_LUC_02.xlsx"

        with pytest.raises(HTTPException) as excinfo:
            asyncio.run(_save_upload(upload, target, limit_bytes=10_000, allowed_suffixes={".xlsx"}))
        assert excinfo.value.status_code == 400
        assert not target.exists()

    def test_save_upload_empty_filename_falls_back_to_target_suffix(self, tmp_path: Path):
        # Một số client gửi filename rỗng: khi đó dựa vào đuôi target để giữ
        # tương thích, không chặn nhầm upload hợp lệ.
        upload = _FakeUploadFile(b"content", filename="")
        target = tmp_path / "000_HSMT.xlsx"

        asyncio.run(_save_upload(upload, target, limit_bytes=10_000, allowed_suffixes={".xlsx"}))
        assert target.exists()


class TestErrorMessageMapping:
    def test_encrypted_or_unknown_underlying_error_falls_back_to_generic_excel_message(self):
        exc = RuntimeError(
            "Không đọc được file '004_NhaThauC.xlsx' (Nhà thầu C): "
            "RuntimeError: Workbook is encrypted and password-protected"
        )
        request = {
            "bidders": [{"file": "004_NhaThauC.xlsx", "original_name": "BaoGia_C_CoMatKhau.xlsx"}]
        }
        msg = format_job_error_message(exc, request)
        assert "BaoGia_C_CoMatKhau.xlsx" in msg
        assert "không đúng định dạng Excel" in msg

    def test_unknown_request_mapping_still_returns_something_safe(self):
        exc = RuntimeError("Không đọc được file 'unknown.xlsx' (???): ValueError: lỗi lạ")
        msg = format_job_error_message(exc, None)
        assert msg  # phải luôn trả một chuỗi, không None / không raise
        assert "unknown.xlsx" in msg


# Chữ ký file giả lập để test chẩn đoán (không cần file Excel thật).
_OLE2 = b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1"
_ENC_MARKER = b"E\x00n\x00c\x00r\x00y\x00p\x00t\x00e\x00d\x00P\x00a\x00c\x00k\x00a\x00g\x00e"


class TestFileSignatureDiagnosis:
    """Mỗi loại lỗi file phải có thông báo cụ thể, dễ hiểu cho người dùng."""

    def _msg_for(self, tmp_path: Path, target_name: str, content: bytes, original: str) -> str:
        (tmp_path / target_name).write_bytes(content)
        exc = RuntimeError(f"Không đọc được file '{target_name}' (Nhà thầu A): BadZipFile: not a zip")
        request = {"bidders": [{"file": target_name, "original_name": original}]}
        return format_job_error_message(exc, request, folder=tmp_path)

    def test_password_protected_file_says_unlock(self, tmp_path: Path):
        content = _OLE2 + b"\x00" * 200 + _ENC_MARKER + b"\x00" * 50
        msg = self._msg_for(tmp_path, "002_x.xlsx", content, "BaoGia_CoMatKhau.xlsx")
        assert "BaoGia_CoMatKhau.xlsx" in msg
        assert "mật khẩu" in msg.lower()
        # KHÔNG còn báo nhầm là "không phải file Excel".
        assert "không phải là file excel" not in msg.lower()

    def test_old_xls_binary_says_save_as_xlsx(self, tmp_path: Path):
        content = _OLE2 + b"Workbook" + b"\x00" * 100  # OLE2 nhưng không có marker mã hóa
        msg = self._msg_for(tmp_path, "002_x.xlsx", content, "KLMT_cu.xls")
        assert "KLMT_cu.xls" in msg
        assert "excel cũ" in msg.lower() or ".xlsx" in msg.lower()
        assert "mật khẩu" not in msg.lower()

    def test_image_renamed_to_xlsx_says_image(self, tmp_path: Path):
        content = b"\x89PNG\r\n\x1a\n" + b"\x00" * 50
        msg = self._msg_for(tmp_path, "002_x.xlsx", content, "anh_chup.xlsx")
        assert "ảnh" in msg.lower()

    def test_pdf_renamed_to_xlsx_says_pdf(self, tmp_path: Path):
        content = b"%PDF-1.7\n" + b"\x00" * 50
        msg = self._msg_for(tmp_path, "002_x.xlsx", content, "scan.xlsx")
        assert "pdf" in msg.lower()

    def test_corrupt_zip_says_corrupt(self, tmp_path: Path):
        content = b"PK\x03\x04" + b"garbage not a real xlsx"
        msg = self._msg_for(tmp_path, "002_x.xlsx", content, "hong.xlsx")
        assert "hỏng" in msg.lower() or "sai cấu trúc" in msg.lower()

    def test_diagnose_returns_none_for_unreadable_path(self, tmp_path: Path):
        from core.excel_io import diagnose_excel_file
        assert diagnose_excel_file(tmp_path / "khong_ton_tai.xlsx") is None


class TestUnknownErrorsAreAlwaysHandled:
    """Bất kỳ lỗi nào ngoài danh sách đã biết cũng phải ra thông báo hợp lệ,
    không raise thêm, không trả None — đảm bảo người dùng luôn nhận phản hồi."""

    @pytest.mark.parametrize("exc", [
        MemoryError("hết bộ nhớ"),
        OSError("ổ đĩa lỗi"),
        ValueError("một lỗi rất lạ chưa từng gặp"),
        RuntimeError("lỗi không theo định dạng nào"),
        Exception(""),                       # message rỗng
        RecursionError("đệ quy quá sâu"),
        ZeroDivisionError("division by zero"),
    ])
    def test_any_exception_returns_nonempty_friendly_string(self, exc):
        for request in (None, {}, {"bidders": [{"file": "x.xlsx", "original_name": "A.xlsx"}]}):
            msg = format_job_error_message(exc, request)
            assert isinstance(msg, str) and msg.strip()

    def test_file_error_with_unrecognized_signature_still_has_message(self, tmp_path: Path):
        # File có chữ ký lạ (không phải zip/ole2/ảnh/pdf) -> vẫn ra thông báo rõ.
        (tmp_path / "002_x.xlsx").write_bytes(b"\x01\x02\x03 noi dung la khong xac dinh")
        exc = RuntimeError("Không đọc được file '002_x.xlsx' (NT A): Exception: la")
        request = {"bidders": [{"file": "002_x.xlsx", "original_name": "Bao_gia.xlsx"}]}
        msg = format_job_error_message(exc, request, folder=tmp_path)
        assert "Bao_gia.xlsx" in msg
        assert "không phải là file excel" in msg.lower()
