"""Kiểm tra nhất quán nội bộ: cùng hạng mục trong MỘT hồ sơ chào giá nhưng được
chào nhiều đơn giá khác nhau.

Với hồ sơ chỉ có một nhà thầu, đây là phép kiểm tra giá duy nhất còn dùng được
vì không có nhà thầu nào khác để so ngang.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from core.excel_reader import load_workbook_items
from core.internal_consistency import (
    annotate_price_inconsistencies,
    find_price_inconsistencies,
)
from core.models import DocumentRole, ItemRecord, RowType, Severity
from core.text_normalizer import normalize_code, normalize_name, normalize_unit

HEADERS = ["STT", "Mã hiệu", "Tên hạng mục", "ĐVT", "KL nhà thầu chào",
           "Đơn giá tổng hợp", "Thành tiền"]


def _item(name, unit="cái", code="", price=100.0, sheet="S1", row=1, path="") -> ItemRecord:
    return ItemRecord(
        source_id="x", role=DocumentRole.HSDT, bidder="b", workbook="w", sheet=sheet,
        row_number=row, item_name=name, unit=unit, item_code=code,
        unit_price_total=price, row_type=RowType.DETAIL,
        normalized_name=normalize_name(name), normalized_unit=normalize_unit(unit),
        normalized_code=normalize_code(code), normalized_path=path,
    )


def test_same_code_and_name_different_price_is_flagged():
    items = [
        _item("Đầu báo khói địa chỉ", code="DI-M9102", price=770_440, row=10),
        _item("Đầu báo khói địa chỉ", code="DI-M9102", price=850_000, row=20),
    ]
    issues = find_price_inconsistencies(items)
    assert len(issues) == 1
    assert issues[0].min_price == 770_440 and issues[0].max_price == 850_000
    assert issues[0].same_sheet is True
    assert issues[0].severity is Severity.WARNING


def test_identical_price_is_not_flagged():
    items = [
        _item("Đầu sprinkler", code="FRD-012", price=235_113, row=10),
        _item("Đầu sprinkler", code="FRD-012", price=235_113, row=20),
    ]
    assert find_price_inconsistencies(items) == []


def test_rounding_difference_is_tolerated():
    items = [
        _item("Ống thép DN100", code="ST-100", price=1_000_000, row=10),
        _item("Ống thép DN100", code="ST-100", price=1_002_000, row=20),   # lệch 0,2%
    ]
    assert find_price_inconsistencies(items) == []


def test_cross_sheet_is_review_not_warning():
    items = [
        _item("Đèn chỉ thị báo cháy", code="C-9314P", price=669_885, sheet="PCCC A1", row=5),
        _item("Đèn chỉ thị báo cháy", code="C-9314P", price=720_000, sheet="PCCC B1", row=5),
    ]
    issue = find_price_inconsistencies(items)[0]
    assert issue.same_sheet is False
    assert issue.severity is Severity.REVIEW
    assert len(issue.sheets) == 2


def test_lump_sum_unit_is_ignored():
    # Đơn vị "Lô" là trọn gói: mỗi dòng một phạm vi riêng, khác giá là bình thường.
    items = [
        _item("Vật tư phụ theo ống", unit="Lô", price=4_349_826, row=10),
        _item("Vật tư phụ theo ống", unit="Lô", price=505_387_183, row=20),
    ]
    assert find_price_inconsistencies(items) == []


def test_section_heading_name_is_ignored():
    # "HỆ THỐNG ĐIỆN" viết hoa toàn bộ là tiêu đề mục, không phải một món hàng.
    items = [
        _item("HỆ THỐNG ĐIỆN", price=395_709_700, row=10),
        _item("HỆ THỐNG ĐIỆN", price=85_852_544_382, row=20),
    ]
    assert find_price_inconsistencies(items) == []


def test_note_text_in_code_column_is_ignored():
    # "CĐT cấp" bị đọc nhầm vào cột mã hiệu; có dấu và khoảng trắng nên không
    # phải mã kỹ thuật, và tên cũng quá ngắn để gom nhóm.
    items = [
        _item("CĐT cấp", code="CĐT cấp", price=45_600, row=10),
        _item("CĐT cấp", code="CĐT cấp", price=4_200_000, row=20),
    ]
    assert find_price_inconsistencies(items) == []


def test_same_code_different_spec_is_not_flagged():
    # Một mã định mức dùng chung cho nhiều quy cách (mã tôn dùng cho mọi cỡ ống
    # gió) — khác quy cách thì khác giá là đúng.
    items = [
        _item("KT 200x200", code="Z08-1.15", price=318_700, row=10),
        _item("KT 1000x1000", code="Z08-1.15", price=2_999_300, row=20),
    ]
    assert find_price_inconsistencies(items) == []


def test_short_name_separated_by_parent_section():
    # "KT: 1500x500" chỉ có nghĩa trong ngữ cảnh mục cha.
    items = [
        _item("KT: 1500x500", price=1_504_058, row=10, path="ong gio ton"),
        _item("KT: 1500x500", price=13_333_616, row=20, path="ong gio mem"),
    ]
    assert find_price_inconsistencies(items) == []


def test_flags_and_warnings_reach_the_record(tmp_path: Path):
    path = tmp_path / "chao_gia.xlsx"
    wb = Workbook(); ws = wb.active; ws.title = "PCCC A1"
    ws.append(HEADERS)
    ws.append(["1", "DI-M9102", "Đầu báo khói địa chỉ", "cái", 71, 770_440, 71 * 770_440])
    ws.append(["2", "DI-M9102", "Đầu báo khói địa chỉ", "cái", 50, 850_000, 50 * 850_000])
    wb.save(path)

    data = load_workbook_items(path, DocumentRole.HSDT, bidder="NT")
    assert any("nhiều đơn giá khác nhau" in w for w in data.warnings)
    flagged = [i for i in data.items if any("nhiều đơn giá" in f for f in i.data_quality_flags)]
    assert len(flagged) == 2, "Cả hai dòng liên quan đều phải được gắn cờ"


def test_reference_appendix_is_not_checked(tmp_path: Path):
    # PL01 là bảng khối lượng mời thầu, không phải hồ sơ chào giá -> bỏ qua.
    path = tmp_path / "pl01.xlsx"
    wb = Workbook(); ws = wb.active; ws.title = "KLMT"
    ws.append(HEADERS)
    ws.append(["1", "DI-M9102", "Đầu báo khói địa chỉ", "cái", 71, 770_440, 0])
    ws.append(["2", "DI-M9102", "Đầu báo khói địa chỉ", "cái", 50, 850_000, 0])
    wb.save(path)

    data = load_workbook_items(path, DocumentRole.HSMT, bidder="")
    assert not any("nhiều đơn giá khác nhau" in w for w in data.warnings)


def test_description_lists_rows_for_manual_check():
    items = [
        _item("Đầu báo khói", code="DI-M9102", price=770_440, sheet="PCCC A1", row=11),
        _item("Đầu báo khói", code="DI-M9102", price=850_000, sheet="PCCC A1", row=25),
    ]
    text = find_price_inconsistencies(items)[0].describe()
    assert "PCCC A1!dòng 11" in text and "PCCC A1!dòng 25" in text
    assert "770,440" in text and "850,000" in text


def test_annotate_returns_empty_when_consistent():
    items = [_item("Đầu báo khói", code="DI-M9102", price=770_440, row=r) for r in (10, 20, 30)]
    assert annotate_price_inconsistencies(items) == []
