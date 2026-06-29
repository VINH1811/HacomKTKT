"""Khi người dùng đặt nhầm vị trí Phụ lục 01 và Phụ lục 02 (tải ngược ô),
hệ thống phải tự nhận diện theo cấu trúc cột và xử lý đúng, không cho kết quả sai.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from core.config import EnterpriseConfig
from core.models import DocumentRole
from core.tender_package import (
    compare_appendices_with_bidders,
    detect_appendix_role,
)


def _cfg() -> EnterpriseConfig:
    cfg = EnterpriseConfig()
    cfg.enable_semantic_matching = False
    cfg.enable_reranker = False
    return cfg


def _make_pl1(path: Path) -> None:
    """Phụ lục 01 - bảng khối lượng mời thầu (có khối lượng/đơn giá)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "KLMT"
    ws.append(["STT", "Mã hiệu", "Tên hạng mục", "ĐVT", "Khối lượng mời thầu", "Đơn giá tổng hợp", "Thành tiền"])
    ws.append(["1", "M-01", "Tủ điện tổng", "Cái", 2, 1_000_000, 2_000_000])
    ws.append(["2", "M-02", "Cáp đồng XLPE 4x10", "m", 100, 50_000, 5_000_000])
    wb.save(path)


def _make_pl2(path: Path) -> None:
    """Phụ lục 02 - yêu cầu vật tư (Vật tư + Thương hiệu - Xuất xứ)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "PL02"
    ws.append(["STT", "Tên vật tư thiết bị", "Thương hiệu - Xuất xứ", "Ghi chú"])
    ws.append(["1", "Tủ điện tổng", "Schneider / ABB - EU", ""])
    ws.append(["2", "Cáp đồng XLPE", "Cadivi / Trần Phú - Việt Nam", ""])
    wb.save(path)


def _make_bidder(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Chao gia"
    ws.append(["STT", "Mã hiệu", "Tên hạng mục", "ĐVT", "Khối lượng nhà thầu", "Đơn giá tổng hợp", "Thành tiền",
               "Thương hiệu", "Xuất xứ"])
    ws.append(["1", "M-01", "Tủ điện tổng", "Cái", 2, 1_050_000, 2_100_000, "Schneider", "EU"])
    ws.append(["2", "M-02", "Cáp đồng XLPE 4x10", "m", 100, 52_000, 5_200_000, "Cadivi", "Việt Nam"])
    wb.save(path)


def test_detect_role_classifies_each_file_correctly(tmp_path: Path):
    pl1 = tmp_path / "pl1.xlsx"
    pl2 = tmp_path / "pl2.xlsx"
    _make_pl1(pl1)
    _make_pl2(pl2)
    cfg = _cfg()

    assert detect_appendix_role(pl1, cfg) == "pl1"
    assert detect_appendix_role(pl2, cfg) == "pl2"


def _missing_count(result) -> int:
    return sum(row.match.kind.value == "missing" for row in result.rows)


def test_swapped_appendices_are_auto_corrected(tmp_path: Path):
    pl1 = tmp_path / "pl1.xlsx"
    pl2 = tmp_path / "pl2.xlsx"
    bidder = tmp_path / "bidder.xlsx"
    _make_pl1(pl1)
    _make_pl2(pl2)
    _make_bidder(bidder)
    cfg = _cfg()

    # Đúng vị trí.
    correct = compare_appendices_with_bidders(
        bidder_files=[("NT A", bidder)],
        output_dir=tmp_path / "out_correct",
        pl1_path=pl1, pl2_path=pl2, config=cfg,
    )
    # Đặt NGƯỢC: PL02 vào ô pl1, PL01 vào ô pl2.
    swapped = compare_appendices_with_bidders(
        bidder_files=[("NT A", bidder)],
        output_dir=tmp_path / "out_swapped",
        pl1_path=pl2, pl2_path=pl1, config=cfg,
    )

    # Kết quả phải tương đương: cùng số hạng mục chuẩn, cùng số dòng thiếu.
    assert swapped.result.summary.total_reference_items == correct.result.summary.total_reference_items
    assert _missing_count(swapped.result) == _missing_count(correct.result)
    # Và có cảnh báo rõ ràng: có thể đã tráo và hệ thống đã tự đổi lại.
    assert any("tráo" in w.lower() and "đổi lại" in w.lower() for w in swapped.result.warnings)


def test_single_pl2_uploaded_into_pl1_slot_is_rerouted(tmp_path: Path):
    pl2 = tmp_path / "pl2.xlsx"
    bidder = tmp_path / "bidder.xlsx"
    _make_pl2(pl2)
    _make_bidder(bidder)
    cfg = _cfg()

    # Người dùng chỉ tải một file PL02 nhưng bỏ vào ô PL01.
    out = compare_appendices_with_bidders(
        bidder_files=[("NT A", bidder)],
        output_dir=tmp_path / "out",
        pl1_path=pl2, pl2_path=None, config=cfg,
    )
    # Hệ thống phải xử lý như PL02 (chế độ PL02-only), không phải PL01.
    assert "PL02" in out.result.audit["mode"]
    assert out.result.audit.get("pl2_requirement_count", 0) >= 1
    assert any("phụ lục 02" in w.lower() for w in out.result.warnings)
