"""Dòng tổng cộng không phải hạng mục để đối chiếu.

Sheet tổng hợp của nhà thầu thường có dòng chỉ ghi "TỔNG :" kèm giá trị cả gói.
Trước đây tên đó không đủ chữ "tổng cộng" nên bị coi là hạng mục thật: tổng
thành tiền bị cộng thêm chính nó (gấp đôi), và báo cáo hiện một "hạng mục phát
sinh ngoài Phụ lục 01" trị giá bằng cả gói thầu.

Ngược lại, "Tổng thầu phụ" hay "Tổng đài điện thoại" là hạng mục thật, không
được loại nhầm.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from core.excel_reader import load_workbook_items
from core.models import DocumentRole, RowType

HEADER = ["STT", "Mã hiệu", "Tên hạng mục", "ĐVT", "KL nhà thầu chào",
          "Đơn giá tổng hợp", "Thành tiền"]
# Vài dòng hạng mục thật để bộ đọc nhận ra đây là bảng khối lượng.
_BASE = [
    ["1", "DI-M9102", "Đầu báo khói địa chỉ", "cái", 10, 500_000, 5_000_000],
    ["2", "VG-50", "Van góc chữa cháy DN50", "cái", 4, 900_000, 3_600_000],
]


def _book(tmp_path: Path, rows: list[list], name: str = "hs.xlsx") -> Path:
    # Không đặt tên sheet là "Tổng hợp": bộ đọc cố ý bỏ qua sheet tổng hợp.
    wb = Workbook(); ws = wb.active; ws.title = "Tong VK"
    ws.append(HEADER)
    for r in _BASE + rows:
        ws.append(r)
    path = tmp_path / name
    wb.save(path)
    return path


def _load(path: Path):
    return load_workbook_items(path, DocumentRole.HSDT, bidder="NT")


def test_bare_total_line_is_not_counted_as_an_item(tmp_path: Path):
    path = _book(tmp_path, [["", "", "TỔNG :", "", None, None, 8_600_000]])
    data = _load(path)
    totals = [i for i in data.items if "TỔNG" in (i.item_name or "")]
    assert totals and all(i.row_type is RowType.SUMMARY for i in totals)

    comparable = [i for i in data.items if i.is_comparable]
    assert sum(i.bid_amount or 0 for i in comparable) == 8_600_000, \
        "Tổng thành tiền không được cộng thêm chính dòng tổng"


def test_total_line_variants(tmp_path: Path):
    for index, label in enumerate(
        ("TỔNG", "TỔNG :", "Tổng cộng", "Tổng tiền", "TỔNG SỐ",
         "Tổng giá trị", "Tổng trước thuế", "THUẾ VAT (8%)")
    ):
        path = _book(tmp_path, [["", "", label, "", None, None, 8_600_000]],
                     name=f"tong{index}.xlsx")
        rows = [i for i in _load(path).items if (i.item_name or "").strip() == label]
        assert rows and rows[0].row_type is RowType.SUMMARY, f"{label!r} phải là dòng tổng"


def test_real_items_starting_with_tong_are_kept(tmp_path: Path):
    # "Tổng thầu phụ", "Tổng đài" là hạng mục thật — tên bắt đầu bằng "tổng"
    # không được làm chúng bị loại khỏi phần đối chiếu.
    for index, label in enumerate(("Tổng thầu phụ phần cơ điện", "Tổng đài điện thoại IP")):
        path = _book(tmp_path, [["3", "", label, "bộ", 2, 2_500_000, 5_000_000]],
                     name=f"that{index}.xlsx")
        rows = [i for i in _load(path).items if (i.item_name or "").strip() == label]
        assert rows and rows[0].row_type is not RowType.SUMMARY, \
            f"{label!r} là hạng mục thật, không phải dòng tổng"
        assert rows[0].is_comparable


def test_amount_only_row_is_a_subtotal_whatever_its_name(tmp_path: Path):
    # Chỉ có thành tiền, không đơn vị/khối lượng/đơn giá thì không đối chiếu
    # được gì — dù tên là hạng mục nghe rất thật. Xem thêm
    # tests/test_section_subtotal_rows.py.
    path = _book(tmp_path, [["", "", "Tổng thầu phụ phần cơ điện", "", None, None, 5_000_000]])
    rows = [i for i in _load(path).items if "Tổng thầu phụ" in (i.item_name or "")]
    assert rows and rows[0].row_type is RowType.SUMMARY


def test_total_line_without_amount_is_not_summary(tmp_path: Path):
    # Không có số tiền thì đó là tiêu đề, để các luật khác xử lý.
    path = _book(tmp_path, [["", "", "TỔNG :", "", None, None, None]])
    rows = [i for i in _load(path).items if "TỔNG" in (i.item_name or "")]
    assert not rows or rows[0].row_type is not RowType.SUMMARY
