from pathlib import Path

from openpyxl import Workbook

from core.excel_reader import _is_numbering_row, load_workbook_items
from core.models import DocumentRole, RowType


def test_column_number_legend_does_not_hide_normal_priced_row():
    assert _is_numbering_row([1, 2, 3, 4, 5, 6])
    assert not _is_numbering_row(["M-01", "Tủ điện", "Tủ", 1, 1000, 1000])


def test_component_without_price_is_not_false_quality_error(tmp_path: Path):
    path = tmp_path / "component.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Điện"
    ws.append(["STT", "Mã hiệu", "Tên hạng mục", "ĐVT", "Khối lượng", "Đơn giá", "Thành tiền"])
    ws.append(["1", "TU-01", "Tủ điện tổng", "Tủ", 1, 1_000_000, 1_000_000])
    ws.append([None, "ACB-01", "ACB 4P 3200A", "Cái", 2, None, None])
    wb.save(path)

    parsed = load_workbook_items(path, DocumentRole.HSDT, bidder="A")
    component = next(item for item in parsed.items if item.row_type is RowType.COMPONENT)
    assert "Thiếu đơn giá tổng hợp" not in component.data_quality_flags
    assert "Thiếu thành tiền" not in component.data_quality_flags
