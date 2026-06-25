from pathlib import Path

from openpyxl import Workbook, load_workbook

from core.annotator import annotate_bidder_workbook
from core.config import EnterpriseConfig
from core.excel_io import read_workbook_matrices, scan_xlsx_issues
from core.excel_reader import load_workbook_items
from core.models import DocumentRole
from core.parallel import WorkbookLoadSpec, load_workbooks_parallel


def _book(path: Path, price: float = 100.0, broken: bool = False) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Điện"
    ws.append(["Mã hiệu", "Tên hạng mục", "ĐVT", "Khối lượng nhà thầu", "Đơn giá tổng hợp", "Thành tiền nhà thầu"])
    ws.append(["M-01", "Cáp điện", "m", 10, price, 10 * price])
    if broken:
        ws["G2"] = "=#REF!"
    wb.save(path)


def _cfg() -> EnterpriseConfig:
    cfg = EnterpriseConfig()
    cfg.enable_semantic_matching = False
    cfg.enable_reranker = False
    cfg.excel_read_engine = "calamine"
    cfg.excel_read_workers = 4
    return cfg


def test_calamine_reader_preserves_rows(tmp_path: Path):
    path = tmp_path / "a.xlsx"
    _book(path)
    data = read_workbook_matrices(path, engine="calamine")
    assert data.engine == "calamine"
    assert data.sheets[0].rows[1][0] == "M-01"


def test_parallel_workbook_load(tmp_path: Path):
    paths = []
    for index in range(4):
        path = tmp_path / f"{index}.xlsx"
        _book(path, 100 + index)
        paths.append(path)
    specs = [WorkbookLoadSpec(str(i), path, DocumentRole.HSDT, f"B{i}") for i, path in enumerate(paths)]
    result = load_workbooks_parallel(specs, _cfg())
    assert set(result) == {"0", "1", "2", "3"}
    assert all(book.read_engine == "calamine" for book in result.values())


def test_ref_is_scanned_and_annotated(tmp_path: Path):
    source = tmp_path / "broken.xlsx"
    output = tmp_path / "marked.xlsx"
    _book(source, broken=True)
    scan = scan_xlsx_issues(source)
    assert any(issue.kind == "FORMULA_ERROR" and issue.cell == "G2" for issue in scan.issues)

    workbook = load_workbook_items(source, DocumentRole.HSDT, bidder="A", read_engine="calamine")
    assert any(issue["cell"] == "G2" for issue in workbook.formula_issues)
    annotate_bidder_workbook(source, output, workbook, [])

    marked = load_workbook(output, data_only=False)
    try:
        assert marked["Điện"]["G2"].fill.fill_type == "solid"
        assert marked["Điện"]["G2"].comment is not None
        values = [marked["AI_KIEM_TRA"].cell(row, 7).value for row in range(2, marked["AI_KIEM_TRA"].max_row + 1)]
        assert any("G2" in str(value) for value in values)
    finally:
        marked.close()
