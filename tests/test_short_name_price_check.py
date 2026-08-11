"""Hạng mục tên ngắn kèm thông số vẫn phải so được giá trong cùng hồ sơ.

"Ống PVC D20" dài 11 ký tự nên từng bị luật "tên quá ngắn = tiêu đề mục" loại
bỏ, không bao giờ so được giá — dù D20 đã đủ để nhận ra đó là món hàng cụ thể.

Nhưng nới ra không thôi thì sinh báo oan: với tên ngắn, thông tin phân biệt thật
sự nằm ở cột QUY CÁCH. Đo trên hồ sơ thật, "Ống D125 UPVC PN10" và "Ống D125
UPVC PN8" bị gộp làm một rồi báo lệch giá, trong khi khác cấp áp lực thì khác
giá là đúng. Vì vậy khóa gom nhóm theo tên phải kèm cả quy cách.
"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from core.excel_reader import load_workbook_items
from core.internal_consistency import _is_generic_name, find_price_inconsistencies
from core.models import DocumentRole

HDR = ["STT", "Mã hiệu", "Tên hạng mục", "ĐVT", "KL nhà thầu chào",
       "Đơn giá tổng hợp", "Thành tiền", "Mô tả/Quy cách"]
# Vài dòng nền để bộ đọc nhận ra đây là bảng khối lượng.
_FILL = [
    ["8", "KH-01", "Công tác vận chuyển vật tư lên tầng", "công", 50, 300_000, 15_000_000, ""],
    ["9", "KH-02", "Vệ sinh công nghiệp sau thi công", "m2", 800, 25_000, 20_000_000, ""],
]


def _book(tmp_path: Path, sheets: list[tuple[str, list[list]]], name: str = "hs.xlsx") -> Path:
    wb = Workbook(); first = True
    for title, rows in sheets:
        ws = wb.active if first else wb.create_sheet()
        ws.title = title; first = False
        ws.append(HDR)
        for row in rows + _FILL:
            ws.append(row)
    path = tmp_path / name
    wb.save(path)
    return path


def _issues(path: Path):
    data = load_workbook_items(path, DocumentRole.HSDT, bidder="NT")
    return find_price_inconsistencies(data.items)


@pytest.mark.parametrize("name", ["Ống PVC D20", "Ống D125", "KT 500x250", "Cáp CV 4x50"])
def test_short_names_with_specs_are_real_items(name: str):
    assert _is_generic_name(name) is False


@pytest.mark.parametrize("name", ["HỆ THỐNG ĐIỆN", "PHẦN NGẦM", "ĐIỆN NHẸ", "Vật tư"])
def test_headings_are_still_treated_as_generic(name: str):
    assert _is_generic_name(name) is True


def test_short_name_same_row_different_price_is_caught(tmp_path: Path):
    path = _book(tmp_path, [("Cấp nước", [
        ["1", "", "Ống PVC D20", "m", 500, 42_000, 21_000_000, ""],
        ["2", "", "Ống PVC D20", "m", 200, 48_000, 9_600_000, ""],
    ])])
    issues = _issues(path)
    assert len(issues) == 1
    assert issues[0].key_label == "Ống PVC D20"
    assert issues[0].same_sheet is True


def test_short_name_across_sheets_is_caught(tmp_path: Path):
    path = _book(tmp_path, [
        ("Cấp nước", [["1", "", "Ống PVC D20", "m", 500, 42_000, 21_000_000, ""]]),
        ("Thoát nước", [["1", "", "Ống PVC D20", "m", 200, 48_000, 9_600_000, ""]]),
    ])
    issues = _issues(path)
    assert len(issues) == 1
    assert issues[0].same_sheet is False, "Khác sheet phải được phân biệt để hạ mức cảnh báo"


def test_different_spec_is_not_flagged(tmp_path: Path):
    # Cùng tên ngắn nhưng khác cấp áp lực -> khác giá là đúng, không được báo.
    path = _book(tmp_path, [("Cấp nước", [
        ["1", "", "Ống D125", "m", 500, 320_800, 160_400_000, "UPVC PN10"],
        ["2", "", "Ống D125", "m", 200, 281_420, 56_284_000, "UPVC PN8"],
    ])])
    assert _issues(path) == []


def test_same_spec_different_price_is_flagged(tmp_path: Path):
    path = _book(tmp_path, [("Thông gió", [
        ["1", "", "KT 500x300", "Cái", 4, 413_500, 1_654_000, "Cửa nan Z kèm lưới chắn côn trùng"],
        ["2", "", "KT 500x300", "Cái", 6, 576_600, 3_459_600, "Cửa nan Z kèm lưới chắn côn trùng"],
    ])])
    issues = _issues(path)
    assert len(issues) == 1 and issues[0].spread_pct > 0.30


def test_code_still_wins_over_name(tmp_path: Path):
    # Có mã hiệu thì gom theo mã, quy cách khác nhau không làm mất cảnh báo.
    path = _book(tmp_path, [("Cấp nước", [
        ["1", "PVC-D20", "Ống nhựa PVC D20", "m", 500, 42_000, 21_000_000, "PN10"],
        ["2", "PVC-D20", "Ống nhựa PVC D20", "m", 200, 48_000, 9_600_000, "PN10"],
    ])])
    issues = _issues(path)
    assert len(issues) == 1 and issues[0].matched_by == "mã hiệu"


def test_rounding_difference_is_ignored(tmp_path: Path):
    path = _book(tmp_path, [("Cấp nước", [
        ["1", "", "Ống PVC D20", "m", 500, 42_000, 21_000_000, ""],
        ["2", "", "Ống PVC D20", "m", 200, 42_126, 8_425_200, ""],
    ])])
    assert _issues(path) == []


def test_different_unit_is_not_compared(tmp_path: Path):
    path = _book(tmp_path, [("Cấp nước", [
        ["1", "", "Ống PVC D20", "m", 500, 42_000, 21_000_000, ""],
        ["2", "", "Ống PVC D20", "cái", 200, 48_000, 9_600_000, ""],
    ])])
    assert _issues(path) == []
