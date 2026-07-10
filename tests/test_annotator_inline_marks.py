"""File đánh dấu: KHÔNG thêm cột AI; chỉ tô màu + chú thích (comment) lên ĐÚNG
những ô có sai lệch. Hạng mục khớp (kể cả khác sheet) không bị đánh dấu.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from core.config import EnterpriseConfig
from core.tender_package import compare_appendices_with_bidders

_AI_COLUMNS = {"AI MỨC ĐỘ", "AI LÝ DO", "AI GHI CHÚ"}
SHEET = "1. HT điện"
NAME_COL, QTY_COL = 3, 6  # Tên hạng mục, Khối lượng nhà thầu chào


def _cfg(**kw) -> EnterpriseConfig:
    cfg = EnterpriseConfig()
    cfg.enable_semantic_matching = False
    cfg.enable_reranker = False
    for k, v in kw.items():
        setattr(cfg, k, v)
    return cfg


def _has_fill(cell) -> bool:
    return bool(cell.fill and cell.fill.patternType and cell.fill.fgColor.rgb not in (None, "00000000"))


def _bidder_header(ws) -> None:
    ws.append(["STT", "Mã hiệu", "Tên hạng mục", "ĐVT", "Khối lượng mời thầu",
               "Khối lượng nhà thầu chào", "Đơn giá tổng hợp", "Thành tiền"])


def _run(tmp_path: Path, pl1_sheet: str, bidder_rows: list[list]) -> Path:
    pl1 = tmp_path / "pl1.xlsx"
    wb = Workbook(); ws = wb.active; ws.title = pl1_sheet
    ws.append(["STT", "Mã hiệu", "Tên hạng mục", "ĐVT", "Khối lượng"])
    ws.append(["1", "M-01", "Tủ điện tổng", "Cái", 2])
    ws.append(["2", "M-02", "Cáp đồng XLPE 4x10", "m", 100])
    ws.append(["3", "M-03", "Máy bơm nước", "Bộ", 1])
    wb.save(pl1)

    bidder = tmp_path / "b.xlsx"
    wb = Workbook(); ws = wb.active; ws.title = SHEET
    _bidder_header(ws)
    for r in bidder_rows:
        ws.append(r)
    wb.save(bidder)

    out = compare_appendices_with_bidders(
        [("NT A", bidder)], tmp_path / "out", pl1_path=pl1,
        config=_cfg(annotate_workers=1, parse_cache_size=0, match_cache_size=0),
    )
    return Path(out.annotated_files["NT A"])


# Cùng tên sheet với PL01 -> ghép chắc chắn theo mã/cấu trúc.
def _annotate_same_sheet(tmp_path: Path) -> Path:
    return _run(tmp_path, SHEET, [
        # Dòng 2: TÊN khác (mã/STT/khối lượng giống) -> chỉ lệch tên.
        ["1", "M-01", "Máy phát điện dự phòng 500kVA", "Cái", 2, 2, 1_000_000, 2_000_000],
        # Dòng 3: KHỐI LƯỢNG chào khác (tên giống) -> chỉ lệch khối lượng.
        ["2", "M-02", "Cáp đồng XLPE 4x10", "m", 100, 130, 50_000, 6_500_000],
        # Dòng 4: khớp hoàn toàn -> KHÔNG đánh dấu gì.
        ["3", "M-03", "Máy bơm nước", "Bộ", 1, 1, 3_000_000, 3_000_000],
    ])


def test_no_ai_columns_added_beside_data(tmp_path: Path):
    ws = load_workbook(_annotate_same_sheet(tmp_path))[SHEET]
    headers = {str(c.value) for c in ws[1] if c.value is not None}
    assert not (headers & _AI_COLUMNS), headers
    assert not any(str(h).startswith("AI ") for h in headers)
    assert ws.max_column == 8  # đúng số cột file gốc, không phình


def test_front_sheets_kept(tmp_path: Path):
    wb = load_workbook(_annotate_same_sheet(tmp_path))
    assert "AI_TONG_QUAN" in wb.sheetnames
    assert "AI_KIEM_TRA" in wb.sheetnames


def test_only_the_problem_cell_is_marked(tmp_path: Path):
    ws = load_workbook(_annotate_same_sheet(tmp_path))[SHEET]

    # Dòng lệch TÊN: chỉ ô tên hạng mục bị bôi + chú thích, ô khối lượng thì không.
    assert ws.cell(2, NAME_COL).comment is not None and _has_fill(ws.cell(2, NAME_COL))
    assert ws.cell(2, QTY_COL).comment is None and not _has_fill(ws.cell(2, QTY_COL))

    # Dòng lệch KHỐI LƯỢNG: chỉ ô khối lượng bị bôi + chú thích, ô tên thì không.
    assert ws.cell(3, QTY_COL).comment is not None and _has_fill(ws.cell(3, QTY_COL))
    assert ws.cell(3, NAME_COL).comment is None and not _has_fill(ws.cell(3, NAME_COL))

    # Dòng khớp hoàn toàn: KHÔNG có ô nào bị bôi hay chú thích.
    assert all(ws.cell(4, c).comment is None and not _has_fill(ws.cell(4, c)) for c in range(1, 9))


def test_matched_but_different_sheet_is_not_marked(tmp_path: Path):
    # PL01 tên sheet KHÁC nhà thầu, hạng mục khớp hoàn toàn -> chỉ là "khác sheet",
    # KHÔNG được tô màu hay chú thích.
    ann = _run(tmp_path, "KLMT", [
        ["1", "M-01", "Tủ điện tổng", "Cái", 2, 2, 1_000_000, 2_000_000],
        ["2", "M-02", "Cáp đồng XLPE 4x10", "m", 100, 100, 50_000, 5_000_000],
    ])
    ws = load_workbook(ann)[SHEET]
    for row in (2, 3):
        assert all(ws.cell(row, c).comment is None and not _has_fill(ws.cell(row, c)) for c in range(1, 9))
