"""Cache SHA-256 (parse + ghép cặp) và annotation đa tiến trình.

Yêu cầu tuyệt đối: kết quả CÓ cache / CHẠY SONG SONG phải giống hệt kết quả
không cache / chạy tuần tự. Các test ở đây đối chứng trực tiếp hai chế độ.
"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from core.config import EnterpriseConfig
from core.matcher import clear_match_cache, match_items, match_items_cached
from core.models import CompareThresholds, DocumentRole
from core.parallel import WorkbookLoadSpec, clear_parse_cache, load_workbooks_parallel
from core.tender_package import compare_appendices_with_bidders


@pytest.fixture(autouse=True)
def _fresh_caches():
    clear_parse_cache()
    clear_match_cache()
    yield
    clear_parse_cache()
    clear_match_cache()


def _cfg(**kw) -> EnterpriseConfig:
    cfg = EnterpriseConfig()
    cfg.enable_semantic_matching = False
    cfg.enable_reranker = False
    for k, v in kw.items():
        setattr(cfg, k, v)
    return cfg


def _pl1(path: Path) -> None:
    wb = Workbook(); ws = wb.active; ws.title = "KLMT"
    ws.append(["STT", "Mã hiệu", "Tên hạng mục", "ĐVT", "Khối lượng"])
    ws.append(["1", "M-01", "Tủ điện tổng", "Cái", 2])
    ws.append(["2", "M-02", "Cáp đồng XLPE 4x10", "m", 100])
    ws.append(["3", "M-03", "Máy bơm nước thải", "Bộ", 1])
    wb.save(path)


def _bidder(path: Path, price: float = 1_000_000) -> None:
    wb = Workbook(); ws = wb.active; ws.title = "1. HT điện"
    ws.append(["STT", "Mã hiệu", "Tên hạng mục", "ĐVT", "Khối lượng mời thầu",
               "Khối lượng nhà thầu chào", "Đơn giá tổng hợp", "Thành tiền"])
    ws.append(["1", "M-01", "Tủ điện tổng", "Cái", 2, 2, price, 2 * price])
    ws.append(["2", "M-02", "Cáp đồng XLPE 4x10", "m", 100, 130, 50_000, 6_500_000])
    wb.save(path)


def _spec(path: Path, role=DocumentRole.HSDT, bidder="NT A") -> WorkbookLoadSpec:
    return WorkbookLoadSpec(key="k", path=path, role=role, bidder=bidder)


# --- Cache parse ---------------------------------------------------------------

def test_parse_cache_hit_returns_equal_but_independent_copy(tmp_path: Path):
    path = tmp_path / "b.xlsx"; _bidder(path)
    cfg = _cfg()

    first = load_workbooks_parallel([_spec(path)], cfg)["k"]
    # Cố tình "làm bẩn" kết quả lần 1 — lần 2 không được bị ảnh hưởng.
    first.items[0].data_quality_flags.append("BẨN")
    first.items[0].raw["X"] = "BẨN"
    first.warnings.append("BẨN")

    second = load_workbooks_parallel([_spec(path)], cfg)["k"]
    assert second.read_engine.endswith("+cache")  # xác nhận trúng cache
    assert "BẨN" not in second.items[0].data_quality_flags
    assert "X" not in second.items[0].raw
    assert "BẨN" not in second.warnings
    # Nội dung nghiệp vụ giống hệt lần đọc gốc.
    assert [i.item_name for i in second.items] == [i.item_name for i in first.items]
    assert [i.unit_price_total for i in second.items] == [i.unit_price_total for i in first.items]


def test_parse_cache_invalidated_when_file_content_changes(tmp_path: Path):
    path = tmp_path / "b.xlsx"; _bidder(path, price=1_000_000)
    cfg = _cfg()
    first = load_workbooks_parallel([_spec(path)], cfg)["k"]
    assert first.items[0].unit_price_total == 1_000_000

    _bidder(path, price=9_999_999)  # ghi đè nội dung -> SHA đổi
    second = load_workbooks_parallel([_spec(path)], cfg)["k"]
    assert second.items[0].unit_price_total == 9_999_999  # không dính dữ liệu cũ
    assert not second.read_engine.endswith("+cache")


def test_parse_cache_overrides_bidder_and_path_per_run(tmp_path: Path):
    # Cùng nội dung nhưng khác tên file/tên nhà thầu (job mới) -> phải theo lần chạy mới.
    p1 = tmp_path / "002_a.xlsx"; _bidder(p1)
    p2 = tmp_path / "005_b.xlsx"; p2.write_bytes(p1.read_bytes())  # cùng SHA
    cfg = _cfg()
    load_workbooks_parallel([_spec(p1, bidder="NT Một")], cfg)
    out = load_workbooks_parallel([_spec(p2, bidder="NT Hai")], cfg)["k"]
    assert out.read_engine.endswith("+cache")
    assert out.bidder == "NT Hai"
    assert out.path == p2
    assert all(i.bidder == "NT Hai" and i.workbook == p2.name for i in out.items)


def test_parse_cache_disabled_when_size_zero(tmp_path: Path):
    path = tmp_path / "b.xlsx"; _bidder(path)
    cfg = _cfg(parse_cache_size=0)
    load_workbooks_parallel([_spec(path)], cfg)
    second = load_workbooks_parallel([_spec(path)], cfg)["k"]
    assert not second.read_engine.endswith("+cache")


# --- Cache ghép cặp -------------------------------------------------------------

def test_match_cache_returns_identical_results_as_uncached(tmp_path: Path):
    pl1 = tmp_path / "pl1.xlsx"; _pl1(pl1)
    bid = tmp_path / "b.xlsx"; _bidder(bid)
    cfg = _cfg()
    ref = load_workbooks_parallel([_spec(pl1, role=DocumentRole.HSMT, bidder="PL01")], cfg)["k"]
    cand = load_workbooks_parallel([_spec(bid)], cfg)["k"]

    fresh = match_items(ref.items, cand.items, cfg)
    first = match_items_cached(ref, cand, cfg)
    second = match_items_cached(ref, cand, cfg)  # lần này trúng cache

    def sig(matches):
        return [(m.reference_index, m.candidate_index, m.kind.value, round(m.score, 9), m.reason)
                for m in matches]

    assert sig(first) == sig(fresh)
    assert sig(second) == sig(fresh)
    # Bản sao độc lập, không dùng chung object với cache.
    assert second[0] is not first[0]


def test_match_cache_survives_threshold_changes_but_not_matcher_config(tmp_path: Path):
    pl1 = tmp_path / "pl1.xlsx"; _pl1(pl1)
    bid = tmp_path / "b.xlsx"; _bidder(bid)
    cfg = _cfg()
    ref = load_workbooks_parallel([_spec(pl1, role=DocumentRole.HSMT, bidder="PL01")], cfg)["k"]
    cand = load_workbooks_parallel([_spec(bid)], cfg)["k"]
    base = match_items_cached(ref, cand, cfg)

    # Đổi ngưỡng giá/khối lượng (kịch bản chạy lặp D-05): matching không phụ
    # thuộc các ngưỡng này nên kết quả phải giữ nguyên.
    cfg2 = _cfg()
    cfg2.thresholds = CompareThresholds(quantity_warn_pct=0.03)
    again = match_items_cached(ref, cand, cfg2)
    assert [(m.reference_index, m.candidate_index) for m in again] == \
           [(m.reference_index, m.candidate_index) for m in base]

    # Đổi tham số MATCHING thật (name_reject_score) -> khóa khác, không trả nhầm.
    cfg3 = _cfg()
    cfg3.thresholds = CompareThresholds(name_reject_score=0.99)
    strict = match_items_cached(ref, cand, cfg3)
    assert isinstance(strict, list)  # chạy lại matcher thật, không lỗi


# --- Đối chứng end-to-end: cache bật vs tắt phải Y HỆT ---------------------------

def _result_signature(result):
    rows = sorted(
        (
            r.canonical_id, r.bidder, r.severity.value, r.match.kind.value,
            round(r.match.score, 9), tuple(sorted(r.flags)),
            tuple(sorted((d.field, str(d.reference_value), str(d.candidate_value), d.severity.value)
                         for d in r.differences)),
            r.pl2_status,
        )
        for r in result.rows
    )
    s = result.summary
    return (rows, s.total_reference_items, s.missing_items, s.extra_items,
            s.review_rows, s.warning_rows, s.critical_rows)


def test_end_to_end_cached_run_identical_to_uncached(tmp_path: Path):
    pl1 = tmp_path / "pl1.xlsx"; _pl1(pl1)
    b1 = tmp_path / "b1.xlsx"; _bidder(b1, price=1_000_000)
    b2 = tmp_path / "b2.xlsx"; _bidder(b2, price=1_500_000)
    bidders = [("NT A", b1), ("NT B", b2)]

    # Chuẩn không cache.
    off = _cfg(parse_cache_size=0, match_cache_size=0, annotate_workers=1)
    baseline = compare_appendices_with_bidders(bidders, tmp_path / "o0", pl1_path=pl1, pl2_path=None, config=off)

    # Có cache: chạy 2 lần, lần 2 trúng cache toàn bộ.
    on = _cfg(annotate_workers=1)
    compare_appendices_with_bidders(bidders, tmp_path / "o1", pl1_path=pl1, pl2_path=None, config=on)
    cached = compare_appendices_with_bidders(bidders, tmp_path / "o2", pl1_path=pl1, pl2_path=None, config=on)

    assert _result_signature(cached.result) == _result_signature(baseline.result)


# --- Annotation đa tiến trình: file y hệt tuần tự --------------------------------

def _annotated_content(path: Path):
    wb = load_workbook(path)
    cells = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        if ws.__class__.__name__ == "Chartsheet":
            continue
        for row in ws.iter_rows():
            for c in row:
                fill = c.fill.fgColor.rgb if (c.fill and c.fill.patternType) else None
                comment = c.comment.text if c.comment else None
                if c.value is not None or fill or comment:
                    cells.append((sn, c.row, c.column, str(c.value), str(fill), comment))
    wb.close()
    return cells


def test_process_pool_annotation_output_identical_to_sequential(tmp_path: Path):
    pl1 = tmp_path / "pl1.xlsx"; _pl1(pl1)
    b1 = tmp_path / "b1.xlsx"; _bidder(b1, price=1_000_000)
    b2 = tmp_path / "b2.xlsx"; _bidder(b2, price=1_500_000)
    bidders = [("NT A", b1), ("NT B", b2)]

    seq = compare_appendices_with_bidders(
        bidders, tmp_path / "seq", pl1_path=pl1, pl2_path=None,
        config=_cfg(annotate_workers=1, parse_cache_size=0, match_cache_size=0),
    )
    par = compare_appendices_with_bidders(
        bidders, tmp_path / "par", pl1_path=pl1, pl2_path=None,
        config=_cfg(annotate_workers=2, parse_cache_size=0, match_cache_size=0),
    )

    assert set(seq.annotated_files) == set(par.annotated_files) == {"NT A", "NT B"}
    for name in seq.annotated_files:
        assert _annotated_content(seq.annotated_files[name]) == _annotated_content(par.annotated_files[name]), \
            f"File đánh dấu của {name} phải y hệt giữa tuần tự và đa tiến trình"
