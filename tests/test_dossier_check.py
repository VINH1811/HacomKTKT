"""Đánh giá tính đầy đủ hồ sơ: nhận diện đầu mục qua tên file/thư mục (không
phân biệt hoa thường và dấu tiếng Việt), ngưỡng số lượng (BCTC 3 năm), phân biệt
bắt buộc/không bắt buộc, và xuất báo cáo ma trận."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from core.dossier_check import (
    STATUS_MISSING,
    STATUS_OK,
    STATUS_OPTIONAL_MISSING,
    STATUS_PARTIAL,
    evaluate_dossier,
    export_dossier_report,
)


def _touch(root: Path, rel: str) -> None:
    p = root / rel
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_bytes(b"x")


def _full_dossier(root: Path) -> None:
    for rel in [
        "1. Đơn chào giá/Đơn chào giá.pdf",
        "2. Bảng chào giá/BOQ_HACOM_MALL.xlsx",
        "3. Vật tư/Danh mục vật tư thiết bị.pdf",
        "4. Tài chính/Báo cáo tài chính năm 2022.pdf",
        "4. Tài chính/Báo cáo tài chính năm 2023.pdf",
        "4. Tài chính/Báo cáo tài chính năm 2024.pdf",
        "5. Năng lực/Giấy chứng nhận đăng ký doanh nghiệp.pdf",
        "6. HĐ/Bảng kê khai hợp đồng tương tự.pdf",
        "7. Nhân sự/Sơ đồ tổ chức.pdf",
        "8. Tiến độ/Bảng tiến độ thi công.pdf",
        "9. BPTC/Biện pháp thi công tổng thể.pdf",
        "10. ATLD/Biện pháp ATLD tại công trình.pdf",
    ]:
        _touch(root, rel)


def _get(result, key):
    return next(c for c in result.categories if c.item.key == key)


def test_full_dossier_all_required_ok(tmp_path: Path):
    _full_dossier(tmp_path)
    res = evaluate_dossier("NT A", tmp_path)
    assert res.missing_required == []
    assert _get(res, "bao_cao_tai_chinh").status == STATUS_OK


def test_missing_and_partial_detected(tmp_path: Path):
    _full_dossier(tmp_path)
    # Xoá đơn chào giá và 2/3 BCTC
    (tmp_path / "1. Đơn chào giá/Đơn chào giá.pdf").unlink()
    (tmp_path / "4. Tài chính/Báo cáo tài chính năm 2022.pdf").unlink()
    (tmp_path / "4. Tài chính/Báo cáo tài chính năm 2023.pdf").unlink()
    res = evaluate_dossier("NT B", tmp_path)
    assert _get(res, "don_chao_gia").status == STATUS_MISSING
    assert _get(res, "bao_cao_tai_chinh").status == STATUS_PARTIAL
    keys = {c.item.key for c in res.missing_required}
    assert {"don_chao_gia", "bao_cao_tai_chinh"} <= keys


def test_diacritics_insensitive_matching(tmp_path: Path):
    # Tên không dấu (kiểu Searefico/VanKhanh) vẫn phải nhận diện được.
    _touch(tmp_path, "1/Thu chao gia.pdf")
    _touch(tmp_path, "2/SRF-Bang chao gia chi tiet.pdf")
    _touch(tmp_path, "3/Bien phap thi cong.pdf")
    res = evaluate_dossier("NT C", tmp_path)
    assert _get(res, "don_chao_gia").status == STATUS_OK
    assert _get(res, "bang_chao_gia").status == STATUS_OK
    assert _get(res, "bien_phap_thi_cong").status == STATUS_OK


def test_optional_category_not_required(tmp_path: Path):
    _full_dossier(tmp_path)  # không có catalogue/ủy quyền
    res = evaluate_dossier("NT D", tmp_path)
    assert _get(res, "catalogue").status == STATUS_OPTIONAL_MISSING
    assert all(c.item.key not in {"catalogue", "uy_quyen"} for c in res.missing_required)


def test_export_matrix_report(tmp_path: Path):
    root_a = tmp_path / "a"; _full_dossier(root_a)
    root_b = tmp_path / "b"; _touch(root_b, "chi co mot file.pdf")
    results = [evaluate_dossier("NT A", root_a), evaluate_dossier("NT B", root_b)]
    out = tmp_path / "checklist.xlsx"
    export_dossier_report(results, out)
    wb = load_workbook(out)
    assert "Checklist hồ sơ" in wb.sheetnames
    assert "BC NT A" in wb.sheetnames and "BC NT B" in wb.sheetnames
    ws = wb["Checklist hồ sơ"]
    values = [ws.cell(r, 3).value for r in range(4, ws.max_row)]
    assert any(v and STATUS_MISSING in str(v) for v in values), "NT B thiếu gần hết phải có ô THIẾU"
