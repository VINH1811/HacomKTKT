"""Vá "vùng xám" format lạ:
1. Mở rộng từ điển từ khóa cột (TT, Đv tính, Số lượng, Hạng mục, Giá...) để các
   file BOQ tiếng Việt viết khác thuật ngữ quen vẫn được đọc và ghép đúng.
2. Cảnh báo khi tỷ lệ ghép cặp thấp bất thường (< 20%) — dấu hiệu file có định
   dạng chưa được nhận diện đầy đủ, tránh người dùng tin nhầm vào các dòng
   'thiếu/phát sinh' giả.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from core.config import EnterpriseConfig
from core.excel_reader import load_workbook_items, map_columns
from core.models import DocumentRole, RowType
from core.tender_package import compare_appendices_with_bidders


def _cfg() -> EnterpriseConfig:
    cfg = EnterpriseConfig()
    cfg.enable_semantic_matching = False
    cfg.enable_reranker = False
    cfg.parse_cache_size = 0
    cfg.match_cache_size = 0
    return cfg


def _pl1(path: Path, n_items: int = 2) -> None:
    wb = Workbook(); ws = wb.active; ws.title = "KLMT"
    ws.append(["STT", "Mã hiệu", "Tên hạng mục", "ĐVT", "Khối lượng"])
    names = ["Tủ điện tổng", "Cáp đồng XLPE 4x10", "Máy bơm nước", "Đèn LED âm trần",
             "Ống thép D100", "Van một chiều D50", "Quạt hướng trục", "Cảm biến khói",
             "Tủ trung thế RMU", "Thang cáp 200x100", "Dây tiếp địa M10", "Aptomat 3P 63A"]
    for i in range(n_items):
        ws.append([str(i + 1), f"M-{i+1:02d}", names[i % len(names)], "Cái", 10])
    wb.save(path)


# --- Phần 1: từ điển từ khóa mở rộng -------------------------------------------

def test_variant_headers_tt_dvtinh_kl_are_mapped():
    headers = ["Đơn giá", "Ghi chú nội bộ", "Nội dung công việc", "TT", "KL", "Đv tính", "Thành tiền"]
    fixed, _ = map_columns(headers, DocumentRole.HSDT)
    fields = set(fixed.values())
    assert "stt" in fields            # "TT"
    assert "unit" in fields           # "Đv tính"
    assert "item_name" in fields      # "Nội dung công việc"
    assert "bid_quantity" in fields   # "KL"
    assert "unit_price_total" in fields


def test_variant_headers_hangmuc_soluong_gia_are_mapped():
    headers = ["TT", "Hạng mục", "Đơn vị", "Số lượng", "Giá"]
    fixed, _ = map_columns(headers, DocumentRole.HSDT)
    fields = set(fixed.values())
    assert "stt" in fields
    assert "item_name" in fields      # "Hạng mục" đứng một mình
    assert "unit" in fields
    assert "bid_quantity" in fields   # "Số lượng"
    assert "unit_price_total" in fields  # "Giá"


def test_variant_terms_do_not_steal_amount_or_material_columns():
    # "Thành tiền hạng mục" phải vẫn là thành tiền, không bị nhận nhầm tên hạng mục;
    # "Mô tả/Quy cách" (kèm quy cách) vẫn là vật tư, không phải tên hạng mục.
    headers = ["STT", "Tên hạng mục", "ĐVT", "Khối lượng", "Đơn giá", "Thành tiền hạng mục", "Mô tả quy cách"]
    fixed, _ = map_columns(headers, DocumentRole.HSDT)
    by_field = {field: col for col, field in fixed.items()}
    assert by_field["item_name"] == 1          # cột "Tên hạng mục" chuẩn thắng
    assert by_field.get("bid_amount") == 5     # "Thành tiền hạng mục"
    assert by_field.get("material") == 6       # "Mô tả quy cách"


def test_grey_zone_file_now_parses_as_detail_and_matches(tmp_path: Path):
    # File kiểu F1 (TT/KL/Đv tính, header ở dòng 4, cột đảo lộn): trước đây dòng bị
    # phân loại COMPONENT và ghép 0; giờ phải là DETAIL và khớp đủ với PL01.
    pl1 = tmp_path / "pl1.xlsx"; _pl1(pl1, n_items=2)
    wb = Workbook(); ws = wb.active; ws.title = "BANG GIA"
    ws.append(["CÔNG TY XYZ"]); ws.append(["Bảng chào giá vật tư"]); ws.append([])
    ws.append(["Đơn giá", "Ghi chú nội bộ", "Nội dung công việc", "TT", "KL", "Đv tính", "Thành tiền"])
    ws.append([1_050_000, "a", "Tủ điện tổng", "1", 10, "Cái", 10_500_000])
    ws.append([52_000, "b", "Cáp đồng XLPE 4x10", "2", 10, "Cái", 520_000])
    nt = tmp_path / "nt.xlsx"; wb.save(nt)

    parsed = load_workbook_items(nt, DocumentRole.HSDT, bidder="NT")
    details = [it for it in parsed.items if it.row_type is RowType.DETAIL]
    assert len(details) == 2 and all(it.stt for it in details)

    out = compare_appendices_with_bidders([("NT", nt)], tmp_path / "o",
                                          pl1_path=pl1, pl2_path=None, config=_cfg())
    s = out.result.summary
    assert s.missing_items == 0 and s.extra_items == 0
    assert s.exact_matches + s.fuzzy_matches == 2


# --- Phần 2: cảnh báo tỷ lệ ghép thấp -------------------------------------------

def _weird_bidder(path: Path) -> None:
    """File dùng thuật ngữ hoàn toàn không nhận diện được cột tên -> chỉ đọc được
    dữ liệu rời rạc, ghép cặp sẽ thất bại gần hết."""
    wb = Workbook(); ws = wb.active; ws.title = "GIA"
    ws.append(["STT", "Tên hạng mục", "ĐVT", "Khối lượng", "Đơn giá"])
    # 12 hạng mục tên hoàn toàn khác PL01 -> ghép 0.
    for i in range(12):
        ws.append([str(i + 1), f"Vật tư hoàn toàn khác biệt số {i}", "Bộ", 1, 999])
    wb.save(path)


def test_low_match_ratio_triggers_warning(tmp_path: Path):
    pl1 = tmp_path / "pl1.xlsx"; _pl1(pl1, n_items=12)
    nt = tmp_path / "nt.xlsx"; _weird_bidder(nt)

    out = compare_appendices_with_bidders([("NT Lạ", nt)], tmp_path / "o",
                                          pl1_path=pl1, pl2_path=None, config=_cfg())
    warns = [w for w in out.result.warnings if "ghép được" in w and "GIẢ" in w]
    assert warns, "Phải cảnh báo tỷ lệ ghép thấp bất thường"
    assert "NT Lạ" in warns[0]


def test_normal_match_ratio_has_no_low_match_warning(tmp_path: Path):
    pl1 = tmp_path / "pl1.xlsx"; _pl1(pl1, n_items=12)
    # Nhà thầu chào đúng 12 hạng mục giống PL01.
    wb = Workbook(); ws = wb.active; ws.title = "BOQ"
    ws.append(["STT", "Mã hiệu", "Tên hạng mục", "ĐVT", "Khối lượng nhà thầu chào", "Đơn giá tổng hợp", "Thành tiền"])
    names = ["Tủ điện tổng", "Cáp đồng XLPE 4x10", "Máy bơm nước", "Đèn LED âm trần",
             "Ống thép D100", "Van một chiều D50", "Quạt hướng trục", "Cảm biến khói",
             "Tủ trung thế RMU", "Thang cáp 200x100", "Dây tiếp địa M10", "Aptomat 3P 63A"]
    for i in range(12):
        ws.append([str(i + 1), f"M-{i+1:02d}", names[i], "Cái", 10, 1000, 10_000])
    nt = tmp_path / "nt.xlsx"; wb.save(nt)

    out = compare_appendices_with_bidders([("NT Chuẩn", nt)], tmp_path / "o",
                                          pl1_path=pl1, pl2_path=None, config=_cfg())
    assert not any("ghép được" in w and "GIẢ" in w for w in out.result.warnings)


def test_small_files_never_trigger_low_match_warning(tmp_path: Path):
    # Bảng chuẩn < 10 hạng mục: không đủ dữ liệu thống kê, không cảnh báo dù ghép 0.
    pl1 = tmp_path / "pl1.xlsx"; _pl1(pl1, n_items=2)
    wb = Workbook(); ws = wb.active; ws.title = "BOQ"
    ws.append(["STT", "Tên hạng mục", "ĐVT", "Khối lượng", "Đơn giá"])
    ws.append(["1", "Thứ gì đó hoàn toàn khác", "Bộ", 1, 999])
    nt = tmp_path / "nt.xlsx"; wb.save(nt)

    out = compare_appendices_with_bidders([("NT", nt)], tmp_path / "o",
                                          pl1_path=pl1, pl2_path=None, config=_cfg())
    assert not any("ghép được" in w and "GIẢ" in w for w in out.result.warnings)
