"""Suy ra vai trò cột từ dữ liệu khi tiêu đề viết theo cách chưa từng gặp.

Đọc cột theo từ khóa tiêu đề luôn có giới hạn: file Excel không có chuẩn nào,
mỗi đơn vị đặt tên một kiểu ("ĐGTH", "Tên vật tư", "Giá trị"...). Khi tiêu đề
trượt thì cả sheet bị đọc thiếu mà không có cách nào cứu.

Bảng khối lượng lại luôn thỏa: khối lượng × đơn giá = thành tiền. Dùng quan hệ
đó để điền vào chỗ tiêu đề bỏ trống — nhưng KHÔNG được đụng vào file mà tiêu đề
đã đọc đúng.
"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from core.column_inference import (
    find_quantity_price_amount,
    find_unit_column,
    looks_like_unit_column,
)
from core.excel_reader import load_workbook_items
from core.models import DocumentRole

BASE = ["STT", "Mã hiệu", "Tên hạng mục", "ĐVT",
        "KL nhà thầu chào", "Đơn giá tổng hợp", "Thành tiền"]


def _rows(count: int = 30) -> list[list]:
    out = []
    for i in range(1, count + 1):
        qty, price = i * 3, 100_000 + i * 7_531
        out.append([str(i), f"HM-{i:03d}", f"Hạng mục thử nghiệm số {i}", "cái",
                    qty, price, qty * price])
    return out


def _book(tmp_path: Path, header: list[str], name: str = "hs.xlsx") -> Path:
    wb = Workbook(); ws = wb.active; ws.title = "PCCC"
    ws.append(header)
    for row in _rows():
        ws.append(row)
    path = tmp_path / name
    wb.save(path)
    return path


def _read(path: Path):
    data = load_workbook_items(path, DocumentRole.HSDT, bidder="NT")
    items = [i for i in data.items if i.is_comparable]
    return items, data.warnings


# ------------------------------------------------------- quan hệ toán học

def test_relation_is_found_without_any_header():
    found = find_quantity_price_amount(_rows(), 7)
    assert found is not None
    (quantity, price, amount), hits, total = found
    assert (quantity, price, amount) == (4, 5, 6)
    assert hits == total == 30


def test_known_columns_anchor_the_search():
    # Biết trước cột đơn giá thì chỉ còn phải tìm hai cột kia.
    found = find_quantity_price_amount(_rows(), 7, known_price=5)
    assert found and found[0] == (4, 5, 6)


def test_random_numbers_do_not_produce_a_false_relation():
    rows = [[str(i), "x", "y", "cái", i, i * 2, i * 5] for i in range(1, 40)]
    # 5 != 1*2 nên bộ (4,5,6) không thỏa; nhưng (0,?) có thể thỏa ngẫu nhiên.
    found = find_quantity_price_amount(rows, 7)
    assert found is None or found[0] != (4, 5, 6)


def test_too_few_rows_is_not_trusted():
    assert find_quantity_price_amount(_rows(4), 7) is None


# --------------------------------------------------------- đơn vị tính

@pytest.mark.parametrize("values, expected", [
    (["cái"] * 10, True),
    (["m", "m2", "m3", "kg", "bộ", "cái", "md", "tấn"], True),
    (["Đầu báo khói địa chỉ"] * 10, False),
    (["1", "2", "3", "4", "5", "6"], False),
])
def test_unit_column_detection(values: list[str], expected: bool):
    assert looks_like_unit_column(values) is expected


def test_unit_column_found_by_value():
    assert find_unit_column(_rows(), 7) == 3


# ------------------------------------------------- tích hợp vào bộ đọc

def test_unreadable_headers_are_rescued(tmp_path: Path):
    path = _book(tmp_path, ["STT", "Mã hiệu", "Tên hạng mục", "ĐVT",
                            "Cột 5", "Cột 6", "Cột 7"])
    items, warnings = _read(path)
    assert len(items) == 30
    assert all(i.bid_quantity is not None for i in items)
    assert all(i.unit_price_total is not None for i in items)
    assert all(i.bid_amount is not None for i in items)
    assert [w for w in warnings if "SUY RA" in w], "Phải nói rõ là đã suy ra, không im lặng"


def test_unit_column_rescued_when_header_is_meaningless(tmp_path: Path):
    path = _book(tmp_path, ["STT", "Mã hiệu", "Tên hạng mục", "X1",
                            "Cột 5", "Cột 6", "Cột 7"])
    items, warnings = _read(path)
    assert all((i.unit or "").strip() for i in items)
    assert [w for w in warnings if "đơn vị tính" in w and "SUY RA" in w]


def test_correct_headers_are_left_alone(tmp_path: Path):
    items, warnings = _read(_book(tmp_path, BASE))
    assert len(items) == 30
    assert not [w for w in warnings if "SUY RA" in w], \
        "File đọc đúng theo tiêu đề thì không được suy luận gì"


def test_inference_never_overrides_a_mapped_column(tmp_path: Path):
    # Đơn giá nhận được từ tiêu đề; chỉ thành tiền phải suy ra.
    path = _book(tmp_path, ["STT", "Mã hiệu", "Tên hạng mục", "ĐVT",
                            "KL nhà thầu chào", "Đơn giá tổng hợp", "Cột 7"])
    items, warnings = _read(path)
    assert all(i.unit_price_total == 100_000 + (n + 1) * 7_531
               for n, i in enumerate(items))
    assert all(i.bid_amount is not None for i in items)
    suy = [w for w in warnings if "SUY RA" in w]
    assert suy and "bid_amount" in suy[0] and "unit_price_total" not in suy[0]
