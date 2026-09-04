"""Bảng làm rõ đặt tên cột kiểu khác vẫn phải nhận ra.

Hồ sơ thật dùng bảng "ĐÁNH GIÁ & LÀM RÕ" với cột yêu cầu tên là "CĐT YÊU CẦU
LÀM RÕ" chứ không phải "Ý kiến CĐT". Hai lỗi khiến cả file bị bỏ qua:

- không có mẫu nào khớp "CĐT yêu cầu làm rõ";
- mỗi khóa tự lấy cột đầu tiên khớp từ khóa của nó, nên "NỘI DUNG HỒ SƠ YÊU CẦU"
  bị cả khóa nội dung lẫn khóa yêu cầu giành, trong khi cột yêu cầu làm rõ thật
  thì không khóa nào nhận.
"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from core.models import UserFacingError
from core.rfi_tracker import (
    STATUS_ANSWERED,
    STATUS_UNANSWERED,
    _find_header,
    parse_rfi_file,
    track_rfi,
)

REAL_HEADER = ("TT", "NỘI DUNG HỒ SƠ YÊU CẦU", "YÊU CẦU",
               "NỘI DUNG HỒ SƠ CHÀO GIÁ", "ĐÁNH GIÁ", "CĐT YÊU CẦU LÀM RÕ")
ROWS = [
    ("I", "TÍNH HỢP LỆ", "", "", "", ""),
    ("1", "Hiệu lực hồ sơ chào giá", "90 ngày", "60 ngày", "Chưa đạt",
     "Đề nghị gia hạn hiệu lực lên 90 ngày"),
    ("2", "Hợp đồng tương tự", "02 hợp đồng", "03 hợp đồng", "Cần làm rõ",
     "Đề nghị bổ sung bản sao công chứng"),
]


def _book(path: Path, response_header: str | None = None,
          answers: tuple[str, ...] = ()) -> Path:
    wb = Workbook(); ws = wb.active; ws.title = "DANH GIA & LAM RO"
    ws.append(["DỰ ÁN MẪU"]); ws.append([])
    ws.append(list(REAL_HEADER) + ([response_header] if response_header else []))
    index = 0
    for row in ROWS:
        line = list(row)
        if response_header:
            if row[5]:
                line.append(answers[index] if index < len(answers) else "")
                index += 1
            else:
                line.append("")
        ws.append(line)
    wb.save(path)
    return path


def test_header_maps_every_column_exactly_once():
    _, mapping = _find_header([REAL_HEADER])
    assert mapping["content"] == 1, "Cột nội dung phải là 'NỘI DUNG HỒ SƠ YÊU CẦU'"
    assert mapping["requirement"] == 2
    assert mapping["cdt_request"] == 5, "'CĐT YÊU CẦU LÀM RÕ' phải là cột yêu cầu làm rõ"
    assert len(set(mapping.values())) == len(mapping), "Không cột nào bị hai khóa giành"


def test_requests_are_read_from_the_real_layout(tmp_path: Path):
    items = parse_rfi_file(_book(tmp_path / "cdt.xlsx"))
    assert len(items) == 2
    assert items[0].cdt_request.startswith("Đề nghị gia hạn")
    assert items[0].content == "Hiệu lực hồ sơ chào giá"


@pytest.mark.parametrize("response_header", [
    "NHÀ THẦU PHẢN HỒI",
    "Nhà thầu trả lời làm rõ",
    "NỘI DUNG BỔ SUNG",
    "GIẢI TRÌNH CỦA NHÀ THẦU",
    "Nhà thầu giải trình",
])
def test_response_column_variants(tmp_path: Path, response_header: str):
    cdt = _book(tmp_path / "cdt.xlsx")
    nt = _book(tmp_path / f"nt_{abs(hash(response_header))}.xlsx", response_header,
               ("Đồng ý gia hạn lên 90 ngày", ""))
    result = track_rfi(cdt, nt, "Nhà thầu A")
    assert len(result.items) == 2
    assert result.count(STATUS_ANSWERED) == 1
    assert result.count(STATUS_UNANSWERED) == 1


def test_old_layout_still_works(tmp_path: Path):
    wb = Workbook(); ws = wb.active; ws.title = "PL 1"
    ws.append(["STT", "Nội dung đánh giá theo HSYC", "Yêu cầu",
               "Nhà thầu kê khai", "Ý kiến CĐT", "Nhà thầu trả lời làm rõ"])
    ws.append(["1", "Hiệu lực hồ sơ", "90 ngày", "60 ngày",
               "Đề nghị gia hạn", "Nhà thầu đồng ý"])
    path = tmp_path / "cu.xlsx"; wb.save(path)
    result = track_rfi(path, path, "Nhà thầu A")
    assert len(result.items) == 1 and result.count(STATUS_ANSWERED) == 1


def test_error_lists_the_accepted_headers(tmp_path: Path):
    wb = Workbook(); ws = wb.active; ws.title = "RFI"
    ws.append(["STT", "CÂU HỎI NHÀ THẦU", "TVTK TRẢ LỜI", "CHỦ ĐẦU TƯ TRẢ LỜI"])
    ws.append(["1", "Đề nghị làm rõ phạm vi", "", "Theo thiết kế"])
    path = tmp_path / "hsyc.xlsx"; wb.save(path)
    with pytest.raises(UserFacingError) as excinfo:
        track_rfi(path, path, "Nhà thầu A")
    message = str(excinfo.value)
    assert "CĐT yêu cầu làm rõ" in message, "Phải nêu các tiêu đề được chấp nhận"
    assert "Tiêu đề đọc được" in message
